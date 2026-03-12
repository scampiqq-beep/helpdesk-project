from datetime import datetime, timedelta, UTC, date, time
from zoneinfo import ZoneInfo


# UTC helper (Python 3.14+: utcnow() deprecated)
def utcnow():
    """Naive UTC datetime for DB fields that historically stored UTC without tzinfo."""
    # Важно: это именно UTC без tzinfo (legacy-хранение в БД).
    return datetime.now(UTC).replace(tzinfo=None)


# Системное время интерфейса (Екатеринбург, UTC+5)
SYSTEM_TZ_NAME = 'Asia/Yekaterinburg'
SYSTEM_TZ = ZoneInfo(SYSTEM_TZ_NAME)


def system_now() -> datetime:
    """Timezone-aware datetime в системном часовом поясе."""
    return datetime.now(SYSTEM_TZ)


def get_runtime_timezone_name() -> str:
    try:
        return get_setting('system.timezone', SYSTEM_TZ_NAME) or SYSTEM_TZ_NAME
    except Exception:
        return SYSTEM_TZ_NAME


def get_runtime_timezone() -> ZoneInfo:
    try:
        return ZoneInfo(get_runtime_timezone_name())
    except Exception:
        return SYSTEM_TZ


def to_local(dt):
    if not dt:
        return None
    try:
        tz = get_runtime_timezone()
    except Exception:
        tz = SYSTEM_TZ
    if dt.tzinfo is None:
        # Основные даты в проекте хранятся как naive UTC (legacy SQLite без tzinfo).
        # Поэтому naive datetime сначала считаем UTC, а затем переводим в системный часовой пояс.
        dt = dt.replace(tzinfo=UTC)
    return dt.astimezone(tz)


def format_local(dt, fmt='%d.%m.%Y %H:%M'):
    if not dt:
        return ''
    return to_local(dt).strftime(fmt)
from flask import Flask, render_template, session, request, redirect, url_for, flash, jsonify, send_from_directory, g, Response, stream_with_context
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from flask_mail import Mail, Message
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from markupsafe import escape
from helpdesk_app.models import (
    db,
    SupportTicket,
    User,
    Department,
    FAQ,
    TicketAttachment,
    TicketHistory,
    TicketMessage,
    CommentLike,
    BitrixSettings,
    TicketTemplate,
    ResponseTemplate,
    Settings,
    TicketCategory,
    Tag,
    TicketStatus,
    TicketCloseReason,
    TicketPriority,
    ticket_tags,
    user_departments,
    Notification,
    NotificationGlobalSettings,
    TicketPresence,
    TicketOperatorChatMessage,
    TicketOperatorChatRead,
    KnowledgeBaseCategory,
    KnowledgeBaseArticle,
    KnowledgeBaseFavorite,
    UserUIState,
    WorkCalendarDay,
)
import os
from services.calendar_service import fetch_year as fetch_production_calendar_year
from utils.timezone_helper import to_local as helper_to_local, format_local as helper_format_local
from helpdesk_app.services.sla_service import SLAService
import json, yaml
import sqlite3
import re
import unicodedata
from openpyxl import Workbook

# =========================
# Input validation helpers
# =========================

# Очень базовый список стоп-слов. Держим коротким: главная цель — отсечь явные оскорбления.
# При желании можно расширять/вынести в настройки.
BAD_WORDS = {
    'петух', 'дурак', 'идиот', 'козёл', 'козел', 'сука', 'мразь', 'ублюдок', 'шлюха'
}

_URL_RE = re.compile(
    r"(?i)(?:https?://|www\.|\b[a-z0-9\-]+\.(?:ru|com|net|org|io|dev|app|info|biz|su)\b)"
)


def _norm_text(s: str) -> str:
    """Нормализация: NFC + схлопывание пробелов."""
    if s is None:
        return ''
    s = unicodedata.normalize('NFC', str(s))
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _contains_url_like(s: str) -> bool:
    if not s:
        return False
    return bool(_URL_RE.search(s))


def _contains_html_like(s: str) -> bool:
    if not s:
        return False
    if '<' in s or '>' in s:
        return True
    if re.search(r"(?i)\b(script|iframe|object|embed)\b", s):
        return True
    return False


def _contains_bad_words(s: str) -> bool:
    if not s:
        return False
    low = s.lower()
    return any(w in low for w in BAD_WORDS)


_PERSON_RE = re.compile(r"^[A-Za-zА-Яа-яЁёІіЇїЄєҐґ'\- ]{2,60}$")


def validate_person_part(value: str) -> tuple[bool, str]:
    """Проверка части ФИО (фамилия/имя/отчество)."""
    v = _norm_text(value)
    if not v:
        return False, "Поле обязательно"
    if _contains_url_like(v) or _contains_html_like(v):
        return False, "Недопустимые символы"
    if _contains_bad_words(v):
        return False, "Недопустимое значение"
    if not _PERSON_RE.fullmatch(v):
        return False, "Только буквы, пробелы и дефис (2–60 символов)"
    return True, v


_ORG_RE = re.compile(r"^[0-9A-Za-zА-Яа-яЁёІіЇїЄєҐґ№\"'\-\.,\(\)/ ]{2,200}$")
_ADDR_RE = re.compile(r"^[0-9A-Za-zА-Яа-яЁёІіЇїЄєҐґ№\"'\-\.,:;\(\)/ ]{3,300}$")


def validate_org(value: str) -> tuple[bool, str]:
    v = _norm_text(value)
    if not v:
        return True, ''
    if _contains_url_like(v) or _contains_html_like(v):
        return False, "Ссылки и HTML запрещены"
    if _contains_bad_words(v):
        return False, "Недопустимое значение"
    if not _ORG_RE.fullmatch(v):
        return False, "Слишком много спецсимволов"
    return True, v


def normalize_org_name(value: str) -> str:
    """Нормализация названия организации для сравнения.
    Используем, когда ИНН отсутствует у профиля или у заявки.
    """
    v = _norm_text(value or '').lower().replace('ё', 'е')
    if not v:
        return ''
    # убираем пунктуацию/кавычки, оставляем буквы/цифры/пробел
    v = re.sub(r"[^0-9a-zа-я ]+", " ", v, flags=re.IGNORECASE)
    v = re.sub(r"\s+", " ", v).strip()
    return v


def validate_address(value: str, required: bool = False) -> tuple[bool, str]:
    v = _norm_text(value)
    if not v:
        return (False, "Адрес обязателен") if required else (True, '')
    if _contains_url_like(v) or _contains_html_like(v):
        return False, "Ссылки и HTML запрещены"
    if not _ADDR_RE.fullmatch(v):
        return False, "Слишком много спецсимволов"
    return True, v


def normalize_phone(value: str) -> tuple[bool, str]:
    v = _norm_text(value)
    if not v:
        return True, ''
    v = re.sub(r"[^0-9\+]", "", v)
    if '+' in v and not v.startswith('+'):
        v = v.replace('+', '')
    digits = re.sub(r"\D", "", v)
    if len(digits) < 10 or len(digits) > 15:
        return False, "Телефон должен содержать 10–15 цифр"
    return True, v


def validate_inn_ru(value: str, required: bool = False) -> tuple[bool, str]:
    """Проверка ИНН.

    В Bitrix часто нет строгой проверки контрольных цифр, поэтому здесь делаем
    мягкую валидацию: только цифры и длина 10-12.

    Это убирает ложные ошибки при создании заявки, если ИНН уже сохранен в профиле.
    """
    v = _norm_text(value)
    if not v:
        return (False, "ИНН обязателен") if required else (True, '')
    if _contains_url_like(v) or _contains_html_like(v):
        return False, "ИНН должен содержать только цифры"
    if not re.fullmatch(r"\d{10,12}", v):
        return False, "ИНН должен содержать 10-12 цифр"
    return True, v

# =========================
# Ticket state indicators (Bitrix-like)
# =========================

_INDICATOR_LABELS = {
    'orange': 'последний раз в обращение писал клиент техподдержки (вы ответственный)',
    'yellow': 'последний раз в обращение писал клиент техподдержки (вы не ответственный)',
    'green':  'последний раз в обращение писали вы',
    'blue':   'последний раз в обращение писал сотрудник техподдержки',
    'black':  'обращение закрыто',
}


def ticket_is_closed(t) -> bool:
    """Bitrix-like: closed indicator when ticket is closed/resolved."""
    try:
        st = (getattr(t, 'status', '') or '').strip().lower()
        if st.startswith('закры') or st.startswith('заверш'):
            return True
    except Exception:
        pass
    if getattr(t, 'is_resolved', False):
        return True
    # if any closure timestamps are present
    for f in ('closed_at', 'auto_closed_at', 'marked_as_completed_at'):
        if getattr(t, f, None) is not None:
            return True
    return False


def ticket_responsible_id_value(t):
    """Responsible operator for indicator logic."""
    return getattr(t, 'assigned_to_id', None) or getattr(t, 'locked_by', None) or getattr(t, 'created_by_operator_id', None)


def compute_ticket_indicator(t, last_msg_is_operator, last_msg_user_id, current_user_id):
    """Return (code, title) where code in orange/yellow/green/blue/black.

    Logic is Bitrix-like:
    - black  -> ticket closed
    - orange -> last client activity and current operator is responsible
    - yellow -> last client activity and current operator is not responsible
    - green  -> last operator activity belongs to current operator
    - blue   -> last operator activity belongs to another operator

    For tickets without threaded messages the latest activity is the initial client request,
    so they are treated as "client last" instead of green.
    """
    if ticket_is_closed(t):
        code = 'black'
        return code, _INDICATOR_LABELS[code]

    rid = ticket_responsible_id_value(t)

    # No threaded replies yet -> initial request belongs to client.
    if last_msg_is_operator is None:
        code = 'orange' if (rid and current_user_id and rid == current_user_id) else 'yellow'
        return code, _INDICATOR_LABELS[code]

    # Last message from operator.
    if bool(last_msg_is_operator):
        code = 'green' if (current_user_id and last_msg_user_id and int(last_msg_user_id) == int(current_user_id)) else 'blue'
        return code, _INDICATOR_LABELS[code]

    # Last message from client.
    code = 'orange' if (rid and current_user_id and rid == current_user_id) else 'yellow'
    return code, _INDICATOR_LABELS[code]
from io import BytesIO
import requests
import urllib3
from dotenv import load_dotenv
import threading
from sqlalchemy.orm import joinedload, aliased
from sqlalchemy import desc
from sqlalchemy import func, or_
from sqlalchemy.exc import OperationalError, IntegrityError
import time as pytime
import os, re
import secrets
from itsdangerous import URLSafeTimedSerializer
from forms import MessageForm
from werkzeug.exceptions import RequestEntityTooLarge
import pandas as pd
# WebSocket (Socket.IO) для живых уведомлений
try:
    from flask_socketio import SocketIO, join_room
except Exception:
    SocketIO = None
    join_room = None


# Получаем абсолютный путь к папке проекта
basedir = os.path.abspath(os.path.dirname(__file__))

# === Настройки ===
load_dotenv()
UPLOAD_FOLDER = os.path.join('static', 'uploads')
ALLOWED_FILE_EXTENSIONS = {'image': {'png', 'jpg', 'jpeg'},'document': {'docx', 'xlsx', 'pdf'}}
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'zip', 'rar'}
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
# upload folder will be created after Flask app config is initialized

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect('instance/support.db')
        g.db.row_factory = sqlite3.Row  # позволяет обращаться по имени колонки
    return g.db

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# --- Приоритеты (храним строкой на русском) ---
PRIORITIES_RU = ["Низкий", "Средний", "Высокий"]
RU_TO_PRIORITY_CODE = {"Низкий": "low", "Средний": "medium", "Высокий": "high"}

PRIORITY_NORMALIZE = {
    "low": "Низкий",
    "medium": "Средний",
    "high": "Высокий",
    "низкий": "Низкий",
    "обычный": "Средний",  # совместимость со старым UI
    "средний": "Средний",
    "высокий": "Высокий",
    "Низкий": "Низкий",
    "Средний": "Средний",
    "Обычный": "Средний",
    "Высокий": "Высокий",
}

URGENT_RE = re.compile(r'\bсрочно\b', re.IGNORECASE)

def normalize_priority(value: str | None, default: str = "Низкий") -> str:
    if not value:
        return default
    v = str(value).strip()
    if v in PRIORITY_NORMALIZE:
        return PRIORITY_NORMALIZE[v]
    v_low = v.lower()
    return PRIORITY_NORMALIZE.get(v_low, default)



def priority_variants(value: str) -> list[str]:
    """Для совместимости: если в БД остались старые коды (high/medium/low),
    фильтруем сразу по двум вариантам."""
    ru = normalize_priority(value, default="Низкий")
    code = RU_TO_PRIORITY_CODE.get(ru)
    return [ru] + ([code] if code else [])

def apply_auto_priority(subject: str | None, message: str | None, current_priority: str | None = None) -> str:
    """Автоприоритет: если в теме/тексте есть слово 'срочно' — ставим 'Высокий'.
    Возвращает новое значение priority (на русском).
    """
    text = f"{subject or ''}\n{message or ''}"
    if URGENT_RE.search(text):
        return "Высокий"
    return normalize_priority(current_priority, default="Низкий")


def update_env_file(key, value):
    """Обновляет значение в .env-файле"""
    env_path = '.env'
    if not os.path.exists(env_path):
        with open(env_path, 'w', encoding='utf-8') as f:
            f.write(f"{key}={value}\n")
        return
    with open(env_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    updated = False
    for i, line in enumerate(lines):
        if line.startswith(f"{key}="):
            lines[i] = f"{key}={value}\n"
            updated = True
            break
    if not updated:
        lines.append(f"{key}={value}\n")
    with open(env_path, 'w', encoding='utf-8') as f:
        f.writelines(lines)

app = Flask(__name__)

# === Автоматическое закрытие заявок ===
from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime

def log_client_feedback(ticket_id, user_id, feedback_type, comment=""):
    """Логирует обратную связь от клиента (принятие/требует доработки)"""
    
    if feedback_type == 'accepted':
        field = 'client_feedback'
        old_value = 'ожидает'
        new_value = 'принял решение'
        note = 'Клиент подтвердил: проблема решена'
    elif feedback_type == 'rework':
        field = 'client_feedback'
        old_value = 'ожидает'
        new_value = 'запросил доработку'
        note = 'Клиент запросил: требуется обработка'
    else:
        return
    
    history_entry = TicketHistory(
        ticket_id=ticket_id,
        user_id=user_id,
        field=field,
        old_value=old_value,
        new_value=new_value,
        note=note
    )
    db.session.add(history_entry)

def start_scheduler():
    scheduler = BackgroundScheduler()
    # Запускать функцию auto_close_tickets каждый час
    scheduler.add_job(
        func=auto_close_tickets,
        trigger="interval",
        hours=1,
        id='auto_close_old_tickets'
    )
    scheduler.start()
    print("✅ Планировщик авто-закрытия заявок запущен")

def auto_close_tickets():
    """Автоматически завершает заявки, на которые клиент не отреагировал 24ч.

    Важно: APScheduler работает вне request-контекста, поэтому нужен app.app_context().
    """
    from datetime import datetime, timedelta

    with app.app_context():
        cutoff = utcnow() - timedelta(hours=24)

        tickets = SupportTicket.query.filter(
            SupportTicket.waiting_for_client_feedback == True,
            SupportTicket.marked_as_completed_at < cutoff,
            SupportTicket.status != 'Завершена'
        ).all()

        for ticket in tickets:
            old_status = ticket.status

            ticket.status = 'Завершена'
            try:
                ticket.close_reason = 'no_response'
            except Exception:
                pass
            ticket.is_spam = False
            ticket.auto_closed_at = utcnow()
            ticket.waiting_for_client_feedback = False
            ticket.closed_at = utcnow()

            # Логируем изменение статуса
            log_ticket_change(
                ticket.id,
                0,
                'status',
                old_status,
                'Завершена',
                note='Авто-закрытие: клиент не ответил за 24ч'
            )

            # Логируем "отказ" клиента дать обратную связь
            history = TicketHistory(
                ticket_id=ticket.id,
                user_id=0,  # системное событие
                field='client_feedback',
                old_value='ожидает',
                new_value='нет ответа',
                note='Клиент не предоставил обратную связь за 24 часа'
            )
            db.session.add(history)

            # Отправляем уведомление клиенту
            try:
                if ticket.end_user_rel and ticket.end_user_rel.email:
                    msg = Message(
                        subject=f"Заявка #{ticket.id} автоматически завершена",
                        recipients=[ticket.end_user_rel.email],
                        html=f"""
                        <p>Здравствуйте!</p>
                        <p>Ваша заявка #{ticket.id} была автоматически завершена, так как вы не предоставили обратную связь в течение 24 часов после решения проблемы.</p>
                        <p>Если проблема не решена, вы можете возобновить заявку по ссылке:
                        <a href="{url_for('ticket_detail', ticket_id=ticket.id, _external=True)}">Возобновить заявку #{ticket.id}</a></p>
                        <p>С уважением,<br>Техническая поддержка</p>
                        """
                    )
                    mail.send(msg)
            except Exception as e:
                print(f"Ошибка отправки email при авто-закрытии: {e}")

        if tickets:
            db.session.commit()
            print(f"[AUTO-CLOSE] Завершено {len(tickets)} заявок по истечению 24ч")

@app.template_filter('from_json')
def from_json_filter(value):
    if not value:
        return []
    try:
        return json.loads(value)
    except:
        return []

@app.template_filter('localtime')
def localtime_filter(value):
    return format_local(value)


@app.template_filter('localdate')
def localdate_filter(value):
    return format_local(value, '%d.%m.%Y')



@app.get('/api/system_time')
def api_system_time():
    """Текущее системное время (Екатеринбург, UTC+5).

    Используется в футере и на странице админки → Система.
    """
    now = datetime.now(get_runtime_timezone())
    return jsonify({
        'tz': get_runtime_timezone_name(),
        'iso': now.isoformat(),
        'date': now.strftime('%d.%m.%Y'),
        'time': now.strftime('%H:%M'),
        'display': now.strftime('%d.%m.%Y %H:%M')
    })

# Верификация по Email (NEW*3.01.2026)

def generate_confirmation_token(email):
    serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    return serializer.dumps(email, salt='email-confirmation-salt')

def confirm_token(token, expiration=3600):  # 1 час
    serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    try:
        email = serializer.loads(
            token,
            salt='email-confirmation-salt',
            max_age=expiration
        )
    except:
        return False
    return email

# Фукнция отправки Email (NEW*3.01.2026)


def send_email_verification(user):
    if not user.email:
        return

    token = generate_confirmation_token(user.email)
    confirm_url = url_for('confirm_email', token=token, _external=True)
    
    subject = "Подтвердите ваш email"
    body = f"Здравствуйте! Перейдите по ссылке для подтверждения:\n{confirm_url}"

    msg = Message(subject=subject, recipients=[user.email], body=body)
    try:
        mail.send(msg)
        print(f"✅ Письмо отправлено на {user.email}")
    except Exception as e:
        print(f"❌ Ошибка: {e}")

# Обработчики ошибок
@app.errorhandler(404)
def page_not_found(e):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('errors/500.html'), 500

@app.errorhandler(403)
def forbidden(e):
    return render_template('errors/403.html'), 403

@app.errorhandler(401)
def unauthorized(e):
    return render_template('errors/401.html'), 401


# Роут подтверждения отправки Email (NEW*3.01.2026)
@app.route('/confirm/<token>')
def confirm_email(token):
    email = confirm_token(token)
    if not email:
        flash('Ссылка недействительна или устарела.', 'error')
        return redirect(url_for('login'))

    print(f"🔍 Ищем email: '{email}'")
    
    # Единая таблица users
    user = User.query.filter(db.func.lower(User.email) == email.lower()).first()

    if not user:
        flash('Пользователь не найден.', 'error')
        return redirect(url_for('login'))

    if user.email_verified:
        flash('Ваш email уже подтверждён.', 'info')
    else:
        user.email_verified = True
        db.session.commit()
        flash('Email успешно подтверждён!', 'success')

    return redirect(url_for('login'))
# Отключаем экранирование Unicode в JSON
if hasattr(app, 'json'):
    app.json.ensure_ascii = False
else:
    app.config['JSON_AS_ASCII'] = False  # для старых версий Flask

app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Инициализация Socket.IO (живые уведомления)
socketio = None
if SocketIO is not None:
    try:
        socketio = SocketIO(app, async_mode='threading', cors_allowed_origins='*')
    except Exception as _e:
        print('SocketIO init error:', _e)
        socketio = None

app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024
app.config['MAX_FILES_PER_UPLOAD'] = 10
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(basedir, "instance", "support.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_size': 10,
    'max_overflow': 20,
    'pool_pre_ping': True,
    'pool_recycle': 3600,
}

# Email
app.config['MAIL_SERVER'] = os.getenv('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.getenv('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.getenv('MAIL_USE_TLS', 'True').lower() == 'true'
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = app.config['MAIL_USERNAME']


# Прикрепление файла в комментариях
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'zip', 'rar'}

# SLA
SLA_ATTENTION_HOURS = int(os.getenv('SLA_ATTENTION_HOURS', 1))
SLA_DEFAULT_HOURS = int(os.getenv('SLA_DEFAULT_HOURS', 48))

# Bitrix
BITRIX_WEBHOOK_URL = os.getenv('BITRIX_WEBHOOK_URL')
BITRIX_UPDATE_WEBHOOK_URL = os.getenv('BITRIX_UPDATE_WEBHOOK_URL', BITRIX_WEBHOOK_URL)
BITRIX_COMMENT_WEBHOOK_URL = os.getenv('BITRIX_COMMENT_WEBHOOK_URL')
BITRIX_CONTAKT_WEBHOOK_URL = os.getenv('BITRIX_CONTAKT_WEBHOOK_URL')

db.init_app(app)
# Flask-Migrate (для flask db ...)
migrate = Migrate(app, db)
mail = Mail(app)
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    """Загрузка пользователя из единой таблицы users (без legacy op_*/user_*)."""
    try:
        if user_id is None:
            return None
        uid = int(str(user_id))
        return db.session.get(User, uid)
    except Exception:
        return None

        # Новый формат: просто числовой id
        if str(user_id).isdigit():
            return db.session.get(User, int(user_id))
    except Exception:
        return None
    return None

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# --- Справочники по умолчанию (используются только для первичной инициализации) ---
DEFAULT_DEPARTMENTS = [
    "Требуется обработка",
    "1ая линия ТП",
    "2ая линия ТП",
    "Отдел аппаратного обеспечения",
    "Отдел сетевого обеспечения",
    "СЭД ДЕЛО",
    "Отдел мониторинга",
    "Отдел информационной безопасности",
]

# Категории (тип) заявок по умолчанию
DEFAULT_TICKET_CATEGORIES = [
    ("issue", "Обращение", 10),
    ("task", "Задача", 20),
]

# Роли в системе фиксированные: admin / operator / client
USER_ROLES = {
    'admin': 'Администратор',
    'operator': 'Оператор',
    'client': 'Клиент'
}

# Статусы (Вариант A): показываем только основные состояния.
# Подстатусы закрытия живут в справочнике TicketCloseReason.
STATUSES = ["Новая", "В работе", "Ожидание", "Завершена"]


def get_active_statuses():
    """Возвращает список статусов для фильтров/списков.

    В проекте исторически был справочник ticket_statuses, но для новой модели
    (фиксированные состояния) используем константу STATUSES.
    """
    return STATUSES


def get_active_close_reasons():
    """Активные причины закрытия (подстатусы для 'Завершена')."""
    try:
        rows = TicketCloseReason.query.filter_by(is_active=True).order_by(TicketCloseReason.sort_order, TicketCloseReason.name).all()
        return rows
    except Exception:
        return []

def get_active_priorities():
    """Возвращает активные приоритеты как список dict: {code,name}."""
    defaults = [
        {'code': 'low', 'name': 'Низкий'},
        {'code': 'medium', 'name': 'Средний'},
        {'code': 'high', 'name': 'Высокий'},
        {'code': 'urgent', 'name': 'Критический'},
    ]
    try:
        rows = TicketPriority.query.filter_by(is_active=True).order_by(TicketPriority.sort_order).all()
        data = [{'code': r.code, 'name': r.name} for r in rows]
        return data or defaults
    except Exception:
        return defaults

def get_active_tags():
    try:
        return Tag.query.filter_by(is_active=True).order_by(Tag.name).all()
    except Exception:
        return []


def status_badge_class(status: str | None) -> str:
    """CSS-класс для статуса (единый по всему проекту)."""
    s = (status or '').strip().lower()
    if not s:
        return 'status-default'
    if s in ('new',) or 'нов' in s:
        return 'status-new'
    if 'принят' in s or s in ('accepted',):
        return 'status-accepted'
    if 'в работе' in s or 'работ' in s or s in ('in_work', 'inwork'):
        return 'status-inwork'
    if 'ожидает' in s or 'подтверж' in s or 'confirm' in s:
        return 'status-wait_client'
    if 'спам' in s:
        return 'status-spam'
    if 'дублик' in s:
        return 'status-duplicate'
    if 'ошиб' in s:
        return 'status-mistake'
    if 'отозв' in s or 'withdraw' in s:
        return 'status-withdrawn'
    if 'заверш' in s or 'закры' in s or 'решен' in s or 'выполн' in s or s in ('resolved', 'done', 'closed'):
        return 'status-completed'
    if 'отклон' in s or 'cancel' in s or 'rejected' in s:
        return 'status-default'
    return 'status-default'


def status_icon(status: str | None) -> str:
    """Bootstrap icon class for ticket status (optional helper for templates)."""
    s = (status or '').strip().lower()
    if not s:
        return 'bi bi-dot'
    if s in ('new',) or 'нов' in s:
        return 'bi bi-plus-circle'
    if 'принят' in s or s in ('accepted',):
        return 'bi bi-play-circle'
    if 'в работе' in s or 'работ' in s or s in ('in_work', 'inwork'):
        return 'bi bi-gear'
    if 'ожидает' in s or 'подтверж' in s or 'confirm' in s:
        return 'bi bi-pause-circle'
    if 'спам' in s or s in ('spam',):
        return 'bi bi-ban'
    if 'дублик' in s or s in ('duplicate',):
        return 'bi bi-files'
    if 'ошиб' in s or s in ('mistake',):
        return 'bi bi-x-octagon'
    if 'решен' in s or 'выполн' in s or s in ('resolved', 'done', 'closed'):
        return 'bi bi-check-circle'
    if 'отклон' in s or 'cancel' in s or 'rejected' in s:
        return 'bi bi-x-circle'
    return 'bi bi-dot'

def priority_badge_class(priority: str | None) -> str:
    """CSS-класс для приоритета. Поддерживаем и код (low/medium/high/urgent), и RU-значения."""
    p = (priority or '').strip().lower()
    if p in ('urgent', 'critical') or 'крит' in p:
        return 'priority-urgent'
    if p in ('high',) or 'высок' in p:
        return 'priority-high'
    if p in ('low',) or 'низк' in p:
        return 'priority-low'
    # medium/normal
    if p in ('medium', 'normal') or 'обыч' in p or 'средн' in p:
        return 'priority-medium'
    return 'priority-medium'






def priority_label(priority: str | None) -> str:
    """Человекочитаемый приоритет (Низкий/Средний/Высокий)."""
    p = (priority or '').strip().lower()
    if p in ('high',) or 'высок' in p:
        return 'Высокий'
    if p in ('low',) or 'низк' in p:
        return 'Низкий'
    if p in ('medium', 'normal') or 'обыч' in p or 'средн' in p:
        return 'Средний'
    return (priority or '—')


def bbcode_to_html(text_in: str) -> str:
    """Мини-bbcode (b/i/s/url/list/quote/code) + экранирование HTML.

    Важно: используем безопасное преобразование в HTML (escape), затем подставляем
    ограниченный набор тегов.
    """
    import html

    # нормализуем переносы строк
    t = (text_in or '').replace('\r\n', '\n').replace('\r', '\n')

    # Если в базе уже лежит HTML (старые комментарии/закреплённый результат),
    # не экранируем его повторно — аккуратно чистим и показываем как есть.
    raw = (text_in or '')
    if ('<' in raw and '>' in raw) and (not re.search(r"\[[a-z]", raw, re.I)):
        s = raw.replace('\r\n', '\n').replace('\r', '\n')
        s = re.sub(r"(?is)<(script|style)[^>]*>.*?</\1>", '', s)
        # div/p -> перенос строки
        s = re.sub(r"(?i)</?(div|p)[^>]*>", '<br>', s)
        allowed = {'br','strong','em','s','a','ul','li','blockquote','pre','code'}

        def _clean_tag(m):
            slash = m.group(1)
            tag = (m.group(2) or '').lower()
            attrs = m.group(3) or ''
            if tag not in allowed:
                return ''
            if tag == 'br':
                return '<br>'
            if slash:
                return f'</{tag}>'
            if tag == 'a':
                href_m = re.search(r'href\s*=\s*"([^"]+)"', attrs, re.I)
                href = href_m.group(1) if href_m else '#'
                return f'<a href="{href}" target="_blank" rel="noopener">'
            if tag == 'ul':
                return '<ul class="bb-list">'
            if tag == 'blockquote':
                return '<blockquote class="bb-quote">'
            if tag == 'pre':
                return '<pre class="bb-code">'
            return f'<{tag}>'

        s = re.sub(r"<\s*(/?)\s*([a-zA-Z0-9]+)([^>]*)>", _clean_tag, s)
        return s

    # экранируем HTML
    t = html.escape(t)

    # [code]...[/code]
    def repl_code(m):
        code = m.group(1)
        return f'<pre class="bb-code"><code>{code}</code></pre>'

    t = re.sub(r"\[code\](.*?)\[/code\]", repl_code, t, flags=re.S | re.I)

    # [quote]...[/quote]
    t = re.sub(r"\[quote\](.*?)\[/quote\]", r'<blockquote class="bb-quote">\1</blockquote>', t, flags=re.S | re.I)

    # базовые стили
    t = re.sub(r"\[b\](.*?)\[/b\]", r'<strong>\1</strong>', t, flags=re.S | re.I)
    t = re.sub(r"\[i\](.*?)\[/i\]", r'<em>\1</em>', t, flags=re.S | re.I)
    t = re.sub(r"\[s\](.*?)\[/s\]", r'<s>\1</s>', t, flags=re.S | re.I)

    # ссылки
    t = re.sub(
        r"\[url=(.*?)\](.*?)\[/url\]",
        r'<a href="\1" target="_blank" rel="noopener">\2</a>',
        t,
        flags=re.S | re.I,
    )
    t = re.sub(
        r"\[url\](.*?)\[/url\]",
        r'<a href="\1" target="_blank" rel="noopener">\1</a>',
        t,
        flags=re.S | re.I,
    )

    # списки: [list][*]a[*]b[/list]
    def repl_list(m):
        body = m.group(1)
        items = re.split(r"\[\*\]", body)
        items = [i.strip() for i in items if i.strip()]
        if not items:
            return ''
        li = ''.join(f'<li>{i}</li>' for i in items)
        return f'<ul class="bb-list">{li}</ul>'

    t = re.sub(r"\[list\](.*?)\[/list\]", repl_list, t, flags=re.S | re.I)

    # авто-ссылки
    t = re.sub(
        r"(https?://[\w\-\._~:/\?#\[\]@!\$&'\(\)\*\+,;=%]+)",
        r'<a href="\1" target="_blank" rel="noopener">\1</a>',
        t,
    )

    # переносы строк -> <br>
    t = t.replace('\n', '<br>')
    return t


def html_to_bbcode(html_in: str) -> str:
    """Грубое обратное преобразование HTML (сгенерированного bbcode_to_html) обратно в BBCode.

    Нужно для режима редактирования комментариев: чтобы в textarea попадал именно BBCode,
    а не плоский текст без тегов.
    """
    import html as _html

    s = (html_in or '')

    # Нормализуем переносы строк
    s = s.replace('\r\n', '\n').replace('\r', '\n')

    # Сначала разворачиваем списки
    def repl_list(m):
        body = m.group(1)
        items = re.findall(r"<li>(.*?)</li>", body, flags=re.S | re.I)
        items = [i.strip() for i in items if i.strip()]
        if not items:
            return ''
        out = "[list]\n" + "\n".join(f"[*]{i}" for i in items) + "\n[/list]"
        return out

    s = re.sub(r"<ul class=\"bb-list\">(.*?)</ul>", repl_list, s, flags=re.S | re.I)

    # code/quote
    s = re.sub(r"<pre class=\"bb-code\"><code>(.*?)</code></pre>", r"[code]\1[/code]", s, flags=re.S | re.I)
    s = re.sub(r"<blockquote class=\"bb-quote\">(.*?)</blockquote>", r"[quote]\1[/quote]", s, flags=re.S | re.I)

    # базовые теги
    s = re.sub(r"<strong>(.*?)</strong>", r"[b]\1[/b]", s, flags=re.S | re.I)
    s = re.sub(r"<em>(.*?)</em>", r"[i]\1[/i]", s, flags=re.S | re.I)
    s = re.sub(r"<s>(.*?)</s>", r"[s]\1[/s]", s, flags=re.S | re.I)

    # ссылки [url]
    s = re.sub(r"<a[^>]*href=\"(.*?)\"[^>]*>(.*?)</a>", r"[url=\1]\2[/url]", s, flags=re.S | re.I)

    # <br> -> \n
    s = re.sub(r"<br\s*/?>", "\n", s, flags=re.I)

    # Убираем случайные теги (если остались)
    s = re.sub(r"<[^>]+>", "", s)

    # HTML entities -> normal
    s = _html.unescape(s)

    # Чуть подчистим
    return s.strip()


@app.template_filter('priority_label')
def priority_label_filter(value):
    return priority_label(value)


@app.template_filter('bbcode')
def bbcode_filter(value):
    return bbcode_to_html(value or '')

def ticket_id_pad(value) -> str:
    """Форматирование ID заявки.

    Раньше показывали 6 знаков с нулями (000001). По требованиям —
    показываем «нормальные» номера (1,2,3...).
    """
    try:
        return str(int(value))
    except Exception:
        return str(value)

@app.template_filter('ticket_id')
def ticket_id_filter(value):
    return ticket_id_pad(value)



def _notif_room_for(obj):
    try:
        rtype, rid = _recipient_key(obj)
        return f"notif_{rtype}_{rid}"
    except Exception:
        return None

if socketio is not None:
    @socketio.on('connect')
    def _socket_connect():
        try:
            if not getattr(current_user, 'is_authenticated', False):
                return False
            room = _notif_room_for(current_user)
            if room and join_room is not None:
                join_room(room)
        except Exception:
            return False
        return True

    @socketio.on('join_ticket')
    def _join_ticket(data):
        """Join per-ticket room for realtime UI (typing, etc.)."""
        try:
            if not getattr(current_user, 'is_authenticated', False):
                return
            tid = None
            if isinstance(data, dict):
                tid = data.get('ticket_id')
            if tid is None:
                return
            try:
                tid = int(tid)
            except Exception:
                return
            room = f"ticket_{tid}"
            if join_room is not None:
                join_room(room)
        except Exception:
            return

    @socketio.on('ticket_typing')
    def _ticket_typing(data):
        """Operator/admin typing indicator in ticket comments."""
        try:
            if not getattr(current_user, 'is_authenticated', False):
                return
            # Только оператор/админ могут отправлять индикатор
            if not (getattr(current_user, 'is_operator', False) or getattr(current_user, 'is_admin', False)):
                return
            if not isinstance(data, dict):
                return
            tid = data.get('ticket_id')
            is_typing = bool(data.get('is_typing'))
            try:
                tid = int(tid)
            except Exception:
                return

            room = f"ticket_{tid}"
            uname = getattr(current_user, 'full_name', None) or getattr(current_user, 'name', None) or getattr(current_user, 'username', 'Оператор')
            payload = {
                'ticket_id': tid,
                'user_id': getattr(current_user, 'id', None),
                'user_name': uname,
                'is_typing': is_typing,
            }
            socketio.emit('ticket_typing', payload, room=room)
        except Exception:
            return

@app.context_processor
def inject_badge_helpers():
    return {
        'status_badge_class': status_badge_class,
        'status_icon': status_icon,
        'priority_badge_class': priority_badge_class,
        # доступ/роль helpers для шаблонов
        'is_tp_operator': is_tp_operator,
    }

def get_setting(key: str, default: str | None = None) -> str | None:
    s = Settings.query.filter_by(key=key).first()
    return s.value if s else default

def set_setting(key: str, value: str) -> None:
    s = Settings.query.filter_by(key=key).first()
    if not s:
        s = Settings(key=key, value=value)
        db.session.add(s)
    else:
        s.value = value


def flash_msg(message: str, category: str = 'info') -> None:
    """Единая точка для flash-сообщений.

    Нормализуем категории, чтобы в шаблонах не плодить условия.
    """
    cat = (category or 'info').lower()
    if cat in ('err', 'error', 'danger'):
        cat = 'danger'
    elif cat in ('warn', 'warning'):
        cat = 'warning'
    elif cat in ('ok', 'success'):
        cat = 'success'
    else:
        cat = 'info'
    flash(message, cat)


# === Уведомления (In-app) ===

def _recipient_key(user_obj):
    if getattr(user_obj, 'role', None) == 'client':
        return ('end_user', user_obj.id)
    return ('user', user_obj.id)

def get_notification_settings():
    try:
        return NotificationGlobalSettings.get_or_create()
    except Exception:
        # если БД ещё не готова
        return None

def _is_notifications_enabled_for(recipient_type: str, recipient_obj) -> bool:
    s = get_notification_settings()
    if not s or not s.enabled:
        return False
    if recipient_type == 'user':
        return bool(s.enabled_for_operators) and bool(getattr(recipient_obj, 'notify_inapp_enabled', True))
    return bool(s.enabled_for_clients) and bool(getattr(recipient_obj, 'notify_inapp_enabled', True))

def _is_event_enabled(recipient_obj, event_type: str, title: str = '', body: str = '') -> bool:
    s = get_notification_settings()
    if not s or not s.enabled:
        return False

    event_type = (event_type or '').strip().lower()
    title_lower = (title or '').strip().lower()
    body_lower = (body or '').strip().lower()

    if event_type == 'assigned':
        # Исторически через assigned шли сразу несколько сценариев:
        # создание заявки, попадание в отдел, назначение исполнителя.
        if 'новая заявка' in title_lower or 'создана' in title_lower or 'в отделе' in title_lower:
            return bool(getattr(s, 'event_new_ticket', True)) and bool(getattr(recipient_obj, 'notify_event_assigned', True))
        return bool(getattr(s, 'event_assigned', True)) and bool(getattr(recipient_obj, 'notify_event_assigned', True))

    if event_type == 'customer_reply':
        # Для клиента это обычно ответ оператора, для оператора — ответ клиента.
        if getattr(recipient_obj, 'role', None) == 'client':
            return bool(getattr(s, 'event_operator_reply', True)) and bool(getattr(recipient_obj, 'notify_event_customer_reply', True))
        return bool(getattr(s, 'event_customer_reply', True)) and bool(getattr(recipient_obj, 'notify_event_customer_reply', True))

    if event_type == 'status':
        if ('sla' in title_lower) or ('sla' in body_lower) or ('просроч' in title_lower) or ('просроч' in body_lower):
            return bool(getattr(s, 'event_sla', True)) and bool(getattr(recipient_obj, 'notify_event_status', True))
        return bool(getattr(s, 'event_status', True)) and bool(getattr(recipient_obj, 'notify_event_status', True))

    if event_type == 'priority':
        if 'важн' in title_lower or 'важн' in body_lower:
            return bool(getattr(s, 'event_important', True))
        return bool(getattr(s, 'event_priority', True))

    if event_type == 'opchat':
        return bool(getattr(s, 'event_opchat', True))

    return True


def _operators_with_access_to_ticket(ticket: 'SupportTicket') -> list['User']:
    """Список операторов/админов, у которых есть доступ к тикету.
    Используется для рассылки уведомлений (комментарии, приоритет, оп.чат).
    """
    try:
        if not ticket:
            return []

        recipients: list[User] = []

        # 1) Админы
        try:
            recipients.extend(User.query.filter_by(role='admin').all())
        except Exception:
            pass

        # 2) Назначенный оператор
        try:
            if getattr(ticket, 'assigned_to_id', None):
                u = db.session.get(User, ticket.assigned_to_id)
                if u:
                    recipients.append(u)
        except Exception:
            pass

        # 3) Основной отдел
        try:
            if getattr(ticket, 'department_id', None):
                dept = Department.query.get(ticket.department_id)
                if dept:
                    recipients.extend(list(getattr(dept, 'users', []) or []))
                    recipients.extend(list(getattr(dept, 'operators', []) or []))
        except Exception:
            pass

        # 4) Доп. отделы (shared)
        try:
            if getattr(ticket, 'shared_departments_rel', None):
                for d in (ticket.shared_departments_rel or []):
                    if not d:
                        continue
                    recipients.extend(list(getattr(d, 'users', []) or []))
                    recipients.extend(list(getattr(d, 'operators', []) or []))
        except Exception:
            pass

        # 5) ТП (1/2 линия) — глобальный доступ
        try:
            tp_ops = User.query.filter(User.role.in_(['operator', 'admin'])).all()
            for u in tp_ops:
                try:
                    if is_tp_operator(u):
                        recipients.append(u)
                except Exception:
                    continue
        except Exception:
            pass

        # unique + оставляем только операторов/админов, у кого реально есть доступ
        out: list[User] = []
        seen = set()
        for u in recipients:
            if not u or getattr(u, 'id', None) is None:
                continue
            if u.id in seen:
                continue
            seen.add(u.id)
            if getattr(u, 'role', None) not in ('operator', 'admin'):
                continue
            if _user_has_access_to_ticket_for_notifications(u, ticket):
                out.append(u)
        return out
    except Exception:
        return []

def create_inapp_notification(
    recipient_obj,
    event_type: str,
    title: str,
    body: str = '',
    url: str | None = None,
    dedupe_key: str | None = None,
    ticket_id: int | None = None,
    commit: bool = True,
    dedupe_minutes: int = 10,
):
    try:
        recipient_type, recipient_id = _recipient_key(recipient_obj)
        if not _is_notifications_enabled_for(recipient_type, recipient_obj):
            return
        if not _is_event_enabled(recipient_obj, event_type, title=title, body=body):
            return

        # ticket_id: предпочтительно передавать явно, но поддерживаем обратную совместимость через url
        if ticket_id is None and url:
            try:
                m = _TICKET_URL_RE.search(url) or _TICKET_URL_RE_ALT.search(url)
                if m:
                    ticket_id = int(m.group(1))
            except Exception:
                ticket_id = None

        # дедуп (опционально): не создаём дубликат в течение dedupe_minutes
        if dedupe_key:
            try:
                cutoff = utcnow() - timedelta(minutes=max(int(dedupe_minutes or 0), 0))
                q = (Notification.query
                     .filter_by(
                         recipient_type=recipient_type,
                         recipient_id=recipient_id,
                         dedupe_key=dedupe_key,
                         is_read=False
                     ))
                if dedupe_minutes:
                    q = q.filter(Notification.created_at >= cutoff)
                recent = q.order_by(Notification.created_at.desc()).first()
                if recent:
                    return
            except Exception:
                pass

        n = Notification(
            recipient_type=recipient_type,
            recipient_id=recipient_id,
            event_type=event_type,
            title=(title or '')[:200],
            body=body or '',
            url=url,
            ticket_id=ticket_id,
            dedupe_key=dedupe_key
        )
        db.session.add(n)
        if commit:
            db.session.commit()
        else:
            db.session.flush()
        # WebSocket push
        try:
            if socketio is not None:
                payload = {
                    'id': n.id,
                    'title': n.title,
                    'body': n.body or '',
                    'url': n.url or '',
                    'created_at': n.created_at.strftime('%H:%M') if n.created_at else ''
                }
                socketio.emit('notification', payload, room=f"notif_{recipient_type}_{recipient_id}")
        except Exception:
            pass
    except Exception:
        db.session.rollback()
        return


# --- Фильтрация уведомлений по доступу к заявкам ---
_TICKET_URL_RE = re.compile(r"/tickets/(\d+)(?:\b|/|\?|#)")
_TICKET_URL_RE_ALT = re.compile(r"[?&]ticket_id=(\d+)(?:\b|&|#)")

def _notification_ticket_id(n: 'Notification') -> int | None:
    """Извлекаем ticket_id из Notification.url (если это уведомление по заявке)."""
    try:
        # 1) новый путь: ticket_id хранится в колонке
        tid = getattr(n, 'ticket_id', None)
        if tid:
            return int(tid)

        # 2) обратная совместимость: парсим из url
        url = (getattr(n, 'url', None) or '').strip()
        if not url:
            return None
        m = _TICKET_URL_RE.search(url) or _TICKET_URL_RE_ALT.search(url)
        if m:
            return int(m.group(1))
    except Exception:
        return None
    return None

def _user_has_access_to_ticket_for_notifications(user_obj, ticket: 'SupportTicket') -> bool:
    """Проверка доступа к заявке для фильтра уведомлений (как на странице заявки)."""
    try:
        if not user_obj or not ticket:
            return False

        # Клиент
        if getattr(user_obj, 'role', None) == 'client':
            return bool(ticket.email) and bool(user_obj.email) and ticket.email == user_obj.email

        # Оператор / админ
        if isinstance(user_obj, User):
            if getattr(user_obj, 'role', None) == 'admin':
                return True

            # 1/2 линия ТП — глобальный доступ
            try:
                if is_tp_operator(user_obj):
                    return True
            except Exception:
                pass

            if getattr(ticket, 'assigned_to_id', None) == user_obj.id:
                return True

            dep_ids = set(user_department_ids(user_obj))
            if getattr(ticket, 'department_id', None) and int(ticket.department_id) in dep_ids:
                return True

            # shared departments (если есть)
            try:
                if getattr(ticket, 'shared_departments_rel', None):
                    for d in ticket.shared_departments_rel:
                        if d and getattr(d, 'id', None) in dep_ids:
                            return True
            except Exception:
                pass

        return False
    except Exception:
        return False


def _accessible_ticket_ids_for_notifications(user_obj, max_ids: int = 5000) -> set[int] | None:
    """Быстрая выборка доступных ticket_id для фильтрации уведомлений.

    Возвращает:
      - None: если у пользователя глобальный доступ (admin / 1-2 линия ТП)
      - set(ids): если доступ ограничен
    """
    try:
        if not user_obj:
            return set()

        # Клиент
        if getattr(user_obj, 'role', None) == 'client':
            if not getattr(user_obj, 'email', None):
                return set()
            ids = (db.session.query(SupportTicket.id)
                   .filter(SupportTicket.email == user_obj.email)
                   .order_by(SupportTicket.id.desc())
                   .limit(int(max_ids))
                   .all())
            return {int(r[0]) for r in ids}

        # Оператор / админ
        if isinstance(user_obj, User):
            if getattr(user_obj, 'role', None) == 'admin':
                return None
            try:
                if is_tp_operator(user_obj):
                    return None
            except Exception:
                pass

            dep_ids = [int(x) for x in (user_department_ids(user_obj) or []) if str(x).isdigit()]
            ids: set[int] = set()

            # assigned
            rows = (db.session.query(SupportTicket.id)
                    .filter(SupportTicket.assigned_to_id == user_obj.id)
                    .order_by(SupportTicket.id.desc())
                    .limit(int(max_ids))
                    .all())
            ids.update(int(r[0]) for r in rows)

            # by department
            if dep_ids:
                rows = (db.session.query(SupportTicket.id)
                        .filter(SupportTicket.department_id.in_(dep_ids))
                        .order_by(SupportTicket.id.desc())
                        .limit(int(max_ids))
                        .all())
                ids.update(int(r[0]) for r in rows)

                # shared departments (ticket_shared_departments)
                try:
                    # SQLite-safe IN (...)
                    placeholders = ",".join([f":d{i}" for i in range(len(dep_ids))])
                    params = {f"d{i}": int(dep_ids[i]) for i in range(len(dep_ids))}
                    params["lim"] = int(max_ids)
                    q = db.text(
                        f"SELECT DISTINCT ticket_id FROM ticket_shared_departments WHERE department_id IN ({placeholders}) LIMIT :lim"
                    )
                    rows = db.session.execute(q, params).fetchall()
                    ids.update(int(r[0]) for r in (rows or []))
                except Exception:
                    pass

            # hard cap
            if len(ids) > int(max_ids):
                return set(list(ids)[:int(max_ids)])
            return ids

        return set()
    except Exception:
        return set()

def _filter_notifications_for_user(user_obj, notifications: list['Notification']) -> list['Notification']:
    """Оставляем только уведомления, относящиеся к доступным заявкам.

    Оптимизация: сначала собираем ticket_id и вытаскиваем тикеты одним запросом,
    чтобы избежать N+1.

    Уведомления без ticket_id (глобальные) — оставляем.
    """
    notifications = notifications or []
    # ticket ids
    tids: set[int] = set()
    for n in notifications:
        try:
            tid = _notification_ticket_id(n)
            if tid:
                tids.add(int(tid))
        except Exception:
            continue

    ticket_map: dict[int, SupportTicket] = {}
    if tids:
        try:
            tickets = SupportTicket.query.filter(SupportTicket.id.in_(list(tids))).all()
            ticket_map = {int(t.id): t for t in tickets}
        except Exception:
            ticket_map = {}

    out: list[Notification] = []
    for n in notifications:
        tid = _notification_ticket_id(n)
        if not tid:
            out.append(n)
            continue
        t = ticket_map.get(int(tid))
        if t and _user_has_access_to_ticket_for_notifications(user_obj, t):
            out.append(n)
    return out

@app.context_processor
def inject_nav_notifications():
    """Данные для колокольчика в навбаре.
    ВАЖНО: показываем только уведомления по тем заявкам, к которым у пользователя есть доступ.
    """
    try:
        if not getattr(current_user, 'is_authenticated', False):
            return {}

        recipient_type, recipient_id = _recipient_key(current_user)
        # Берём "окно" последних уведомлений для меню (дальше фильтруем по доступу)
        raw = (Notification.query
               .filter_by(recipient_type=recipient_type, recipient_id=recipient_id)
               .order_by(Notification.created_at.desc())
               .limit(200)
               .all())
        visible = _filter_notifications_for_user(current_user, raw)
        items = visible[:5]

        # unread считаем отдельно (берём с запасом только непрочитанные)
        raw_unread = (Notification.query
                      .filter_by(recipient_type=recipient_type, recipient_id=recipient_id, is_read=False)
                      .order_by(Notification.created_at.desc())
                      .limit(800)
                      .all())
        unread = len(_filter_notifications_for_user(current_user, raw_unread))

        return dict(nav_unread_count=unread, nav_notifications=items)
    except Exception:
        return dict(nav_unread_count=0, nav_notifications=[])




def get_profile_enforcement_mode() -> str:
    """Режим обязательности профиля для клиентов.

    strict: пускаем только в профиль/выход/подтверждение почты/bitrix lookup.
    soft: можно смотреть, но нельзя создавать/отправлять (POST) до заполнения.
    off: выключено.
    """
    mode = (get_setting('profile_enforcement_mode', 'strict') or 'strict').strip().lower()
    return mode if mode in ('strict', 'soft', 'off') else 'strict'

def get_default_intake_department_id() -> int | None:
    v = get_setting('default_intake_department_id')
    if v and str(v).isdigit():
        return int(v)
    return None

def user_department_ids(user: User) -> list[int]:
    # Поддерживаем both: многие отделы (user.departments) и старый "основной отдел" (department_id)
    ids = set()
    if hasattr(user, 'departments') and user.departments:
        ids.update([d.id for d in user.departments])
    if getattr(user, 'department_id', None):
        ids.add(int(user.department_id))
    return list(ids)

def is_tp_operator(user: User) -> bool:
    """Операторы техподдержки (1/2 линия).

    Раньше было строгое сравнение с '1ая линия ТП'/'2ая линия ТП', из‑за чего права
    ломались при любом отличии в названии отдела. Теперь проверка гибкая:
    допускаем варианты '1 линия ТП', '2-я линия техподдержки' и т.п.
    """
    try:
        names = set()
        for d in (getattr(user, 'departments', []) or []):
            if getattr(d, 'name', None):
                names.add(d.name)
        if getattr(user, 'department', None) and getattr(user.department, 'name', None):
            names.add(user.department.name)

        # re используется в _match
        def _match(line_num: str, name: str) -> bool:
            s = (name or '').lower().replace('ё', 'е')
            if 'линия' not in s and 'line' not in s:
                return False
            if line_num == '1':
                return bool(re.search(r'(\b1\b|1\s*[-я]?|перва)', s))
            if line_num == '2':
                return bool(re.search(r'(\b2\b|2\s*[-я]?|втора)', s))
            return False

        for nm in names:
            if _match('1', nm) or _match('2', nm):
                return True
        return False
    except Exception:
        return False



def ensure_sqlite_schema():
    """Для SQLite: автоматически добавляем недостающие колонки, чтобы проект не падал,
    даже если миграции не были применены."""
    uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
    if not uri.startswith('sqlite'):
        return
    # sqlite:////abs/path.db или sqlite:///relative.db
    db_file = uri.replace('sqlite:///', '', 1)
    if not os.path.isabs(db_file):
        db_file = os.path.join(app.root_path, db_file)

    try:
        conn = sqlite3.connect(db_file)
        cur = conn.cursor()

        def has_column(table, col):
            cur.execute(f"PRAGMA table_info({table});")
            cols = [r[1] for r in cur.fetchall()]
            return col in cols

        def has_table(table):
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?;", (table,))
            return cur.fetchone() is not None


        # support_tickets
        if not has_column('support_tickets', 'ticket_type'):
            cur.execute("ALTER TABLE support_tickets ADD COLUMN ticket_type VARCHAR(20) DEFAULT 'issue'")
        if not has_column('support_tickets', 'category_id'):
            cur.execute("ALTER TABLE support_tickets ADD COLUMN category_id INTEGER")
        if not has_column('support_tickets', 'inn'):
            cur.execute("ALTER TABLE support_tickets ADD COLUMN inn VARCHAR(20)")
        if not has_column('support_tickets', 'address'):
            cur.execute("ALTER TABLE support_tickets ADD COLUMN address TEXT")
        if not has_column('support_tickets', 'organization'):
            cur.execute("ALTER TABLE support_tickets ADD COLUMN organization VARCHAR(255)")
        if not has_column('support_tickets', 'is_spam'):
            cur.execute("ALTER TABLE support_tickets ADD COLUMN is_spam INTEGER DEFAULT 0")
        if not has_column('support_tickets', 'close_reason'):
            cur.execute("ALTER TABLE support_tickets ADD COLUMN close_reason TEXT")

        # users UI theme
        if not has_column('users', 'ui_theme'):
            cur.execute("ALTER TABLE users ADD COLUMN ui_theme VARCHAR(12) DEFAULT 'light'")

        # users: единая модель (email_verified, position, role)
        if not has_column('users', 'email_verified'):
            cur.execute("ALTER TABLE users ADD COLUMN email_verified BOOLEAN DEFAULT 0")
        if not has_column('users', 'position'):
            cur.execute("ALTER TABLE users ADD COLUMN position VARCHAR(100)")
        if not has_column('users', 'role'):
            cur.execute("ALTER TABLE users ADD COLUMN role VARCHAR(20) DEFAULT 'client'")

        # users: предложенные реквизиты (из CRM)
        if not has_column('users', 'suggested_organization'):
            cur.execute("ALTER TABLE users ADD COLUMN suggested_organization VARCHAR(200)")
        if not has_column('users', 'suggested_inn'):
            cur.execute("ALTER TABLE users ADD COLUMN suggested_inn VARCHAR(12)")
        if not has_column('users', 'suggested_address'):
            cur.execute("ALTER TABLE users ADD COLUMN suggested_address VARCHAR(300)")

        # support_tickets: ссылка на клиента в users
        if not has_column('support_tickets', 'client_id'):
            cur.execute("ALTER TABLE support_tickets ADD COLUMN client_id INTEGER")

        # Справочник причин закрытия (подстатусы для 'Завершена')
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='ticket_close_reasons';")
        if not cur.fetchone():
            cur.execute("""
                CREATE TABLE ticket_close_reasons (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    code VARCHAR(32) UNIQUE NOT NULL,
                    name VARCHAR(80) NOT NULL,
                    sort_order INTEGER NOT NULL DEFAULT 0,
                    is_active BOOLEAN NOT NULL DEFAULT 1,
                    require_comment BOOLEAN NOT NULL DEFAULT 0
                );
            """)
            cur.execute("CREATE INDEX idx_ticket_close_reasons_active ON ticket_close_reasons(is_active, sort_order);")

        # Дефолтные причины закрытия (можно редактировать в админке)
        try:
            cur.execute("INSERT OR IGNORE INTO ticket_close_reasons(code,name,sort_order,is_active,require_comment) VALUES('spam','Спам',10,1,0);")
            cur.execute("INSERT OR IGNORE INTO ticket_close_reasons(code,name,sort_order,is_active,require_comment) VALUES('duplicate','Дубликат',20,1,0);")
            cur.execute("INSERT OR IGNORE INTO ticket_close_reasons(code,name,sort_order,is_active,require_comment) VALUES('wrong','Ошибочная',30,1,0);")
            cur.execute("INSERT OR IGNORE INTO ticket_close_reasons(code,name,sort_order,is_active,require_comment) VALUES('withdrawn','Отозванная',40,1,0);")
            # Системная причина (обычно не показываем): авто-закрытие после ожидания
            cur.execute("INSERT OR IGNORE INTO ticket_close_reasons(code,name,sort_order,is_active,require_comment) VALUES('no_response','Нет ответа клиента (24ч)',900,0,0);")
        except Exception:
            pass

        # Нормализация исторических статусов (чтобы в списках было 4 состояния)
        try:
            # Принята -> В работе
            cur.execute("UPDATE support_tickets SET status='В работе' WHERE status='Принята';")
            # Ожидает подтверждения клиента -> Ожидание
            cur.execute("UPDATE support_tickets SET status='Ожидание' WHERE status='Ожидает подтверждения клиента';")
            # Старые подстатусы закрытия -> Завершена + причина
            cur.execute("UPDATE support_tickets SET close_reason='spam', is_spam=1, status='Завершена' WHERE status='Спам';")
            cur.execute("UPDATE support_tickets SET close_reason='duplicate', status='Завершена' WHERE status='Дубликат';")
            cur.execute("UPDATE support_tickets SET close_reason='wrong', status='Завершена' WHERE status IN ('Ошибочная','Ошибочно');")
            cur.execute("UPDATE support_tickets SET close_reason='withdrawn', status='Завершена' WHERE status='Отозвано';")
        except Exception:
            pass



        # ticket_messages
        if not has_column('ticket_messages', 'edited_at'):
            cur.execute("ALTER TABLE ticket_messages ADD COLUMN edited_at TEXT")
        if not has_column('ticket_messages', 'edited_by_id'):
            cur.execute("ALTER TABLE ticket_messages ADD COLUMN edited_by_id INTEGER")

        # users: настройки уведомлений (in-app)
        for table in ('users',):
            if not has_column(table, 'notify_inapp_enabled'):
                cur.execute(f"ALTER TABLE {table} ADD COLUMN notify_inapp_enabled BOOLEAN DEFAULT 1")
            if not has_column(table, 'notify_event_assigned'):
                cur.execute(f"ALTER TABLE {table} ADD COLUMN notify_event_assigned BOOLEAN DEFAULT 1")
            if not has_column(table, 'notify_event_customer_reply'):
                cur.execute(f"ALTER TABLE {table} ADD COLUMN notify_event_customer_reply BOOLEAN DEFAULT 1")
            if not has_column(table, 'notify_event_status'):
                cur.execute(f"ALTER TABLE {table} ADD COLUMN notify_event_status BOOLEAN DEFAULT 1")

        # --- Миграция данных: legacy end_users -> users(role='client') (если таблица ещё есть) ---
        # 1) создаём клиентов в users по email (если ещё нет)
        try:
            if has_table('end_users'):
                cur.execute("SELECT id, name, last_name, patronymic, email, password, is_active, phone, organization, position, created_at, inn, address, email_verified, ui_theme, notify_inapp_enabled, notify_event_assigned, notify_event_customer_reply, notify_event_status FROM end_users")
                rows = cur.fetchall()
                for r in rows:
                    eu_id, name, last_name, patronymic, email, password, is_active, phone, organization, position, created_at, inn, address, email_verified, ui_theme, n1, n2, n3, n4 = r
                    if not email:
                        continue
                    cur.execute("SELECT id FROM users WHERE lower(email)=lower(?) LIMIT 1", (email,))
                    if cur.fetchone():
                        continue
                    # username для клиентов = email
                    cur.execute(
                        """INSERT INTO users(username,name,last_name,patronymic,email,password,role,is_active,phone,organization,position,created_at,inn,address,email_verified,ui_theme,notify_inapp_enabled,notify_event_assigned,notify_event_customer_reply,notify_event_status)
                             VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            email, name, last_name, patronymic, email, password, 'client', is_active or 1,
                            phone, organization, position, created_at, inn, address, email_verified or 0,
                            ui_theme or 'light', n1 if n1 is not None else 1, n2 if n2 is not None else 1,
                            n3 if n3 is not None else 1, n4 if n4 is not None else 1
                        )
                    )
        except Exception:
            pass

        # 2) проставляем support_tickets.client_id по legacy user_id (end_users.id)
        try:
            if has_table('end_users'):
                cur.execute(
                    """UPDATE support_tickets
                           SET client_id = (
                             SELECT u.id FROM users u
                              WHERE lower(u.email)=lower(support_tickets.email)
                              LIMIT 1
                           )
                         WHERE client_id IS NULL AND email IS NOT NULL"""
                )
        except Exception:
            pass
        # 3) после миграции таблица end_users больше не нужна — удаляем
        try:
            if has_table('end_users'):
                cur.execute("DROP TABLE IF EXISTS end_users;")
        except Exception:
            pass



        # notifications table (если нет — создаём)
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='notifications';")
        if not cur.fetchone():
            cur.execute("""
                CREATE TABLE notifications (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    recipient_type VARCHAR(20) NOT NULL,
                    recipient_id INTEGER NOT NULL,
                    event_type VARCHAR(50) NOT NULL,
                    title VARCHAR(200) NOT NULL,
                    body TEXT,
                    url VARCHAR(400),
                    ticket_id INTEGER,
                    is_read BOOLEAN NOT NULL DEFAULT 0,
                    created_at DATETIME NOT NULL,
                    dedupe_key VARCHAR(120)
                );
            """)
            cur.execute("CREATE INDEX idx_notifications_recipient ON notifications(recipient_type, recipient_id);")
            cur.execute("CREATE INDEX idx_notifications_unread ON notifications(recipient_type, recipient_id, is_read);")
            cur.execute("CREATE INDEX idx_notifications_created ON notifications(created_at);")
            cur.execute("CREATE INDEX idx_notifications_ticket_id ON notifications(ticket_id);")
        else:
            # существующая таблица: добавляем ticket_id (если не было)
            if not has_column('notifications', 'ticket_id'):
                cur.execute("ALTER TABLE notifications ADD COLUMN ticket_id INTEGER")
                try:
                    cur.execute("CREATE INDEX IF NOT EXISTS idx_notifications_ticket_id ON notifications(ticket_id);")
                except Exception:
                    pass

            # backfill ticket_id из url для старых строк
            try:
                cur.execute("SELECT id, url FROM notifications WHERE ticket_id IS NULL AND url IS NOT NULL;")
                rows = cur.fetchall() or []
                for nid, url in rows:
                    try:
                        s = (url or '').strip()
                        if not s:
                            continue
                        m = _TICKET_URL_RE.search(s) or _TICKET_URL_RE_ALT.search(s)
                        if not m:
                            continue
                        tid = int(m.group(1))
                        cur.execute("UPDATE notifications SET ticket_id = ? WHERE id = ?", (tid, int(nid)))
                    except Exception:
                        continue
            except Exception:
                pass

        # notification_global_settings table
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='notification_global_settings';")
        if not cur.fetchone():
            cur.execute("""
                CREATE TABLE notification_global_settings (
                    id INTEGER PRIMARY KEY,
                    enabled BOOLEAN NOT NULL DEFAULT 1,
                    enabled_for_operators BOOLEAN NOT NULL DEFAULT 1,
                    enabled_for_clients BOOLEAN NOT NULL DEFAULT 1,
                    event_assigned BOOLEAN NOT NULL DEFAULT 1,
                    event_customer_reply BOOLEAN NOT NULL DEFAULT 1,
                    event_operator_reply BOOLEAN NOT NULL DEFAULT 1,
                    event_status BOOLEAN NOT NULL DEFAULT 1,
                    event_priority BOOLEAN NOT NULL DEFAULT 1,
                    event_opchat BOOLEAN NOT NULL DEFAULT 1,
                    event_new_ticket BOOLEAN NOT NULL DEFAULT 1,
                    event_important BOOLEAN NOT NULL DEFAULT 1,
                    event_sla BOOLEAN NOT NULL DEFAULT 1,
                    updated_at DATETIME NOT NULL
                );
            """)
            # seed singleton row
            cur.execute("""
                INSERT INTO notification_global_settings
                (id, enabled, enabled_for_operators, enabled_for_clients, event_assigned, event_customer_reply, event_operator_reply, event_status, event_priority, event_opchat, event_new_ticket, event_important, event_sla, updated_at)
                VALUES (1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, CURRENT_TIMESTAMP);
            """)
        else:
            if not has_column('notification_global_settings', 'event_operator_reply'):
                cur.execute("ALTER TABLE notification_global_settings ADD COLUMN event_operator_reply BOOLEAN NOT NULL DEFAULT 1")
            if not has_column('notification_global_settings', 'event_priority'):
                cur.execute("ALTER TABLE notification_global_settings ADD COLUMN event_priority BOOLEAN NOT NULL DEFAULT 1")
            if not has_column('notification_global_settings', 'event_opchat'):
                cur.execute("ALTER TABLE notification_global_settings ADD COLUMN event_opchat BOOLEAN NOT NULL DEFAULT 1")
            if not has_column('notification_global_settings', 'event_new_ticket'):
                cur.execute("ALTER TABLE notification_global_settings ADD COLUMN event_new_ticket BOOLEAN NOT NULL DEFAULT 1")
            if not has_column('notification_global_settings', 'event_important'):
                cur.execute("ALTER TABLE notification_global_settings ADD COLUMN event_important BOOLEAN NOT NULL DEFAULT 1")
            if not has_column('notification_global_settings', 'event_sla'):
                cur.execute("ALTER TABLE notification_global_settings ADD COLUMN event_sla BOOLEAN NOT NULL DEFAULT 1")


        conn.commit()
    except Exception as e:
        print(f"[DB] ensure_sqlite_schema error: {e}")
    finally:
        try:
            conn.close()
        except Exception:
            pass


# === Инициализация (таблицы + первичные данные) ===
with app.app_context():
    db.create_all()
    ensure_sqlite_schema()

    # 1) Админ по умолчанию
    if not User.query.filter_by(username='admin').first():
        admin = User(
            username='admin',
            name='Системы',
            last_name='Администратор',
            patronymic='Техподдержки',
            email='monitoring@ciktrb.ru',
            password=generate_password_hash('admin_cjs', method='pbkdf2:sha256'),
            role='admin',
            department_id=None
        )
        db.session.add(admin)

    # 2) Отделы по умолчанию (если их ещё нет)
    if Department.query.count() == 0:
        for dept_name in DEFAULT_DEPARTMENTS:
            db.session.add(Department(name=dept_name))

    db.session.commit()

    # 3) BitrixSettings: гарантируем наличие записи для каждого отдела
    for dept in Department.query.all():
        if not BitrixSettings.query.filter_by(department=dept.name).first():
            db.session.add(BitrixSettings(department=dept.name))
    db.session.commit()

    # 4) Категории заявок по умолчанию
    if TicketCategory.query.count() == 0:
        for code, name, sort_order in DEFAULT_TICKET_CATEGORIES:
            db.session.add(TicketCategory(code=code, name=name, sort_order=sort_order, is_active=True))
        db.session.commit()

    # 5) Настройка "входящего" отдела (куда падают все новые заявки)
    # ВАЖНО: должна устанавливаться независимо от того, создавались ли категории прямо сейчас
    if not Settings.query.filter_by(key='default_intake_department_id').first():
        first_line = Department.query.filter_by(name="1ая линия ТП").first()
        if first_line:
            set_setting('default_intake_department_id', str(first_line.id))
            db.session.commit()

    # 6) Демо-шаблон ответа
    instr_file = "ТП_Инструкция по первичной диагностике сетевых неисправностей (2).docx"
    if not ResponseTemplate.query.filter_by(title="Инструкция: диагностика сети").first():
        tmpl = ResponseTemplate(
            title="Инструкция: диагностика сети",
            body=(
                "Здравствуйте!\n"
                "Для ускорения диагностики сетевой проблемы, пожалуйста, выполните следующие действия:\n"
                "1. Откройте командную строку (Win+R → введите cmd)\n"
                "2. Выполните команды и пришлите скриншоты:\n"
                "   - ipconfig /all\n"
                "   - ping [адрес]\n"
                "   - tracert [адрес]\n"
                "Подробная инструкция во вложении."
            ),
            files=json.dumps([instr_file]),
            department="Все"
        )
        db.session.add(tmpl)
        db.session.commit()

    # 7) База знаний: минимальный сид (чтобы было что посмотреть)
    if KnowledgeBaseCategory.query.count() == 0:
        net = KnowledgeBaseCategory(title='Сеть и доступ', sort_order=10, is_active=True)
        pc = KnowledgeBaseCategory(title='ПК и Windows', sort_order=20, is_active=True)
        db.session.add(net)
        db.session.add(pc)
        db.session.commit()

    if KnowledgeBaseArticle.query.count() == 0:
        net_cat = KnowledgeBaseCategory.query.filter_by(title='Сеть и доступ').first()
        a = KnowledgeBaseArticle(
            title='Диагностика сети: ping / tracert / ipconfig',
            category_id=(net_cat.id if net_cat else None),
            tags='ping, tracert, ipconfig, сеть',
            summary=(
                "Здравствуйте!\n"
                "Пожалуйста, выполните шаги ниже и пришлите результат в ответном сообщении.\n\n"
                "[b]1) ipconfig[/b]\n"
                "[code]ipconfig /all[/code]\n"
                "[b]2) ping[/b]\n"
                "[code]ping 8.8.8.8 -n 20[/code]\n"
                "[b]3) tracert[/b]\n"
                "[code]tracert 8.8.8.8[/code]\n"
                "Если адрес назначения другой — замените 8.8.8.8 на нужный адрес."
            ),
            body=(
                "[b]Подробная инструкция[/b]\n\n"
                "1) Откройте командную строку: Win+R → cmd → Enter.\n"
                "2) Выполните команды по очереди.\n"
                "3) Скопируйте вывод (правой кнопкой мыши → Выделить → Enter) или сделайте скриншоты.\n\n"
                "[b]Что означают результаты[/b]\n"
                "• ping показывает задержку и потери пакетов.\n"
                "• tracert показывает маршрут до узла и место, где возникают задержки/обрывы.\n"
                "• ipconfig помогает проверить настройки сети (IP, шлюз, DNS)."
            ),
            is_published=True,
        )
        db.session.add(a)
        db.session.commit()




# =========================
# CLI команды
# =========================
@app.cli.command("seed")
def seed_command():
    """Инициализировать/обновить справочники (идемпотентно)."""
    with app.app_context():
        # Повторяем то, что делается при старте — безопасно.
        # Если ты отключишь автосид на старте, эта команда станет основным способом.
        db.create_all()
        # Админ по умолчанию
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                name='Системы',
                last_name='Администратор',
                patronymic='Техподдержки',
                email='monitoring@ciktrb.ru',
                password=generate_password_hash('admin_cjs', method='pbkdf2:sha256'),
                role='admin',
                department_id=None
            )
            db.session.add(admin)
        if Department.query.count() == 0:
            for dept_name in DEFAULT_DEPARTMENTS:
                db.session.add(Department(name=dept_name))
        db.session.commit()
        for dept in Department.query.all():
            if not BitrixSettings.query.filter_by(department=dept.name).first():
                db.session.add(BitrixSettings(department=dept.name))
        db.session.commit()
        if TicketCategory.query.count() == 0:
            for code, name, sort_order in DEFAULT_TICKET_CATEGORIES:
                db.session.add(TicketCategory(code=code, name=name, sort_order=sort_order, is_active=True))
            db.session.commit()
        if not Settings.query.filter_by(key='default_intake_department_id').first():
            first_line = Department.query.filter_by(name="1ая линия ТП").first()
            if first_line:
                set_setting('default_intake_department_id', str(first_line.id))
                db.session.commit()
        print("✅ Seed выполнен")


@app.cli.command("mail-parser")
def mail_parser_command():
    """Запустить парсер почты в CLI (в отдельном процессе, без потоков и reloader)."""
    print("📨 Mail parser запущен (Ctrl+C чтобы остановить)")
    backoff = 30
    while True:
        try:
            _run_mail_check_once()
            backoff = 30
        except KeyboardInterrupt:
            print("\n🛑 Mail parser остановлен")
            break
        except Exception as e:
            print(f"[MAIL PARSER] Ошибка: {e}. Следующая попытка через {backoff} сек.")
            backoff = min(backoff * 2, 300)
        pytime.sleep(backoff)


def create_bitrix_task(ticket):
    if not BITRIX_WEBHOOK_URL:
        print("[BITRIX] Вебхук не настроен (BITRIX_WEBHOOK_URL пустой)")
        return False, "Вебхук Битрикс не настроен"
    
    dept_name = ticket.department.name if ticket.department else None
    if not dept_name:
        print(f"[BITRIX] Отдел не указан для заявки #{ticket.id}")
        return False, "Отдел не указан"
    
    setting = BitrixSettings.query.filter_by(department=dept_name).first()
    if not setting:
        print(f"[BITRIX] Настройки не найдены для отдела: {dept_name}")
        return False, f"Настройки для отдела '{dept_name}' не найдены"
    
    # Получаем сопровождающих безопасно
    try:
        accomplices = setting.get_accomplices_list() if hasattr(setting, 'get_accomplices_list') else []
    except:
        accomplices = []
        if setting.accomplices:
            try:
                accomplices = [int(x.strip()) for x in setting.accomplices.split(',') if x.strip().isdigit()]
            except:
                accomplices = []
    
    params = {
        "responsible": setting.responsible_id or '4519',
        "group_id": 82,
        "accomplices": accomplices,
        "auditors": [4038],
        "tags": ["Внутренние"]
    }
    
    task_data = {
        "fields": {
            "TITLE": f"ТП #{ticket.id}: {ticket.subject[:100]}",
            "DESCRIPTION": f"Заявка от: {ticket.name} ({ticket.email})\n\nОписание:\n{ticket.message}",
            "RESPONSIBLE_ID": params["responsible"],
            "GROUP_ID": params["group_id"],
            "ACCOMPLICES": params["accomplices"],
            "AUDITORS": params["auditors"],
            "CREATED_BY": 4519,
            "PRIORITY": "1",
            "ALLOW_CHANGE_DEADLINE": "Y",
            "ALLOW_TIME_TRACKING": "Y",
            "TASK_CONTROL": "Y",
            "TAGS": params.get("tags", [])
        }
    }
    
    try:
        print(f"[BITRIX] Отправка запроса к {BITRIX_WEBHOOK_URL}")
        response = requests.post(BITRIX_WEBHOOK_URL, json=task_data, verify=False, timeout=15)
        print(f"[BITRIX] Ответ: {response.status_code} {response.text[:200]}")
        
        if response.status_code == 200:
            result = response.json()
            task_id = result.get("result", {}).get("task", {}).get("id")
            if task_id:
                ticket.bitrix_task_id = str(task_id)
                db.session.commit()
                print(f"[BITRIX] ✅ Задача создана, ID: {task_id}")
                return True, str(task_id)
            else:
                error_msg = f"Не найден task.id в ответе: {result}"
                print(f"[BITRIX] ❌ {error_msg}")
                return False, error_msg
        else:
            error_msg = f"Ошибка {response.status_code}: {response.text}"
            print(f"[BITRIX] ❌ {error_msg}")
            return False, error_msg
    except Exception as e:
        error_msg = f"Исключение: {e}"
        print(f"[BITRIX] ❌ {error_msg}")
        import traceback
        traceback.print_exc()
        return False, error_msg

def log_ticket_change(ticket_id, user_id, field, old_value, new_value, note=None):
    """Логирует изменение в заявке"""
    if old_value == new_value:
        return
    
    history_entry = TicketHistory(
        ticket_id=ticket_id,
        user_id=user_id,
        field=field,
        old_value=str(old_value) if old_value is not None else None,
        new_value=str(new_value) if new_value is not None else None,
        note=note  # ← ДОБАВЛЕНО
    )
    db.session.add(history_entry)

def update_bitrix_task_status(ticket):
    if not ticket.bitrix_task_id or not BITRIX_UPDATE_WEBHOOK_URL:
        print("[BITRIX] Вебхук обновления не настроен или нет bitrix_task_id")
        return False
    update_data = {
        "taskId": ticket.bitrix_task_id,
        "fields": {
            "mark": "P" if ticket.status == "Завершена" else "N"
        }
    }
    try:
        response = requests.post(BITRIX_UPDATE_WEBHOOK_URL, json=update_data, verify=False, timeout=10)
        print(f"[BITRIX] Ответ при обновлении статуса: {response.status_code} {response.text}")
        return response.status_code == 200
    except Exception as e:
        print(f"[BITRIX] Ошибка при обновлении статуса: {e}")
        return False

# ===== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ РАБОТЫ С БИТРИКС24 =====

def get_company_by_id(base_url, company_id):
    """Получение данных компании по её ID с извлечением ИНН и адреса из любых полей"""
    try:
        company_url = f"{base_url}crm.company.get.json"
        company_resp = requests.post(company_url, json={"id": company_id}, verify=False, timeout=10)
        
        if company_resp.status_code == 200:
            company = company_resp.json().get("result", {})
            if company:
                # Извлекаем название
                company_name = company.get("TITLE", "").strip()
                
                # 🔑 УМНЫЙ ПОИСК ИНН: сканируем ВСЕ поля компании
                inn = ""
                # Сначала ищем по названию поля (регистронезависимо)
                for key, value in company.items():
                    if not value:
                        continue
                    key_lower = str(key).lower()
                    # Проверяем название поля на ключевые слова
                    if any(word in key_lower for word in ['inn', 'инн', 'tax', ' taxpayer', 'ипн', 'кпп']):
                        value_str = str(value).strip().replace(' ', '').replace('-', '').replace('_', '').replace('№', '')
                        # Проверяем формат ИНН/КПП (9-14 цифр)
                        if len(value_str) in (9, 10, 12, 14) and value_str.isdigit():
                            inn = value_str
                            print(f"[Bitrix] ИНН найден в поле {key}: {inn}")
                            break
                
                # Если не нашли по названию поля - ищем по формату во всех значениях
                if not inn:
                    for key, value in company.items():
                        if not value:
                            continue
                        value_str = str(value).strip().replace(' ', '').replace('-', '').replace('_', '').replace('№', '')
                        if len(value_str) in (9, 10, 12, 14) and value_str.isdigit():
                            # Дополнительная проверка: исключаем похожие на даты или номера
                            if not any(word in str(key).lower() for word in ['date', 'time', 'phone', 'version', 'id']):
                                inn = value_str
                                print(f"[Bitrix] ИНН найден по формату в поле {key}: {inn}")
                                break
                
                # 🔑 УМНЫЙ ПОИСК АДРЕСА: сканируем ВСЕ поля
                address = ""
                # Приоритетные поля
                address_fields = [
                    'ADDRESS', 'ADDRESS_LEGAL', 'ADDRESS_FULL', 'REG_ADDRESS', 
                    'UF_CRM_ADDRESS', 'UF_CRM_1678901234', 'UF_CRM_COMPANY_ADDRESS'
                ]
                for field in address_fields:
                    if field in company and company[field]:
                        addr_value = str(company[field]).strip()
                        if len(addr_value) > 5:  # Минимальная длина адреса
                            address = addr_value
                            print(f"[Bitrix] Адрес найден в поле {field}: {address}")
                            break
                
                # Если не нашли - сканируем все поля на ключевые слова
                if not address:
                    for key, value in company.items():
                        if not value:
                            continue
                        value_str = str(value).lower()
                        # Ищем признаки адреса (улица, дом, город, республика и т.д.)
                        if any(word in value_str for word in ['улиц', 'ул.', 'проспект', 'пр.', 'дом', 'д.', 
                                                              'корпус', 'к.', 'строение', 'стр.', 'город', 
                                                              'г.', 'республик', 'респ.', 'область', 'ул',
                                                              'адрес', 'адрес:', 'фактическ', 'юридическ',
                                                              'уфа', 'башкортостан', 'рб', 'республика']):
                            addr_value = str(company[key]).strip()
                            if len(addr_value) > 10:  # Минимальная длина для адреса
                                address = addr_value
                                print(f"[Bitrix] Адрес найден по ключевым словам в поле {key}: {address}")
                                break
                
                # Если адрес всё ещё пустой, но есть поле АДРЕС - берём его целиком
                if not address and 'АДРЕС' in company:
                    address = str(company['АДРЕС']).strip()
                elif not address and 'ADDRESS' in company:
                    address = str(company['ADDRESS']).strip()
                
                return {
                    "company": company_name,
                    "inn": inn,
                    "address": address
                }
        return None
    except Exception as e:
        print(f"[Bitrix] Ошибка получения компании ID {company_id}: {e}")
        import traceback
        traceback.print_exc()
        return None

def search_company_by_email(base_url, email):
    """Поиск компании напрямую по email (проверяем все возможные поля с email)"""
    try:
        # Стандартные поля для поиска email в компаниях
        email_fields = [
            "EMAIL", 
            "UF_CRM_EMAIL", 
            "UF_CRM_COMPANY_EMAIL",
            "UF_CRM_1678901234",
            "UF_CRM_CONTACT_EMAIL"
        ]
        
        for field in email_fields:
            search_url = f"{base_url}crm.company.list.json"
            search_data = {
                "order": {"DATE_CREATE": "DESC"},
                "filter": {field: email},
                "select": ["ID", "TITLE"]
            }
            try:
                resp = requests.post(search_url, json=search_data, verify=False, timeout=10)
                if resp.status_code == 200:
                    companies = resp.json().get("result", [])
                    if companies:
                        company_id = companies[0].get("ID")
                        print(f"[Bitrix] Компания найдена по полю {field}, ID={company_id}")
                        return get_company_by_id(base_url, company_id)
            except:
                continue
        
        return None
    except Exception as e:
        print(f"[Bitrix] Ошибка поиска компании по email: {e}")
        return None

def search_company_by_domain(base_url, domain):
    """Поиск компании по домену (последний шанс)"""
    try:
        # Поиск по названию (часто содержит домен)
        search_url = f"{base_url}crm.company.list.json"
        search_data = {
            "order": {"DATE_CREATE": "DESC"},
            "filter": {"%TITLE": domain},
            "select": ["ID", "TITLE"]
        }
        resp = requests.post(search_url, json=search_data, verify=False, timeout=10)
        
        if resp.status_code == 200:
            companies = resp.json().get("result", [])
            if companies:
                company_id = companies[0].get("ID")
                print(f"[Bitrix] Компания найдена по домену в названии, ID={company_id}")
                return get_company_by_id(base_url, company_id)
        
        return None
    except Exception as e:
        print(f"[Bitrix] Ошибка поиска компании по домену {domain}: {e}")
        return None

# ===== ОСНОВНАЯ ФУНКЦИЯ ПОИСКА КОМПАНИИ ПО EMAIL =====

def get_company_from_bitrix(email):
    """
    Универсальный поиск данных компании по email через контакт
    """
    if not BITRIX_CONTAKT_WEBHOOK_URL:
        print("[Bitrix] BITRIX_CONTAKT_WEBHOOK_URL не задан")
        return {
            "company": "",
            "inn": "",
            "address": ""
        }

    # Формируем базовый URL (без методов)
    base_url = BITRIX_CONTAKT_WEBHOOK_URL.rstrip('/')
    for method in ['crm.contact.add', 'crm.contact.list', 'crm.company.get', 'crm.company.list']:
        if base_url.endswith(method):
            base_url = base_url.replace(method, '').rstrip('/')
    base_url = base_url.rstrip('/') + '/'

    try:
        # Шаг 1: Найти контакт по email
        contact_url = f"{base_url}crm.contact.list.json"
        contact_data = {
            "filter": {"EMAIL": email},
            "select": ["ID", "COMPANY_ID", "NAME", "LAST_NAME", "EMAIL"]
        }
        contact_resp = requests.post(contact_url, json=contact_data, verify=False, timeout=10)
        contacts = contact_resp.json().get("result", [])
        
        if not contacts:
            print(f"[Bitrix] Контакт с email {email} не найден")
            return {
                "company": "",
                "inn": "",
                "address": ""
            }
        
        contact = contacts[0]
        company_id = contact.get("COMPANY_ID")
        
        if not company_id or company_id in ("0", "None", None):
            print(f"[Bitrix] У контакта нет привязанной компании")
            return {
                "company": "",
                "inn": "",
                "address": ""
            }
        
        # Шаг 2: Получить компанию по ID
        company_url = f"{base_url}crm.company.get.json"
        company_resp = requests.post(company_url, json={"id": company_id}, verify=False, timeout=10)
        company = company_resp.json().get("result", {})
        
        if not company:
            print(f"[Bitrix] Компания с ID {company_id} не найдена")
            return {
                "company": "",
                "inn": "",
                "address": ""
            }
        
        # === ИЗВЛЕЧЕНИЕ ДАННЫХ С УМНЫМ ПОИСКОМ ===
        company_name = company.get("TITLE", "").strip()
        
        # 🔑 ПОИСК ИНН: сканируем ВСЕ поля компании на предмет 14-значного кода
        inn = ""
        for key, value in company.items():
            if not value:
                continue
            # Преобразуем в строку и очищаем
            value_str = str(value).strip().replace(' ', '').replace('-', '').replace('_', '').replace('№', '')
            # Ищем 14-значный ИНН (как в вашем примере 02741935165999)
            if len(value_str) == 14 and value_str.isdigit():
                inn = value_str
                print(f"[Bitrix] ИНН найден в поле {key}: {inn}")
                break
            # Также проверяем 10 и 12 цифр для других форматов
            elif len(value_str) in (10, 12) and value_str.isdigit():
                # Дополнительная проверка: исключаем похожие на даты или номера
                if not any(word in str(key).lower() for word in ['date', 'time', 'phone', 'version', 'id', 'number']):
                    inn = value_str
                    print(f"[Bitrix] ИНН найден по формату в поле {key}: {inn}")
                    break
        
        # 🔑 ПОИСК АДРЕСА: ищем по ключевым словам в значениях
        address = ""
        for key, value in company.items():
            if not value:
                continue
            value_str = str(value).lower()
            # Ищем признаки адреса (улица, дом, город, республика и т.д.)
            if any(word in value_str for word in ['улиц', 'ул.', 'дом', 'д.', 'город', 'г.', 
                                                  'республик', 'респ.', 'башкортостан', 'уфа',
                                                  'адрес', 'фактическ', 'юридическ', 'шафиев']):
                address = str(company[key]).strip()
                print(f"[Bitrix] Адрес найден в поле {key}: {address}")
                break
        
        # Если не нашли по ключевым словам - берем из стандартных полей
        if not address:
            address = str(company.get("ADDRESS", "")).strip() or str(company.get("ADDRESS_LEGAL", "")).strip()
        
        return {
            "company": company_name,
            "inn": inn,
            "address": address
        }
        
    except Exception as e:
        print(f"[Bitrix ERROR] Ошибка при поиске компании для email {email}: {e}")
        import traceback
        traceback.print_exc()
        return {
            "company": "",
            "inn": "",
            "address": ""
        }

def suggest_templates(message_text):
    message_lower = message_text.lower()
    suggestions = []
    if any(word in message_lower for word in ['ping', 'интернет', 'сеть', 'tracert', 'ipconfig']):
        suggestions.append("Инструкция: диагностика сети")
    if any(word in message_lower for word in ['битрикс', 'корпоративный портал', 'портал', 'crm']):
        suggestions.append("Инструкция: работа в Битрикс24")
    return suggestions

def add_bitrix_comment(ticket, comment_text, is_internal=False):
    if not ticket.bitrix_task_id or not BITRIX_COMMENT_WEBHOOK_URL:
        print("[BITRIX] Вебхук комментариев не настроен")
        return False
    prefix = "Служебный комментарий: " if is_internal else "Ответ оператора: "
    full_text = f"{prefix}{comment_text}"
    comment_data = {
        "TASKID": ticket.bitrix_task_id,
        "FIELDS": {
            "POST_MESSAGE": full_text
        }
    }
    try:
        response = requests.post(BITRIX_COMMENT_WEBHOOK_URL, json=comment_data, verify=False, timeout=10)
        print(f"[BITRIX] Ответ при добавлении комментария: {response.status_code} {response.text}")
        return response.status_code == 200
    except Exception as e:
        print(f"[BITRIX] Ошибка при добавлении комментария: {e}")
        return False
def translate_priority(priority):
    """Переводит приоритет с английского на русский"""
    if not priority:
        return '—'
    
    translations = {
        'low': 'Низкий',
        'medium': 'Средний', 
        'high': 'Высокий'
    }
    
    return translations.get(priority.lower(), priority)

def translate_status(status):
    """Переводит статусы на русский"""
    if not status:
        return '—'
    
    translations = {
        'new': 'Новая',
        'Новая': 'Новая',
        'accepted': 'Принята',
        'Принята': 'Принята',
        'in_progress': 'В работе',
        'В работе': 'В работе',
        'completed': 'Завершена',
        'Завершена': 'Завершена',
        'Ожидает подтверждения клиента': 'Ожидает подтверждения клиента',
        'waiting_for_client': 'Ожидает подтверждения клиента'
    }
    
    return translations.get(status, status)

def translate_field_name(field):
    """Переводит названия полей на русский"""
    translations = {
        'priority': 'Приоритет',
        'status': 'Статус',
        'department': 'Отдел',
        'subject': 'Тема',
        'message': 'Описание',
        'response': 'Ответ',
        'client_feedback': 'Обратная связь клиента',
        'system_note': 'Системное уведомление',
        'created_at': 'Дата создания',
        'updated_at': 'Дата обновления',
        'closed_at': 'Дата закрытия',
        'assigned_to': 'Назначено',
        'locked_by': 'Заблокировано',
        'locked_at': 'Время блокировки'
    }
    
    return translations.get(field, field)

# === Маршруты ===
@app.route('/api/upload', methods=['POST'])
@login_required
def upload_file_api():
    """Универсальная загрузка файлов для заявок и комментариев"""
    
    # Проверка наличия файлов
    if 'files' not in request.files:
        return jsonify({'success': False, 'error': 'Нет файлов для загрузки'}), 400
    
    files = request.files.getlist('files')
    
    # Проверка количества файлов
    if len(files) > 10:
        return jsonify({'success': False, 'error': 'Максимум 10 файлов за раз'}), 400
    
    uploaded_files = []
    total_size = 0
    
    # Разрешенные расширения
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt'}
    
    for file in files:
        if not file or not file.filename.strip():
            continue
        
        # Проверка расширения
        ext = file.filename.rsplit('.', 1)[1].lower()
        if ext not in allowed_extensions:
            return jsonify({
                'success': False,
                'error': f'Файл {file.filename} имеет недопустимый формат. '
                        f'Разрешены: PNG, JPG, PDF, DOC, DOCX, XLS, XLSX, TXT'
            }), 400
        
        # Проверка размера (5 МБ)
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > 5 * 1024 * 1024:
            return jsonify({
                'success': False,
                'error': f'Файл {file.filename} превышает 5 МБ'
            }), 400
        
        total_size += file_size
        if total_size > 50 * 1024 * 1024:  # 50 МБ общий лимит
            return jsonify({
                'success': False,
                'error': 'Общий размер файлов превышает лимит (50 МБ)'
            }), 400
        
        # Сохранение файла
        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        
        upload_folder = os.path.join('static', 'uploads', 'attachments')
        os.makedirs(upload_folder, exist_ok=True)
        
        file_path = os.path.join(upload_folder, unique_filename)
        file.save(file_path)
        
        # Определяем тип файла
        is_image = ext in {'png', 'jpg', 'jpeg', 'gif'}
        
        uploaded_files.append({
            'original_name': filename,
            'saved_name': unique_filename,
            'url': url_for('static', filename=f'uploads/attachments/{unique_filename}'),
            'type': 'image' if is_image else 'document',
            'size': file_size,
            'size_formatted': format_file_size(file_size)
        })
    
    if not uploaded_files:
        return jsonify({'success': False, 'error': 'Нет допустимых файлов'}), 400
    
    return jsonify({
        'success': True,
        'files': uploaded_files,
        'count': len(uploaded_files)
    })

def format_file_size(size_bytes):
    """Форматирование размера файла"""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} MB"

@app.route('/api/bitrix/company-by-email')
def bitrix_company_by_email():
    email = request.args.get('email', '').strip()
    if not email or '@' not in email:
        return jsonify({"success": False, "error": "Некорректный email"}), 400

    try:
        # Используем глобальный вебхук из .env
        company_data = get_company_from_bitrix(email)
        if company_data:
            return jsonify({
                "success": True,
                "company_name": company_data["company"],
                "inn": company_data["inn"],
                "address": company_data["address"]
            })
        else:
            return jsonify({
                "success": False,
                "error": "Компания не найдена"
            }), 404
    except Exception as e:
        print(f"[ERROR] Bitrix lookup failed for {email}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": "Ошибка при запросе к Битрикс24"}), 500


@app.route('/api/user-by-email')
@login_required
def api_user_by_email():
    """Поиск клиента (User.role='client') по email.

    Используется на единой странице создания заявки: если оператор вводит email клиента,
    подтягиваем ФИО/организацию/ИНН/адрес/телефон из локальной БД.
    """
    email = (request.args.get('email') or '').strip().lower()
    if not email or '@' not in email:
        return jsonify({"success": False, "error": "Некорректный email"}), 400

    try:
        u = User.query.filter(db.func.lower(User.email) == email, User.role == 'client').first()
        if not u:
            return jsonify({"success": False, "error": "Пользователь не найден"}), 404

        fio = " ".join([x for x in [u.last_name, u.name, u.patronymic] if x]).strip()
        fio = fio or u.email

        return jsonify({
            "success": True,
            "user_id": u.id,
            "full_name": fio,
            "phone": u.phone or "",
            "organization": u.organization or "",
            "inn": u.inn or "",
            "address": u.address or "",
        })
    except Exception as e:
        print(f"[ERROR] user-by-email failed for {email}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": "Ошибка при запросе к базе"}), 500
    


# === PRESENCE: кто смотрит/пишет в заявке ===
def _presence_user_key(user):
    if getattr(user, 'role', None) == 'client':
        return f"client_{user.id}", "client"
    return f"op_{user.id}", (user.role or "operator")

def _presence_display_name(user):
    if getattr(user, 'role', None) == 'client':
        fio = " ".join([x for x in [user.last_name, user.name, user.patronymic] if x]).strip()
        return fio or user.email or f"Клиент #{user.id}"
    fio = " ".join([x for x in [getattr(user, 'last_name', None), getattr(user, 'name', None), getattr(user, 'patronymic', None)] if x]).strip()
    return fio or getattr(user, 'username', None) or getattr(user, 'email', None) or f"Оператор #{user.id}"

@app.route('/api/ticket/<int:ticket_id>/presence', methods=['GET'])
@login_required
def api_ticket_presence(ticket_id):
    cutoff = utcnow() - timedelta(seconds=20)
    rows = TicketPresence.query.filter(
        TicketPresence.ticket_id == ticket_id,
        TicketPresence.last_seen >= cutoff
    ).all()

    viewers = []
    typers = []
    for r in rows:
        viewers.append({
            "display_name": r.display_name,
            "role": r.role,
            "is_typing": bool(r.is_typing)
        })
        if r.is_typing:
            typers.append(r.display_name)

    return jsonify({
        "success": True,
        "viewers": viewers,
        "typers": typers
    })

@app.route('/api/ticket/<int:ticket_id>/presence/stream')
@login_required
def api_ticket_presence_stream(ticket_id):
    # Server-Sent Events: даёт “онлайн” обновления без SocketIO/Redis
    def _gen():
        # ВАЖНО: генератор должен выполняться в рамках request/app context,
        # иначе Flask-SQLAlchemy падает с "Working outside of application context".
        try:
            while True:
                cutoff = utcnow() - timedelta(seconds=20)
                rows = TicketPresence.query.filter(
                    TicketPresence.ticket_id == ticket_id,
                    TicketPresence.last_seen >= cutoff
                ).all()

                viewers = []
                typers = []
                for r in rows:
                    viewers.append({
                        "display_name": r.display_name,
                        "role": r.role,
                        "is_typing": bool(r.is_typing)
                    })
                    if r.is_typing:
                        typers.append(r.display_name)

                payload = {"success": True, "viewers": viewers, "typers": typers}
                yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"
                pytime.sleep(2.5)
        except GeneratorExit:
            return

    resp = Response(stream_with_context(_gen()), mimetype='text/event-stream')
    # Чтобы SSE стабильно работал за прокси/браузерами
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Accel-Buffering'] = 'no'
    return resp


@app.route('/api/ticket/<int:ticket_id>/presence/heartbeat', methods=['POST'])
@login_required
def api_ticket_presence_heartbeat(ticket_id):
    data = request.get_json(silent=True) or {}
    is_typing = bool(data.get('is_typing', False))

    user_key, role = _presence_user_key(current_user)
    display_name = _presence_display_name(current_user)

    row = TicketPresence.query.filter_by(ticket_id=ticket_id, user_key=user_key).first()
    if not row:
        row = TicketPresence(ticket_id=ticket_id, user_key=user_key, role=role, display_name=display_name)

    row.is_typing = is_typing
    row.last_seen = utcnow()

    db.session.add(row)
    try:
        db.session.commit()
    except OperationalError:
        db.session.rollback()
        return jsonify({"success": False, "error": "readonly_db"}), 200
    return jsonify({"success": True})


# === OPERATOR CHAT: отдельный чат на каждый тикет (только для операторов) ===
def _ticket_is_closed(t: SupportTicket) -> bool:
    s = (t.status or '').strip().lower()
    # В финальных статусах операторский чат должен быть закрыт
    return s in [
        'завершена',
        'спам',
        'дубликат',
        'ошибочно',
        'ошибочная',
        'закрыта',
        'закрыто',
        'closed',
    ]


@app.route('/api/ticket/<int:ticket_id>/opchat/messages', methods=['GET'])
@login_required
def api_opchat_messages(ticket_id):
    if not isinstance(current_user, User):
        return jsonify({'success': False, 'error': 'forbidden'}), 403

    ticket = SupportTicket.query.get_or_404(ticket_id)
    limit = min(int(request.args.get('limit', 50)), 200)
    before_id = request.args.get('before_id', type=int)

    q = TicketOperatorChatMessage.query.filter_by(ticket_id=ticket.id).order_by(TicketOperatorChatMessage.id.desc())
    if before_id:
        q = q.filter(TicketOperatorChatMessage.id < before_id)
    rows = q.limit(limit).all()
    rows.reverse()

    items = []
    for m in rows:
        u = m.user
        fio = " ".join([x for x in [getattr(u, 'last_name', None), getattr(u, 'name', None), getattr(u, 'patronymic', None)] if x]).strip()
        items.append({
            'id': m.id,
            'user_id': m.user_id,
            'author': fio or getattr(u, 'username', None) or f"Оператор #{m.user_id}",
            'message': m.message,
            'created_at': m.created_at.strftime('%d.%m.%Y %H:%M') if m.created_at else ''
        })

    # Также вернём текущий unread для этого пользователя (по тикету)
    try:
        read_row = TicketOperatorChatRead.query.filter_by(ticket_id=ticket.id, user_id=current_user.id).first()
        last_read = int(getattr(read_row, 'last_read_message_id', 0) or 0)
        unread = (TicketOperatorChatMessage.query
                  .filter(TicketOperatorChatMessage.ticket_id == ticket.id,
                          TicketOperatorChatMessage.id > last_read,
                          TicketOperatorChatMessage.user_id != current_user.id)
                  .count())
    except Exception:
        unread = 0

    return jsonify({'success': True, 'ticket_closed': _ticket_is_closed(ticket), 'items': items, 'unread': unread})


@app.route('/api/ticket/<int:ticket_id>/opchat/mark-read', methods=['POST'])
@login_required
def api_opchat_mark_read(ticket_id):
    """Отметить операторский чат тикета прочитанным до last_id включительно."""
    if not isinstance(current_user, User):
        return jsonify({'success': False, 'error': 'forbidden'}), 403

    ticket = SupportTicket.query.get_or_404(ticket_id)
    data = request.get_json(silent=True) or {}
    last_id = int(data.get('last_id') or 0)
    if last_id < 0:
        last_id = 0

    row = TicketOperatorChatRead.query.filter_by(ticket_id=ticket.id, user_id=current_user.id).first()
    if not row:
        row = TicketOperatorChatRead(ticket_id=ticket.id, user_id=current_user.id, last_read_message_id=last_id)
    else:
        row.last_read_message_id = max(int(row.last_read_message_id or 0), last_id)
    db.session.add(row)
    try:
        db.session.commit()
    except OperationalError:
        db.session.rollback()
        return jsonify({"success": False, "error": "readonly_db"}), 200
    return jsonify({"success": True})


@app.route('/api/ticket/<int:ticket_id>/opchat/unread-count', methods=['GET'])
@login_required
def api_opchat_unread_count(ticket_id):
    """Количество непрочитанных сообщений операторского чата в конкретном тикете."""
    if not isinstance(current_user, User):
        return jsonify({'success': False, 'error': 'forbidden'}), 403

    ticket = SupportTicket.query.get_or_404(ticket_id)
    row = TicketOperatorChatRead.query.filter_by(ticket_id=ticket.id, user_id=current_user.id).first()
    last_read = int(getattr(row, 'last_read_message_id', 0) or 0)

    unread = (TicketOperatorChatMessage.query
              .filter(TicketOperatorChatMessage.ticket_id == ticket.id,
                      TicketOperatorChatMessage.id > last_read,
                      TicketOperatorChatMessage.user_id != current_user.id)
              .count())
    return jsonify({'success': True, 'unread': unread})


@app.route('/api/ticket/<int:ticket_id>/opchat/send', methods=['POST'])
@login_required
def api_opchat_send(ticket_id):
    if not isinstance(current_user, User):
        return jsonify({'success': False, 'error': 'forbidden'}), 403

    ticket = SupportTicket.query.get_or_404(ticket_id)
    if _ticket_is_closed(ticket):
        return jsonify({'success': False, 'error': 'ticket_closed'}), 400

    data = request.get_json(silent=True) or {}
    text = (data.get('message') or '').strip()
    if not text:
        return jsonify({'success': False, 'error': 'empty'}), 400
    if len(text) > 2000:
        text = text[:2000]

    msg = TicketOperatorChatMessage(ticket_id=ticket.id, user_id=current_user.id, message=text)
    db.session.add(msg)
    db.session.commit()

    # --- In-app уведомления: новое сообщение в операторском чате ---
    try:
        # имя автора
        author = ""
        try:
            u = db.session.get(User, msg.user_id)
            if u:
                author = " ".join([x for x in [getattr(u, 'last_name', None), getattr(u, 'name', None), getattr(u, 'patronymic', None)] if x]).strip() or (getattr(u, 'username', None) or '')
        except Exception:
            author = ""

        preview = (text or '').strip()
        if len(preview) > 120:
            preview = preview[:120] + '…'
        body = (author + ': ' if author else '') + preview

        for op in _operators_with_access_to_ticket(ticket):
            if isinstance(current_user, User) and op.id == current_user.id:
                continue
            create_inapp_notification(
                op,
                'opchat',
                f"Новое сообщение в оп. чате тикета #{ticket.id}",
                body,
                url_for('ticket_detail', ticket_id=ticket.id, open_opchat=1),
                dedupe_key=f"opchat:{ticket.id}:{msg.id}:op:{op.id}"
            )
    except Exception:
        pass

    return jsonify({'success': True, 'id': msg.id})


@app.route('/api/ticket/<int:ticket_id>/opchat/stream')
@login_required
def api_opchat_stream(ticket_id):
    if not isinstance(current_user, User):
        return Response('data: {"success":false}\n\n', mimetype='text/event-stream'), 403

    ticket = SupportTicket.query.get_or_404(ticket_id)
    after = request.args.get('after', default=0, type=int)

    def _gen():
        last_id = after
        try:
            while True:
                # Если тикет уже в финальном статусе — сообщаем и закрываем стрим.
                try:
                    fresh = SupportTicket.query.get(ticket.id)
                    if fresh is not None and _ticket_is_closed(fresh):
                        payload = {'success': True, 'ticket_closed': True, 'items': [], 'last_id': last_id}
                        yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"
                        break
                except Exception:
                    pass

                q = (TicketOperatorChatMessage.query
                     .filter(TicketOperatorChatMessage.ticket_id == ticket.id,
                             TicketOperatorChatMessage.id > last_id)
                     .order_by(TicketOperatorChatMessage.id.asc()))
                rows = q.all()

                items = []
                for m in rows:
                    u = m.user
                    fio = " ".join([x for x in [getattr(u, 'last_name', None), getattr(u, 'name', None), getattr(u, 'patronymic', None)] if x]).strip()
                    items.append({
                        'id': m.id,
                        'user_id': m.user_id,
                        'author': fio or getattr(u, 'username', None) or f"Оператор #{m.user_id}",
                        'message': m.message,
                        'created_at': m.created_at.strftime('%d.%m.%Y %H:%M') if m.created_at else ''
                    })
                    last_id = max(last_id, m.id)

                payload = {
                    'success': True,
                    'ticket_closed': _ticket_is_closed(ticket),
                    'items': items,
                    'last_id': last_id,
                }
                yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"
                pytime.sleep(2.0)
        except GeneratorExit:
            return

    return Response(stream_with_context(_gen()), mimetype='text/event-stream')


# ---------------- Shared departments quick actions ----------------
@app.route('/api/ticket/<int:ticket_id>/shared_departments/remove', methods=['POST'])
@login_required
def api_shared_departments_remove(ticket_id):
    """Быстро снять дополнительный отдел без смены основного department_id."""
    # определяем тип пользователя
    is_client = getattr(current_user, 'role', None) == 'client'
    is_operator = (not is_client) and isinstance(current_user, User) and (
        current_user.role in ['operator', 'admin'] or is_tp_operator(current_user)
    )
    if is_client or not is_operator:
        return jsonify({'success': False, 'error': 'forbidden'}), 403

    ticket = SupportTicket.query.get_or_404(ticket_id)
    data = request.get_json(silent=True) or {}
    dept_id = data.get('department_id')
    try:
        dept_id_int = int(dept_id)
    except Exception:
        return jsonify({'success': False, 'error': 'bad_department_id'}), 400

    # нельзя снимать основной отдел
    if ticket.department_id and int(ticket.department_id) == dept_id_int:
        return jsonify({'success': False, 'error': 'cannot_remove_main_department'}), 400

    # снять из shared
    try:
        rel = getattr(ticket, 'shared_departments_rel', None)
        if rel is None:
            return jsonify({'success': False, 'error': 'no_shared_departments'}), 400
        to_remove = None
        for d in list(rel):
            if getattr(d, 'id', None) == dept_id_int:
                to_remove = d
                break
        if to_remove:
            rel.remove(to_remove)
            try:
                log_ticket_change(
                    ticket.id,
                    current_user.id,
                    'shared_departments',
                    '—',
                    f"- {to_remove.name}",
                    'Снят дополнительный отдел'
                )
            except Exception:
                pass
            db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/ui/unread-counts', methods=['GET'])
@login_required
def api_ui_unread_counts():
    """Универсальные счётчики непрочитанного для UI (колокольчик, бейджи и т.п.)."""
    # unread notifications
    notif_unread = 0
    opchat_unread = 0

    try:
        recipient_type, recipient_id = _recipient_key(current_user)
        accessible_ids = _accessible_ticket_ids_for_notifications(current_user)

        # Глобальные уведомления (ticket_id is NULL) считаем быстро и точно
        global_unread = (Notification.query
                         .filter_by(recipient_type=recipient_type, recipient_id=recipient_id, is_read=False)
                         .filter(Notification.ticket_id.is_(None))
                         .count())

        if accessible_ids is None:
            # глобальный доступ -> можно просто count()
            ticket_unread = (Notification.query
                             .filter_by(recipient_type=recipient_type, recipient_id=recipient_id, is_read=False)
                             .filter(Notification.ticket_id.isnot(None))
                             .count())
            notif_unread = int(global_unread or 0) + int(ticket_unread or 0)
        else:
            if not accessible_ids:
                notif_unread = int(global_unread or 0)
            else:
                ticket_unread = (Notification.query
                                 .filter_by(recipient_type=recipient_type, recipient_id=recipient_id, is_read=False)
                                 .filter(Notification.ticket_id.in_(list(accessible_ids)))
                                 .count())
                notif_unread = int(global_unread or 0) + int(ticket_unread or 0)

        # safety: если где-то остались старые уведомления без ticket_id, но с url на тикет —
        # они попадут в global_unread; скорректируем на небольшом окне.
        try:
            raw = (Notification.query
                   .filter_by(recipient_type=recipient_type, recipient_id=recipient_id, is_read=False)
                   .order_by(Notification.created_at.desc())
                   .limit(300)
                   .all())
            notif_unread = len(_filter_notifications_for_user(current_user, raw)) if len(raw) < 300 else notif_unread
        except Exception:
            pass
    except Exception:
        notif_unread = 0

    # unread opchat: только для операторов
    if isinstance(current_user, User):
        try:
            # В этой системе 1/2 линия должна видеть все тикеты — считаем по всем тикетам.
            # unread = сообщения не от себя и с id > last_read (если last_read нет — 0)
            from sqlalchemy import func
            Read = TicketOperatorChatRead
            Msg = TicketOperatorChatMessage

            sub_last = (db.session.query(Read.ticket_id.label('t_id'), Read.last_read_message_id.label('lr'))
                        .filter(Read.user_id == current_user.id)
                        .subquery())

            opchat_unread = (db.session.query(func.count(Msg.id))
                             .outerjoin(sub_last, sub_last.c.t_id == Msg.ticket_id)
                             .filter(Msg.user_id != current_user.id)
                             .filter(Msg.id > func.coalesce(sub_last.c.lr, 0))
                             .scalar()) or 0
        except Exception:
            opchat_unread = 0

    total = int(notif_unread or 0) + int(opchat_unread or 0)
    return jsonify({'success': True, 'notifications': int(notif_unread or 0), 'opchat': int(opchat_unread or 0), 'total': total})


def _opchat_unread_threads_for_user(user_id: int, limit: int = 5):
    """Сводка непрочитанных тредов операторского чата по тикетам."""
    try:
        from sqlalchemy import func
        Read = TicketOperatorChatRead
        Msg = TicketOperatorChatMessage

        sub_last = (db.session.query(Read.ticket_id.label('t_id'), Read.last_read_message_id.label('lr'))
                    .filter(Read.user_id == user_id)
                    .subquery())

        sub_unread = (db.session.query(
                        Msg.ticket_id.label('ticket_id'),
                        func.count(Msg.id).label('unread'),
                        func.max(Msg.id).label('last_id'))
                     .outerjoin(sub_last, sub_last.c.t_id == Msg.ticket_id)
                     .filter(Msg.user_id != user_id)
                     .filter(Msg.id > func.coalesce(sub_last.c.lr, 0))
                     .group_by(Msg.ticket_id)
                     .subquery())

        rows = (db.session.query(sub_unread.c.ticket_id, sub_unread.c.unread, Msg.message, Msg.created_at, Msg.user_id)
                .join(Msg, Msg.id == sub_unread.c.last_id)
                .order_by(sub_unread.c.last_id.desc())
                .limit(limit)
                .all())

        out = []
        for ticket_id, unread, message, created_at, author_id in rows:
            author_name = 'Оператор'
            try:
                au = db.session.get(User, author_id)
                if au:
                    author_name = (au.username or 'Оператор')
            except Exception:
                pass

            out.append({
                'ticket_id': int(ticket_id),
                'unread': int(unread or 0),
                'last_message': (message or '')[:120],
                'last_at': created_at.strftime('%d.%m.%Y %H:%M') if created_at else '',
                'last_author': author_name,
            })
        return out
    except Exception:
        return []


@app.route('/api/ui/dropdown-items', methods=['GET'])
@login_required
def api_ui_dropdown_items():
    """Данные для колокольчика: список уведомлений + непрочитанные треды опер-чата."""
    recipient_type, recipient_id = _recipient_key(current_user)

    accessible_ids = _accessible_ticket_ids_for_notifications(current_user)
    qs = (Notification.query
          .filter_by(recipient_type=recipient_type, recipient_id=recipient_id)
          .order_by(Notification.created_at.desc()))

    if accessible_ids is None:
        raw_notifs = qs.limit(200).all()
    else:
        # берём с запасом: глобальные + по доступным тикетам
        if accessible_ids:
            raw_notifs = (qs.filter(
                or_(Notification.ticket_id.is_(None), Notification.ticket_id.in_(list(accessible_ids)))
            ).limit(250).all())
        else:
            raw_notifs = qs.filter(Notification.ticket_id.is_(None)).limit(250).all()

    # финальная страховка (на случай старых строк без ticket_id)
    notifs = _filter_notifications_for_user(current_user, raw_notifs)[:5]

    notif_items = []
    for n in notifs:
        notif_items.append({
            'id': n.id,
            'title': n.title,
            'is_read': bool(n.is_read),
            'created_at': n.created_at.strftime('%d.%m.%Y %H:%M') if n.created_at else ''
        })

    opchat_threads = []
    if isinstance(current_user, User):
        opchat_threads = _opchat_unread_threads_for_user(current_user.id, limit=5)

    return jsonify({'success': True, 'notifications': notif_items, 'opchat_threads': opchat_threads})


@app.route('/api/ui/theme', methods=['POST'])
@login_required
def api_ui_theme():
    """Сохранение темы интерфейса (light/dark).
    Храним в сессии + в профиле пользователя (если есть колонка ui_theme)."""
    data = request.get_json(silent=True) or {}
    theme = (data.get('theme') or '').strip().lower()
    if theme not in ('light', 'dark'):
        return jsonify({'success': False, 'error': 'invalid_theme'}), 400

    session['ui_theme'] = theme
    try:
        # Пользователи имеют колонку ui_theme (через ensure_sqlite_schema)
        if hasattr(current_user, 'ui_theme'):
            current_user.ui_theme = theme
            db.session.commit()
    except Exception:
        db.session.rollback()

    return jsonify({'success': True, 'theme': theme})
@app.errorhandler(RequestEntityTooLarge)
def handle_large_file(e):
    flash('Ошибка: суммарный размер файлов превышает 50 МБ.', 'danger')
    return redirect(url_for('create_ticket'))

@app.route('/api/suggest-templates', methods=['POST'])
@login_required
def api_suggest_templates():
    data = request.get_json()
    message = data.get('message', '')
    suggestions = suggest_templates(message)
    return jsonify({'suggestions': suggestions})

@app.route('/api/get-template')
@login_required
def get_template():
    name = request.args.get('name')
    if not name:
        return jsonify({'body': '', 'files': []})
    
    template = ResponseTemplate.query.filter(
        (ResponseTemplate.title == name) | 
        (ResponseTemplate.title.contains(name))
    ).first()
    
    if template:
        files_list = []
        
        # Случай 1: files — это JSON-массив
        try:
            if template.files:
                import json
                parsed = json.loads(template.files)
                if isinstance(parsed, list):
                    for item in parsed:
                        if isinstance(item, str):
                            files_list.append({'filename': item, 'original_name': item})
                        elif isinstance(item, dict):
                            files_list.append({
                                'filename': item.get('filename', item.get('original_name', '')),
                                'original_name': item.get('original_name', item.get('filename', ''))
                            })
                elif isinstance(parsed, str):
                    files_list.append({'filename': parsed, 'original_name': parsed})
        except:
            # Случай 2: files — это просто строка (имя одного файла)
            if template.files:
                files_list.append({'filename': template.files, 'original_name': template.files})
        
        return jsonify({
            'body': template.body,
            'files': files_list
        })
    
    return jsonify({'body': '', 'files': []})

@app.route('/')
def index():
    """Главная страница.

    Если пользователь не авторизован — открываем форму входа.
    """
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    return redirect(url_for('ticket_list'))

# === helpful ===
@app.route('/api/ticket/<int:ticket_id>/helpfulness', methods=['POST'])
@login_required
def ticket_helpfulness(ticket_id):
    try:
        ticket = SupportTicket.query.get_or_404(ticket_id)
        if ticket.client_id != current_user.id:
            return jsonify({'error': 'Permission denied'}), 403
        if ticket.helpful is not None:
            return jsonify({'error': 'Оценка уже зафиксирована'}), 400
        data = request.get_json()
        if not data or 'helpful' not in data:
            return jsonify({'error': 'Missing "helpful" field'}), 400
        helpful = data['helpful']
        if not isinstance(helpful, bool):
            return jsonify({'error': '"helpful" must be boolean'}), 400
        if helpful:
            ticket.status = "Завершена"
            ticket.closed_at = utcnow()
        else:
            ticket.status = "Ждёт контроля"
        ticket.helpful = helpful
        ticket.helpful_at = utcnow()

        if not helpful:
            eval_time = datetime.now().strftime('%d.%m.%Y %H:%M')
            eval_entry = f"Клиент оценил: ❌ Не помогло — {eval_time}"
            if ticket.response:
                ticket.response = f"{ticket.response}\n{eval_entry}"
            else:
                ticket.response = eval_entry
        db.session.commit()
        print(f"✅ Установлен статус: {ticket.status}, helpful: {ticket.helpful}")
        return jsonify({'status': 'ok'})
    except Exception as e:
        print("❌ Ошибка при сохранении:", str(e))
        db.session.rollback()
        return jsonify({'error': 'Ошибка сохранения'}), 500

@app.route('/api/ticket/<int:ticket_id>/attachments')
@login_required
def get_ticket_attachments(ticket_id):
    """Возвращает ВСЕ файлы заявки: из заявки + из комментариев"""
    ticket = SupportTicket.query.get_or_404(ticket_id)
    
    # Проверка доступа
    if getattr(current_user, 'role', None) == 'client':
        if ticket.client_id != current_user.id:
            return jsonify({"error": "Доступ запрещён"}), 403
    
    attachments = []
    
    # === ФАЙЛЫ ИЗ ОСНОВНОЙ ЗАЯВКИ (ticket.files) ===
    if ticket.files:
        try:
            file_list = json.loads(ticket.files)
            print(f"DEBUG: Найдено файлов в заявке: {len(file_list)}")
            print(f"DEBUG: Файлы: {file_list}")
            
            if isinstance(file_list, list):
                for filename in file_list:
                    print(f"DEBUG: Обработка файла: {filename}")
                    
                    # Ищем файл в разных возможных местах
                    file_found = False
                    file_path = None
                    url = None
                    
                    # Вариант 1: в папке статики uploads
                    file_path_1 = os.path.join('static', 'uploads', filename)
                    if os.path.exists(file_path_1):
                        file_path = file_path_1
                        url = url_for('static', filename=f'uploads/{filename}')
                        file_found = True
                        print(f"DEBUG: Файл найден в static/uploads: {file_path}")
                    
                    # Вариант 2: в корневой папке uploads
                    if not file_found:
                        file_path_2 = os.path.join('uploads', filename)
                        if os.path.exists(file_path_2):
                            file_path = file_path_2
                            url = f"/uploads/{filename}"
                            file_found = True
                            print(f"DEBUG: Файл найден в uploads: {file_path}")
                    
                    # Вариант 3: в папке редактора
                    if not file_found and 'editor_images' in filename:
                        file_path_3 = os.path.join('static', 'uploads', 'editor_images', filename)
                        if os.path.exists(file_path_3):
                            file_path = file_path_3
                            url = url_for('static', filename=f'uploads/editor_images/{filename}')
                            file_found = True
                            print(f"DEBUG: Файл найден в editor_images: {file_path}")
                    
                    if file_found and file_path:
                        ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
                        is_image = ext in {'png', 'jpg', 'jpeg', 'gif', 'webp'}
                        
                        file_size = os.path.getsize(file_path)
                        print(f"DEBUG: Тип файла: {'изображение' if is_image else 'документ'}, Размер: {file_size} bytes")
                        
                        attachments.append({
                            'original_name': filename,
                            'saved_name': filename,
                            'url': url,
                            'type': 'image' if is_image else 'document',
                            'size': file_size,
                            'size_formatted': format_file_size(file_size),
                            'source': 'ticket'
                        })
                    else:
                        print(f"DEBUG: ⚠️ Файл НЕ найден: {filename}")
        except Exception as e:
            print(f"DEBUG: ❌ Ошибка загрузки файлов заявки: {e}")
            import traceback
            traceback.print_exc()
    
    # === ФАЙЛЫ ИЗ КОММЕНТАРИЕВ ===
    messages = TicketMessage.query.filter_by(ticket_id=ticket_id).all()
    print(f"\nDEBUG: Найдено комментариев: {len(messages)}")
    
    for msg in messages:
        print(f"DEBUG: Комментарий ID {msg.id}: {len(msg.file_attachments)} вложений")
        for att in msg.file_attachments:
            print(f"DEBUG:   - {att.original_name} (ID: {att.id})")
            ext = att.original_name.rsplit('.', 1)[1].lower() if '.' in att.original_name else ''
            is_image = ext in {'png', 'jpg', 'jpeg', 'gif', 'webp'}
            
            attachments.append({
                'original_name': att.original_name,
                'saved_name': att.filename,
                'url': att.url or url_for('static', filename=f'uploads/attachments/{att.filename}'),
                'type': 'image' if is_image else 'document',
                'size': att.size,
                'size_formatted': format_file_size(att.size),
                'source': 'comment',
                'comment_id': msg.id
            })
    
    print(f"\nDEBUG: === ИТОГО НАЙДЕНО ФАЙЛОВ: {len(attachments)} ===")
    for idx, att in enumerate(attachments):
        print(f"DEBUG:   {idx + 1}. {att['original_name']} - {att['type']} ({att['source']})")
    
    return jsonify({
        'success': True,
        'files': attachments,
        'count': len(attachments)
    })
# === ИСТОРИЯ ===
@app.route('/api/ticket/<int:ticket_id>/history')
@login_required
def api_ticket_history(ticket_id):
    if getattr(current_user, 'role', None) == 'client':
        # Клиентам показываем только публичные события
        history = TicketHistory.query.filter_by(
            ticket_id=ticket_id
        ).filter(
            TicketHistory.field.in_(['status', 'response', 'client_feedback', 'priority', 'department']) 
        ).order_by(TicketHistory.timestamp.desc()).all()
    else:
        # Операторам показываем все
        history = TicketHistory.query.filter_by(
            ticket_id=ticket_id
        ).order_by(TicketHistory.timestamp.desc()).all()
    
    result = []
    for h in history:
        # Определяем имя пользователя
        username = 'Система'
        if h.user_id > 0:
            user = db.session.get(User, h.user_id)
            if user:
                full = f"{user.last_name or ''} {user.name or ''} {user.patronymic or ''}".strip()
                username = full or (user.username or user.email or 'Пользователь')
        
        # Локализуем значения полей
        old_value_display = h.old_value
        new_value_display = h.new_value
        
        # Значения по умолчанию
        description = ""
        icon = "bi-info-circle"
        color = "text-muted"
        field_display = translate_field_name(h.field)
        
        if h.field == 'priority':
            old_value_display = translate_priority(h.old_value)
            new_value_display = translate_priority(h.new_value)
            field_display = 'Приоритет'
            icon = "bi-flag"
            color = "history-priority"
            description = f"Приоритет: {old_value_display} → {new_value_display}"
            
        elif h.field == 'status':
            field_display = 'Статус'
            icon = "bi-arrow-right"
            color = "history-status"
            
            # Локализуем значения статусов
            old_status = translate_status(h.old_value) if h.old_value else '—'
            new_status = translate_status(h.new_value) if h.new_value else '—'
            
            # Особые случаи для статусов
            if h.new_value == 'Ожидает подтверждения клиента':
                description = f"📤 Отправлено на проверку клиенту"
                icon = "bi-send"
            elif h.new_value == 'Завершена' and h.old_value == 'Ожидает подтверждения клиента':
                if h.note and 'Авто-закрытие' in str(h.note):
                    description = f"⏰ Автоматическое закрытие (клиент не ответил за 24ч)"
                    icon = "bi-clock"
                    color = "text-secondary"
                else:
                    description = f"✅ Завершено"
                    icon = "bi-check-circle"
                    color = "history-feedback"
            elif h.new_value == 'В работе' and h.old_value == 'Ожидает подтверждения клиента':
                description = f"🔧 Возвращено на доработку"
                icon = "bi-arrow-repeat"
                color = "history-back"
            else:
                description = f"Статус: {old_status} → {new_status}"
                
        elif h.field == 'department':
            field_display = 'Отдел'
            icon = "bi-building"
            color = "history-department"
            description = f"Отдел: {h.old_value or '—'} → {h.new_value or '—'}"
            
        elif h.field == 'client_feedback':
            field_display = 'Обратная связь'
            if h.new_value == 'принял решение':
                description = f"✅ Клиент подтвердил: проблема решена"
                icon = "bi-check-circle"
                color = "history-feedback"
            elif h.new_value == 'запросил доработку':
                description = f"🔧 Клиент запросил: требуется обработка"
                icon = "bi-arrow-repeat"
                color = "history-feedback"
            elif h.new_value == 'нет ответа':
                description = f"⏰ Клиент не предоставил обратную связь"
                icon = "bi-clock"
                color = "history-system"
            elif h.new_value == 'ожидает':
                description = f"⏳ Ожидание подтверждения от клиента"
                icon = "bi-hourglass"
                color = "history-system"
            else:
                description = f"Обратная связь: {h.old_value or '—'} → {h.new_value or '—'}"
                icon = "bi-chat"
                color = "text-info"
                
        elif h.field == 'subject':
            field_display = 'Тема'
            icon = "bi-card-heading"
            color = "history-system"
            # Обрезаем длинные темы
            old_short = h.old_value[:30] + '...' if h.old_value and len(h.old_value) > 30 else h.old_value
            new_short = h.new_value[:30] + '...' if h.new_value and len(h.new_value) > 30 else h.new_value
            description = f"Тема изменена: {old_short or '—'} → {new_short or '—'}"
            
        elif h.field == 'system_note':
            field_display = 'Система'
            icon = "bi-gear"
            color = "text-secondary"
            description = h.new_value or 'Системное событие'
            
        elif h.field == 'response':
            field_display = 'Ответ оператора'
            icon = "bi-chat-text"
            color = "text-info"
            if h.old_value and h.new_value:
                description = f"Ответ оператора обновлён"
            elif h.new_value:
                description = f"📝 Оператор добавил ответ"
            else:
                description = f"Ответ оператора: {h.old_value or '—'} → {h.new_value or '—'}"
                
        elif h.field == 'message':
            field_display = 'Описание'
            icon = "bi-file-text"
            color = "text-primary"
            description = f"Описание заявки обновлено"
            
        else:
            # Для всех остальных полей
            description = f"{field_display}: {h.old_value or '—'} → {h.new_value or '—'}"
        
        # Если description все еще пустой (на всякий случай)
        if not description:
            description = f"{field_display}: {h.old_value or '—'} → {h.new_value or '—'}"
        
        result.append({
            'timestamp': h.timestamp.isoformat(),
            'username': username,
            'field': h.field,
            'field_display': field_display,
            'old_value': h.old_value,
            'new_value': h.new_value,
            'old_value_display': old_value_display,
            'new_value_display': new_value_display,
            'note': h.note,
            'description': description,
            'icon': icon,
            'color': color
        })
    
    return jsonify(result)
# Форма комментариев
messages_store = {}  # { ticket_id: [message1, message2, ...] }

def get_current_operator_name():
    return "Оператор Тест"

## breadcrumbs больше не используются (удалено)

@app.route('/uploads/attachments/<filename>')
@login_required
def download_attachment(filename):
    return send_from_directory(
        os.path.join(app.config['UPLOAD_FOLDER'], 'attachments'),
        filename,
        as_attachment=True
    )

# === АУТЕНТИФИКАЦИЯ ===

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('ticket_list'))

    if request.method == 'POST':
        identifier = request.form['username'].strip()  # поле в форме всё ещё "username"
        password = request.form['password'].strip()

        user = None

        ident = identifier.strip()
        # Сначала пробуем по email
        if '@' in ident and '.' in ident:
            candidate = User.query.filter(db.func.lower(User.email) == ident.lower()).first()
        else:
            candidate = None

        # Затем по username
        if not candidate:
            candidate = User.query.filter_by(username=ident).first()

        if candidate and check_password_hash(candidate.password, password):
            if not candidate.is_active:
                flash("Аккаунт заблокирован", "error")
            elif (candidate.is_client() and not candidate.email_verified):
                login_user(candidate)
                flash("Подтвердите email", "warning")
                return redirect(url_for('unverified'))
            else:
                user = candidate

        if user:
            login_user(user)
            return redirect(url_for('ticket_list'))
        else:
            flash("Неверный email/логин или пароль", "error")

    return render_template('login.html')
@app.before_request
def restrict_unverified_users():
    if not current_user.is_authenticated:
        return
    if hasattr(current_user, 'is_operator') and current_user.is_operator():
        return
    if hasattr(current_user, 'email_verified') and not current_user.email_verified:
        allowed_endpoints = {
            'unverified',
            'resend_verification',
            'confirm_email',
            'logout',
            'static'
        }
        print(f"🔍 endpoint = {request.endpoint}")
        print(f"🔍 allowed = {allowed_endpoints}")
        if request.endpoint not in allowed_endpoints:
            print(f"⚠️ {current_user.email} заблокирован от {request.endpoint}")
            return redirect(url_for('unverified'))

    # Обязательность заполнения профиля для клиентов (тонкая настройка)
    try:
        mode = get_profile_enforcement_mode()
        if mode == 'off':
            return

        missing = []
        if not getattr(current_user, 'email', None):
            missing.append('Почта')
        if not getattr(current_user, 'inn', None):
            missing.append('ИНН')
        if not getattr(current_user, 'last_name', None):
            missing.append('Фамилия')
        if not getattr(current_user, 'name', None):
            missing.append('Имя')
        # Отчество можно требовать — ты просил ФИО, значит требуем
        if not getattr(current_user, 'patronymic', None):
            missing.append('Отчество')
        if not getattr(current_user, 'address', None):
            missing.append('Адрес')

        if missing:
            # Разрешённые точки всегда (чтобы можно было заполнить профиль)
            always_allowed = {
                'user_profile',
                'logout',
                'static',
                'bitrix_company_by_email',
                'unverified',
                'resend_verification',
                'confirm_email',
            }

            endpoint = request.endpoint or ''
            if endpoint.startswith('static'):
                return
            if endpoint in always_allowed:
                return

            if mode == 'soft':
                # В soft-режиме разрешаем только просмотр (GET) базовых страниц,
                # но запрещаем любые действия (POST) и создание заявок.
                view_allowed = {
                    'ticket_list',
                    'ticket_detail',
                    'user_knowledge_base',
                }
                if request.method == 'GET' and endpoint in view_allowed:
                    return

            # strict (и всё остальное) — отправляем в профиль
            flash_msg('Заполни профиль: ' + ', '.join(missing), 'warning')
            return redirect(url_for('user_profile'))
    except Exception:
        pass


# Страница Unverified

@app.route('/unverified')
@login_required
def unverified():
    if getattr(current_user, 'role', None) == 'client' and current_user.email_verified:
        return redirect(url_for('ticket_list'))
    return render_template('unverified.html')

# Кнопка Отправить повторно письмо подтверждения

@app.route('/resend-verification')
@login_required
def resend_verification():
    if getattr(current_user, 'role', None) == 'client' and not current_user.email_verified:
        send_email_verification(current_user)
        flash("Письмо с подтверждением отправлено повторно. Проверьте почту (включая папку «Спам»).", "info")
    return redirect(url_for('unverified'))

# === АДМИНКА ===

@app.route('/admin')
@login_required
def admin():
    # Админка доступна только пользователям с ролью admin
    if getattr(current_user, 'role', None) != 'admin':
        flash("Доступ запрещён", "error")
        # Операторам — максимум канбан (если разрешён), конечным пользователям — их заявки
        if isinstance(current_user, User):
            return redirect(url_for('kanban'))
        return redirect(url_for('ticket_list'))

    new_tickets = SupportTicket.query.filter_by(status='Новая').order_by(SupportTicket.created_at.desc()).all()
    total_tickets = SupportTicket.query.count()
    resolved_count = SupportTicket.query.filter_by(status='Завершена').count()
    operators_count = User.query.filter(User.role.in_(['admin', 'operator'])).count()

    return render_template(
        'admin_dashboard.html',
        new_tickets=new_tickets,
        total_tickets=total_tickets,
        resolved_count=resolved_count,
        operators_count=operators_count
    )

# --- endpoint aliases for backward-compatible templates/links ---
# Some templates historically used url_for('admin_dashboard').
try:
    app.add_url_rule('/admin', endpoint='admin_dashboard', view_func=admin)
except Exception:
    pass


@app.route('/admin/close-reasons')
@login_required
def admin_close_reasons():
    if getattr(current_user, 'role', None) == 'client' or getattr(current_user, 'role', None) != 'admin':
        flash('Доступ запрещён', 'error')
        return redirect(url_for('ticket_list'))

    reasons = TicketCloseReason.query.order_by(TicketCloseReason.sort_order, TicketCloseReason.name).all()
    return render_template('admin_close_reasons.html', reasons=reasons)


@app.route('/admin/close-reasons/new', methods=['GET', 'POST'])
@login_required
def admin_close_reason_new():
    if getattr(current_user, 'role', None) == 'client' or getattr(current_user, 'role', None) != 'admin':
        flash('Доступ запрещён', 'error')
        return redirect(url_for('ticket_list'))

    if request.method == 'POST':
        code = (request.form.get('code') or '').strip()
        name = (request.form.get('name') or '').strip()
        sort_order = request.form.get('sort_order', '0').strip()
        is_active = request.form.get('is_active') == '1'
        require_comment = request.form.get('require_comment') == '1'

        if not code or not name:
            flash('Заполните code и name', 'error')
        else:
            try:
                so = int(sort_order) if str(sort_order).lstrip('-').isdigit() else 0
                r = TicketCloseReason(code=code, name=name, sort_order=so, is_active=is_active, require_comment=require_comment)
                db.session.add(r)
                db.session.commit()
                flash('Причина закрытия добавлена', 'success')
                return redirect(url_for('admin_close_reasons'))
            except Exception as e:
                db.session.rollback()
                flash(f'Ошибка: {e}', 'error')

    return render_template('admin_close_reason_form.html', reason=None)


@app.route('/admin/close-reasons/<int:reason_id>/edit', methods=['GET', 'POST'])
@login_required
def admin_close_reason_edit(reason_id):
    if getattr(current_user, 'role', None) == 'client' or getattr(current_user, 'role', None) != 'admin':
        flash('Доступ запрещён', 'error')
        return redirect(url_for('ticket_list'))

    reason = TicketCloseReason.query.get_or_404(reason_id)
    if request.method == 'POST':
        reason.code = (request.form.get('code') or '').strip()
        reason.name = (request.form.get('name') or '').strip()
        sort_order = request.form.get('sort_order', '0').strip()
        reason.sort_order = int(sort_order) if str(sort_order).lstrip('-').isdigit() else 0
        reason.is_active = request.form.get('is_active') == '1'
        reason.require_comment = request.form.get('require_comment') == '1'
        try:
            db.session.commit()
            flash('Сохранено', 'success')
            return redirect(url_for('admin_close_reasons'))
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка: {e}', 'error')

    return render_template('admin_close_reason_form.html', reason=reason)


@app.route('/admin/close-reasons/<int:reason_id>/delete', methods=['POST'])
@login_required
def admin_close_reason_delete(reason_id):
    if getattr(current_user, 'role', None) == 'client' or getattr(current_user, 'role', None) != 'admin':
        flash('Доступ запрещён', 'error')
        return redirect(url_for('ticket_list'))

    reason = TicketCloseReason.query.get_or_404(reason_id)
    try:
        # мягкое удаление
        reason.is_active = False
        db.session.commit()
        flash('Причина выключена', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'error')
    return redirect(url_for('admin_close_reasons'))
# Дополнительные данные для компактных графиков
@app.route('/api/stats/compact')
@login_required
def api_stats_compact():
    if getattr(current_user, 'role', None) == 'client' or current_user.role != 'admin':
        return jsonify({}), 403
    
    # Рассчитываем показатели как на скриншоте
    total_tickets = SupportTicket.query.count()
    resolved_tickets = SupportTicket.query.filter_by(is_resolved=True).count()
    
    # Среднее время ответа (можно рассчитать реальное)
    avg_response_hours = 8
    avg_response_minutes = 4
    
    # Индекс удовлетворенности (из поля helpful)
    helpful_tickets = SupportTicket.query.filter(SupportTicket.helpful.isnot(None)).count()
    if helpful_tickets > 0:
        positive = SupportTicket.query.filter_by(helpful=True).count()
        satisfaction = round((positive / helpful_tickets) * 100)
    else:
        satisfaction = 71  # по умолчанию
    
    # Просроченные
    overdue = SupportTicket.query.filter(
        SupportTicket.is_resolved == False,
        SupportTicket.sla_deadline < utcnow()
    ).count()
    
    return jsonify({
        'avg_response': f'{avg_response_hours} ч. {avg_response_minutes} м.',
        'satisfaction': f'{satisfaction}%',
        'total_tickets': total_tickets,
        'resolved_tickets': resolved_tickets,
        'overdue_tickets': overdue,
        'positive_reviews': '42%',  # можно рассчитать реальные
        'negative_reviews': '18%',
        'no_reviews': '40%'
    })

@app.route('/api/dashboard/data')
@login_required
def api_dashboard_data():
    if getattr(current_user, 'role', None) == 'client' or current_user.role != 'admin':
        return jsonify({}), 403
    
    # 1. Заявки по месяцам (последние 6 месяцев)
    today = utcnow()
    months_data = []
    month_labels = []
    
    for i in range(5, -1, -1):
        month_start = today.replace(day=1) - timedelta(days=30*i)
        month_end = month_start + timedelta(days=32)
        month_end = month_end.replace(day=1) - timedelta(days=1)
        
        count = SupportTicket.query.filter(
            SupportTicket.created_at >= month_start,
            SupportTicket.created_at <= month_end
        ).count()
        
        months_data.append(count)
        month_labels.append(month_start.strftime('%b %Y'))
    
    # 2. Статусы заявок
    status_counts = {status: 0 for status in STATUSES}
    for ticket in SupportTicket.query.all():
        status_counts[ticket.status] = status_counts.get(ticket.status, 0) + 1
    
    # 3. Доля решенных/нерешенных
    total_tickets = SupportTicket.query.count()
    resolved_tickets = SupportTicket.query.filter_by(is_resolved=True).count()
    unresolved_tickets = total_tickets - resolved_tickets
    
    # 4. SLA
    overdue = SupportTicket.query.filter(
        SupportTicket.is_resolved == False,
        SupportTicket.sla_deadline < utcnow()
    ).count()
    
    # 5. Распределение по отделам (топ-8)
    dept_counts = {}
    for ticket in SupportTicket.query.all():
        dept_name = ticket.department.name if ticket.department else 'Без отдела'
        dept_counts[dept_name] = dept_counts.get(dept_name, 0) + 1
    
    # Сортируем и берем топ-8
    sorted_depts = sorted(dept_counts.items(), key=lambda x: x[1], reverse=True)[:8]
    dept_labels = [dept[0] for dept in sorted_depts]
    dept_values = [dept[1] for dept in sorted_depts]
    
    return jsonify({
        'tickets_by_month': {'labels': month_labels, 'data': months_data},
        'status_distribution': {
            'labels': list(status_counts.keys()),
            'data': list(status_counts.values())
        },
        'resolution_rate': {
            'labels': ['Решены', 'В работе'],
            'data': [resolved_tickets, unresolved_tickets]
        },
        'sla_compliance': {
            'on_time_percent': round((total_tickets - overdue) / total_tickets * 100, 1) if total_tickets else 0,
            'overdue_percent': round(overdue / total_tickets * 100, 1) if total_tickets else 0
        },
        'department_distribution': {
            'labels': dept_labels,
            'data': dept_values
        }
    })
@app.route('/api/ticket/<int:ticket_id>/messages')
@login_required
def get_ticket_messages(ticket_id):
    ticket = SupportTicket.query.get_or_404(ticket_id)
    
    if getattr(current_user, 'role', None) == 'client':
        if ticket.client_id != current_user.id:
            return jsonify({"error": "Доступ запрещён"}), 403
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    messages_query = TicketMessage.query.filter_by(ticket_id=ticket_id)\
        .order_by(TicketMessage.created_at.desc())
    
    pagination = messages_query.paginate(page=page, per_page=per_page, error_out=False)
    messages = pagination.items
    
    comments = []
    for msg in messages:
        # Определяем отображаемое имя
        if msg.is_operator:
            user = db.session.get(User, msg.user_id)
            display_name = user.username if user else "Оператор"
        else:
            user = db.session.get(User, msg.user_id)
            if user:
                display_name = user.email or user.username or (f"{user.last_name or ''} {user.name or ''}".strip()) or "Клиент"
            else:
                display_name = "Клиент"
        
        # Получаем вложения
        attachments = []
        for att in msg.attachments:
            attachments.append({
                'id': att.id,
                'filename': att.filename,
                'original_name': att.original_name,
                'size': att.size,
                'url': att.url or f"/static/uploads/attachments/{att.filename}",
                'created_at': att.created_at.isoformat() if att.created_at else None
            })
        
        comments.append({
            'id': msg.id,
            'message': msg.message,
            'is_operator': msg.is_operator,
            'display_name': display_name,
            'created_at': msg.created_at.isoformat() if msg.created_at else None,
            'likes': 0,  # Добавьте логику лайков если нужно
            'liked': False,
            'attachments': attachments
        })
    
    return jsonify({
        'comments': comments[::-1],  # В обратном порядке для хронологического отображения
        'total': pagination.total,
        'page': page,
        'per_page': per_page,
        'has_next': pagination.has_next
    })

import os
from services.calendar_service import fetch_year as fetch_production_calendar_year
from utils.timezone_helper import to_local as helper_to_local, format_local as helper_format_local
from werkzeug.utils import secure_filename
import uuid

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'zip', 'rar'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# === API ===
# API для лайков
@app.route('/api/ticket/<int:ticket_id>/comment/<int:comment_id>/like', methods=['POST', 'DELETE'])
@login_required
def toggle_comment_like(ticket_id, comment_id):
    ticket = SupportTicket.query.get_or_404(ticket_id)
    
    # Проверка доступа
    if getattr(current_user, 'role', None) == 'client':
        if ticket.client_id != current_user.id:
            return jsonify({"error": "Доступ запрещён"}), 403
    
    comment = TicketMessage.query.get_or_404(comment_id)
    
    if request.method == 'POST':
        # Проверяем, не лайкнул ли уже пользователь
        existing_like = CommentLike.query.filter_by(
            comment_id=comment_id,
            user_id=current_user.id
        ).first()
        
        if not existing_like:
            new_like = CommentLike(
                comment_id=comment_id,
                user_id=current_user.id
            )
            db.session.add(new_like)
            db.session.commit()
        
        return jsonify({"success": True, "liked": True})
    
    elif request.method == 'DELETE':
        # Удаляем лайк
        like = CommentLike.query.filter_by(
            comment_id=comment_id,
            user_id=current_user.id
        ).first()
        
        if like:
            db.session.delete(like)
            db.session.commit()
        
        return jsonify({"success": True, "liked": False})

# API для закрепления комментария
@app.route('/api/ticket/<int:ticket_id>/comment/<int:comment_id>/pin-as-result', methods=['POST'])
@login_required
def pin_comment_result(ticket_id, comment_id):
    """Закрепить комментарий как результат заявки"""
    try:
        ticket = SupportTicket.query.get_or_404(ticket_id)
        message = TicketMessage.query.get_or_404(comment_id)
        
        # Проверка прав
        if not current_user.is_operator:
            return jsonify({'error': 'Доступ запрещен'}), 403
        
        # Проверка, что комментарий принадлежит этой заявке
        if message.ticket_id != ticket.id:
            return jsonify({'error': 'Комментарий не принадлежит этой заявке'}), 400
        
        # Проверка, что комментарий от оператора
        if not message.is_operator:
            return jsonify({'error': 'Можно закреплять только комментарии операторов'}), 400
        
        # Закрепляем комментарий как результат
        ticket.pinned_result_id = comment_id
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/ticket/<int:ticket_id>/unpin-result', methods=['POST'])
@login_required
def unpin_ticket_result(ticket_id):
    """Открепить результат от заявки"""
    try:
        ticket = SupportTicket.query.get_or_404(ticket_id)
        
        # Проверка прав
        if not current_user.is_operator:
            return jsonify({'error': 'Доступ запрещен'}), 403
        
        if not ticket.pinned_result_id:
            return jsonify({'error': 'Нет закрепленного результата'}), 400
        
        # Открепляем результат
        ticket.pinned_result_id = None
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
# API для загрузки файлов
@app.route('/api/ticket/<int:ticket_id>/comment/upload', methods=['POST'])
@login_required
def upload_comment_attachment(ticket_id):
    ticket = SupportTicket.query.get_or_404(ticket_id)
    
    # Проверка доступа
    if getattr(current_user, 'role', None) == 'client':
        if ticket.client_id != current_user.id:
            return jsonify({"error": "Доступ запрещён"}), 403
    
    # Проверка количества файлов
    if 'files[]' not in request.files and 'file' not in request.files:
        return jsonify({"error": "Файлы не найдены"}), 400
    
    files = request.files.getlist('files[]') if 'files[]' in request.files else [request.files['file']]
    
    if len(files) > app.config['MAX_FILES_PER_UPLOAD']:
        return jsonify({
            'error': f'Максимум {app.config["MAX_FILES_PER_UPLOAD"]} файлов за раз'
        }), 400
    
    uploaded_files = []
    total_size = 0
    
    for file in files:
        if not file or not file.filename.strip():
            continue
        
        # Проверка расширения
        if not allowed_file(file.filename):
            return jsonify({
                'error': f'Файл {file.filename} имеет недопустимый формат'
            }), 400
        
        # Проверка размера
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > 5 * 1024 * 1024:
            return jsonify({
                'error': f'Файл {file.filename} превышает 5 МБ'
            }), 400
        
        total_size += file_size
        if total_size > app.config['MAX_CONTENT_LENGTH']:
            return jsonify({'error': 'Общий размер превышает лимит'}), 400
        
        # Сохранение
        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        
        upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'attachments')
        os.makedirs(upload_folder, exist_ok=True)
        
        file_path = os.path.join(upload_folder, unique_filename)
        file.save(file_path)
        
        # Определение типа
        ext = filename.rsplit('.', 1)[1].lower()
        file_type = 'image' if ext in ALLOWED_FILE_EXTENSIONS['image'] else 'document'
        
        uploaded_files.append({
            'original_name': filename,
            'saved_name': unique_filename,
            'url': url_for('static', filename=f'uploads/attachments/{unique_filename}'),
            'type': file_type,
            'size': file_size,
            'size_formatted': format_file_size(file_size)
        })
    
    return jsonify({
        'success': True,
        'files': uploaded_files
    })

@app.route('/api/tickets')
@login_required
def api_tickets():
    if getattr(current_user, 'role', None) == 'client':
        return jsonify([])

    if current_user.role == 'admin':
        tickets = SupportTicket.query.order_by(SupportTicket.created_at.desc()).all()
    else:
        dep_ids = user_department_ids(current_user)
        if dep_ids:
            tickets = SupportTicket.query.filter(SupportTicket.department_id.in_(dep_ids)).order_by(SupportTicket.created_at.desc()).all()
        else:
            tickets = []

    result = []
    for t in tickets:
        # === SLA logic ===
        sla_class = ""
        sla_info = ""
        if not t.is_resolved and t.sla_deadline:
            now = utcnow()
            if now > t.sla_deadline:
                sla_class = "sla-red"
                sla_info = "ПРОСРОЧЕНО!"
            else:
                total = (t.sla_deadline - t.created_at).total_seconds()
                remaining = (t.sla_deadline - now).total_seconds()
                if remaining < total * 0.25:
                    sla_class = "sla-yellow"
                    hrs = int(remaining // 3600)
                    mins = int((remaining % 3600) // 60)
                    sla_info = f"Осталось: ~{hrs} ч {mins} мин"
                else:
                    sla_class = "sla-green"
                    hrs = int(remaining // 3600)
                    mins = int((remaining % 3600) // 60)
                    sla_info = f"До дедлайна: ~{hrs} ч {mins} мин"

        # === Формируем ответ ===
        result.append({
            'id': t.id,
            'name': t.name,
            'email': t.email,
            'subject': t.subject,
            'message': t.message,
            'created_at': t.created_at.isoformat(),
            'department': t.department.name if t.department else t.department,
            'status': t.status,
            'is_resolved': t.is_resolved,
            'response': t.response,
            'helpful': t.helpful,
            'bitrix_task_url': f"https://corportal.ciktrb.ru/workgroups/group/82/tasks/task/view/{t.bitrix_task_id}/" if t.bitrix_task_id else None,
            'files': t.files,
            'internal_comment': t.internal_comment,
            'locked_by': t.locked_by,
            'locked_by_name': t.locked_by and (db.session.get(User, t.locked_by).username if db.session.get(User, t.locked_by) else None) if t.locked_by else None,
            'locked_at': t.locked_at.isoformat() if t.locked_at else None,
            'sla_class': sla_class,
            'sla_info': sla_info,
            'organization': t.organization,
            'inn': t.inn
        })
        
    return jsonify(result)

# === БЛОКИРОВКИ И ЗАГРУЗКИ ===
@app.route('/lock_ticket', methods=['POST'])
@login_required
def lock_ticket():
    if getattr(current_user, 'role', None) == 'client':
        return jsonify({'success': False}), 403
    data = request.get_json()
    ticket_id = data.get('ticket_id')
    ticket = SupportTicket.query.get(ticket_id)
    if ticket:
        ticket.locked_by = current_user.id
        ticket.locked_at = utcnow()
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False})

@app.route('/unlock_ticket', methods=['POST'])
@login_required
def unlock_ticket():
    if getattr(current_user, 'role', None) == 'client':
        return jsonify({'success': False}), 403
    data = request.get_json()
    ticket_id = data.get('ticket_id')
    ticket = SupportTicket.query.get(ticket_id)
    if ticket:
        ticket.locked_by = None
        ticket.locked_at = None
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False})

@app.route('/create_bitrix_task/<int:ticket_id>', methods=['POST'])
@login_required
def create_bitrix_task_api(ticket_id):
    if getattr(current_user, 'role', None) == 'client':
        return jsonify({'success': False, 'error': 'Доступ запрещён'}), 403
    
    ticket = SupportTicket.query.get_or_404(ticket_id)

    # Если задача уже создана — просто вернём ссылку
    if getattr(ticket, 'bitrix_task_id', None):
        task_id = str(ticket.bitrix_task_id)
        url = f"https://corportal.ciktrb.ru/workgroups/group/82/tasks/task/view/{task_id}/"
        return jsonify({'success': True, 'task_id': task_id, 'url': url, 'label': f'Задача Битрикс №{task_id}'})

    success, payload = create_bitrix_task(ticket)

    if success:
        task_id = str(payload)
        url = f"https://corportal.ciktrb.ru/workgroups/group/82/tasks/task/view/{task_id}/"
        return jsonify({'success': True, 'task_id': task_id, 'url': url, 'label': f'Задача Битрикс №{task_id}'})
    else:
        return jsonify({'success': False, 'error': payload or 'Не удалось создать задачу в Bitrix24'})


@app.route('/ticket/<int:ticket_id>/update_priority', methods=['POST'])
@login_required
def update_ticket_priority(ticket_id: int):
    # Operators/Admins only
    if getattr(current_user, 'role', None) == 'client':
        return jsonify({'success': False, 'error': 'Доступ запрещён'}), 403

    ticket = SupportTicket.query.get_or_404(ticket_id)

    # Accept JSON or form
    data = request.get_json(silent=True) or {}
    priority = (data.get('priority') or request.form.get('priority') or '').strip().lower()

    # Normalized priorities used across UI
    allowed = {'низкий', 'средний', 'высокий', 'low', 'medium', 'high'}
    if priority not in allowed:
        return jsonify({'success': False, 'error': 'Некорректный приоритет'}), 400

    # Store in canonical internal values (low/medium/high)
    mapping = {'низкий': 'low', 'средний': 'medium', 'высокий': 'high',
               'low': 'low', 'medium': 'medium', 'high': 'high'}
    old_priority = getattr(ticket, 'priority', None)
    ticket.priority = mapping[priority]
    db.session.commit()

    # --- In-app уведомления: смена приоритета ---
    try:
        if (old_priority or '') != (ticket.priority or ''):
            # Клиенту
            if getattr(ticket, 'client_id', None):
                cu = db.session.get(User, ticket.client_id)
                if cu:
                    create_inapp_notification(
                        cu,
                        'priority',
                        f"Изменён приоритет заявки #{ticket.id}",
                        f"{old_priority or '—'} → {ticket.priority or '—'}",
                        url_for('ticket_detail', ticket_id=ticket.id),
                        dedupe_key=f"priority:{ticket.id}:{old_priority}->{ticket.priority}:eu"
                    )

            # Операторам с доступом (кроме инициатора)
            for op in _operators_with_access_to_ticket(ticket):
                if isinstance(current_user, User) and op.id == current_user.id:
                    continue
                create_inapp_notification(
                    op,
                    'priority',
                    f"Изменён приоритет заявки #{ticket.id}",
                    f"{old_priority or '—'} → {ticket.priority or '—'}",
                    url_for('ticket_detail', ticket_id=ticket.id),
                    dedupe_key=f"priority:{ticket.id}:{old_priority}->{ticket.priority}:op:{op.id}"
                )
    except Exception:
        pass

    # UI-friendly label (ru)
    try:
        display = translate_priority(ticket.priority)
    except Exception:
        display = ticket.priority

    return jsonify({'success': True, 'priority': ticket.priority, 'priority_display': display})


@app.route('/ticket/<int:ticket_id>/toggle_critical', methods=['POST'])
@login_required
def toggle_ticket_critical(ticket_id: int):
    """Отметка "Важная задача".

    Приоритеты как функционал убраны, оставляем только boolean-пометку.
    Для совместимости используем поле SupportTicket.priority:
      - включено: 'critical'
      - выключено: 'normal'
    Также понимаем старое значение 'Критический'.
    """
    ticket = SupportTicket.query.get_or_404(ticket_id)

    is_client = getattr(current_user, 'role', None) == 'client'
    is_operator = (not is_client) and (
        current_user.role in ['operator', 'admin'] or (isinstance(current_user, User) and is_tp_operator(current_user))
    )
    if not is_operator:
        return jsonify({'success': False, 'message': 'Недостаточно прав'}), 403

    old_value = (ticket.priority or '').strip()
    was_on = old_value in ('critical', 'Критический')
    new_is_on = not was_on
    ticket.priority = 'critical' if new_is_on else 'normal'

    # TicketHistory.user_id = NOT NULL — фиксируем автора изменения.
    uid = current_user.id

    db.session.add(TicketHistory(
        ticket_id=ticket.id,
        user_id=uid,
        field='important_task',
        old_value='Да' if was_on else 'Нет',
        new_value='Да' if new_is_on else 'Нет'
    ))

    db.session.commit()
    # is_critical оставляем в ответе для совместимости фронта
    return jsonify({'success': True, 'is_important': new_is_on, 'is_critical': new_is_on})


@app.post('/tickets/<int:ticket_id>/comments/<int:message_id>/edit')
@login_required
def edit_ticket_comment(ticket_id: int, message_id: int):
    is_client = getattr(current_user, 'role', None) == 'client'
    is_operator = (not is_client) and current_user.role in ['operator', 'admin']
    is_admin = is_operator and current_user.role == 'admin'
    can_manage_spam = can_manage_spam_user(current_user)
    can_manage_spam = can_manage_spam_user(current_user)

    if not is_client and not is_operator:
        return redirect(url_for('login'))

    ticket = SupportTicket.query.get_or_404(ticket_id)

    # access checks
    if is_client:
        if ticket.email != current_user.email:
            flash('Доступ запрещён', 'error')
            return redirect(url_for('ticket_list'))
    else:
        if not is_admin:
            has_access = False
            if isinstance(current_user, User) and is_tp_operator(current_user):
                has_access = True
            elif ticket.assigned_to_id == current_user.id:
                has_access = True
            elif ticket.department_id and ticket.department_id in user_department_ids(current_user):
                has_access = True
            else:
                try:
                    if getattr(ticket, 'shared_departments_rel', None):
                        dep_ids = set(user_department_ids(current_user))
                        for d in ticket.shared_departments_rel:
                            if d and getattr(d, 'id', None) in dep_ids:
                                has_access = True
                                break
                except Exception:
                    pass
            if not has_access:
                flash('Доступ к заявке запрещён', 'error')
                return redirect(url_for('ticket_list'))

    msg = TicketMessage.query.get_or_404(message_id)
    if msg.ticket_id != ticket.id:
        abort(404)

    if not is_admin and msg.user_id != current_user.id:
        flash('Недостаточно прав для редактирования', 'error')
        return redirect(url_for('ticket_detail', ticket_id=ticket.id))

    if ticket.status in ('Завершена','Спам','Дубликат','Ошибочная','Ошибочно'):
        flash('Нельзя редактировать комментарии в закрытой заявке', 'error')
        return redirect(url_for('ticket_detail', ticket_id=ticket.id))

    new_text = (request.form.get('message') or '').strip()
    if not new_text:
        flash('Сообщение не может быть пустым', 'error')
        return redirect(url_for('ticket_detail', ticket_id=ticket.id))

    # Обновляем текст
    msg.message = new_text
    msg.edited_at = utcnow()
    msg.edited_by_id = current_user.id

    # Удаление вложений при редактировании
    delete_ids = request.form.getlist('delete_attachment_ids')
    if delete_ids:
        for raw_id in delete_ids:
            try:
                att_id = int(raw_id)
            except Exception:
                continue
            att = TicketAttachment.query.get(att_id)
            if not att or att.message_id != msg.id:
                continue
            # Пробуем удалить файл с диска (не критично)
            try:
                for base in [os.path.join('static', 'uploads', 'attachments'), os.path.join('uploads', 'attachments')]:
                    fp = os.path.join(base, att.filename)
                    if os.path.exists(fp):
                        os.remove(fp)
            except Exception:
                pass
            db.session.delete(att)

    # SQLite может быть read-only: не падаем 500, даём понятное сообщение
    try:
        db.session.commit()
    except OperationalError:
        db.session.rollback()
        flash('Не удалось сохранить изменения: база данных доступна только для чтения (SQLite readonly). Проверьте права на instance/support.db и папку instance/.', 'error')
        return redirect(url_for('ticket_detail', ticket_id=ticket.id, _anchor=f'comment-{msg.id}'))

    flash('Комментарий обновлён', 'success')
    return redirect(url_for('ticket_detail', ticket_id=ticket.id, _anchor=f'comment-{msg.id}'))


@app.post('/tickets/<int:ticket_id>/comments/<int:message_id>/delete')
@login_required
def delete_ticket_comment(ticket_id: int, message_id: int):
    is_client = getattr(current_user, 'role', None) == 'client'
    is_operator = (not is_client) and current_user.role in ['operator', 'admin']
    is_admin = is_operator and current_user.role == 'admin'
    can_manage_spam = can_manage_spam_user(current_user)
    if not is_admin:
        flash('Недостаточно прав для удаления', 'error')
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    ticket = SupportTicket.query.get_or_404(ticket_id)
    msg = TicketMessage.query.get_or_404(message_id)
    if msg.ticket_id != ticket.id:
        abort(404)

    if ticket.pinned_result_id == msg.id:
        ticket.pinned_result_id = None

    db.session.delete(msg)
    db.session.commit()
    flash('Комментарий удалён', 'success')
    return redirect(url_for('ticket_detail', ticket_id=ticket.id))

def update_ticket_priority(ticket_id):
    try:
        if not (current_user.is_operator or current_user.is_admin):
            return jsonify({'success': False, 'error': 'Недостаточно прав'}), 403
        
        data = request.get_json()
        if not data or 'priority' not in data:
            return jsonify({'success': False, 'error': 'Отсутствует приоритет'}), 400
        
        # ИСПРАВЛЕНО: используем SupportTicket вместо Ticket
        ticket = SupportTicket.query.get_or_404(ticket_id)
        
        # Проверяем и нормализуем приоритет
        new_priority = normalize_priority(data.get('priority'), default="Обычный")
        if new_priority not in PRIORITIES_RU:
            return jsonify({'success': False, 'error': 'Недопустимый приоритет'}), 400

        # Сохраняем старое значение
        old_priority = ticket.priority

        
        # Обновляем
        ticket.priority = new_priority
        
        # История
        history = TicketHistory(
            ticket_id=ticket.id,
            user_id=current_user.id,
            field='priority',
            old_value=old_priority,
            new_value=new_priority
        )
        db.session.add(history)
        
        db.session.commit()

        # --- In-app уведомления: смена приоритета (второй путь обновления) ---
        try:
            if (old_priority or '') != (ticket.priority or ''):
                if getattr(ticket, 'client_id', None):
                    cu = db.session.get(User, ticket.client_id)
                    if cu:
                        create_inapp_notification(
                            cu,
                            'priority',
                            f"Изменён приоритет заявки #{ticket.id}",
                            f"{old_priority or '—'} → {ticket.priority or '—'}",
                            url_for('ticket_detail', ticket_id=ticket.id),
                            dedupe_key=f"priority2:{ticket.id}:{old_priority}->{ticket.priority}:client"
                        )

                for op in _operators_with_access_to_ticket(ticket):
                    if isinstance(current_user, User) and op.id == current_user.id:
                        continue
                    create_inapp_notification(
                        op,
                        'priority',
                        f"Изменён приоритет заявки #{ticket.id}",
                        f"{old_priority or '—'} → {ticket.priority or '—'}",
                        url_for('ticket_detail', ticket_id=ticket.id),
                        dedupe_key=f"priority2:{ticket.id}:{old_priority}->{ticket.priority}:op:{op.id}"
                    )
        except Exception:
            pass
        
        return jsonify({
            'success': True,
            'priority': data['priority']
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Ошибка обновления приоритета: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/ticket/<int:ticket_id>/update_department', methods=['POST'])
@login_required
def update_ticket_department(ticket_id):
    try:
        # Проверяем права
        if not (current_user.is_operator or current_user.is_admin):
            return jsonify({'success': False, 'error': 'Недостаточно прав'}), 403
        
        # Получаем данные
        data = request.get_json()
        if not data or 'department' not in data:
            return jsonify({'success': False, 'error': 'Отсутствует отдел'}), 400
        
        # Находим тикет
        ticket = SupportTicket.query.get_or_404(ticket_id)
        old_department_id = getattr(ticket, 'department_id', None)
        
        # Сохраняем старое значение ПЕРЕД изменением
        old_department_name = ticket.department.name if ticket.department else '—'
        old_status = ticket.status  # Сохраняем старый статус
        
        # Находим новый отдел
        new_department = None
        if data['department'] and data['department'].strip():
            department = Department.query.filter_by(name=data['department'].strip()).first()
            if department:
                new_department = department
            else:
                # Если отдела нет, можно оставить None
                pass
        
        # Обновляем тикет
        ticket.department_id = new_department.id if new_department else None
        
        # ✅ ВАЖНО: Меняем статус на "Принята" при назначении отдела
        if new_department and ticket.status == 'Новая':
            ticket.status = 'Принята'
            print(f"Статус тикета {ticket.id} изменен с '{old_status}' на '{ticket.status}'")
        
        # Добавляем запись в историю для отдела
        history_department = TicketHistory(
            ticket_id=ticket.id,
            user_id=current_user.id,
            field='department',
            old_value=old_department_name,
            new_value=new_department.name if new_department else '—'
        )
        db.session.add(history_department)
        
        # Добавляем запись в историю для статуса (если он изменился)
        if new_department and ticket.status != old_status:
            history_status = TicketHistory(
                ticket_id=ticket.id,
                user_id=current_user.id,
                field='status',
                old_value=old_status,
                new_value=ticket.status
            )
            db.session.add(history_status)
            print(f"Добавлена запись в историю: статус изменен на '{ticket.status}'")
        
        # Сохраняем изменения
        db.session.commit()

        # --- In-app уведомления: заявка появилась/перешла в отдел ---
        try:
            if new_department and (int(old_department_id or 0) != int(new_department.id)):
                recipients = []
                recipients.extend(list(getattr(new_department, 'users', []) or []))
                recipients.extend(list(getattr(new_department, 'operators', []) or []))

                seen = set()
                for op in recipients:
                    if not op or getattr(op, 'id', None) is None:
                        continue
                    if op.id in seen:
                        continue
                    seen.add(op.id)
                    if getattr(op, 'role', None) not in ('operator', 'admin'):
                        continue
                    if isinstance(current_user, User) and op.id == current_user.id:
                        continue
                    # Доп. проверка доступа (на всякий)
                    if not _user_has_access_to_ticket_for_notifications(op, ticket):
                        continue
                    create_inapp_notification(
                        op,
                        'assigned',
                        f"Заявка #{ticket.id} в отделе «{new_department.name}»",
                        ticket.subject or '',
                        url_for('ticket_detail', ticket_id=ticket.id),
                        dedupe_key=f"dept_move:{ticket.id}:{new_department.id}:op:{op.id}"
                    )
        except Exception:
            pass
        
        return jsonify({
            'success': True,
            'department': new_department.name if new_department else '—',
            'status': ticket.status,
            'status_changed': ticket.status != old_status
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Ошибка обновления отдела: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# === СЛУЖЕБНЫЕ ===
@app.route('/debug/clear_locks')
@login_required
def debug_clear_locks():
    tickets = SupportTicket.query.filter(SupportTicket.locked_by.isnot(None)).all()
    for ticket in tickets:
        ticket.locked_by = None
        ticket.locked_at = None
    db.session.commit()
    return "Блокировки очищены"

# === ПОЛЬЗОВАТЕЛИ ===
@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('login'))


# === Уведомления: открыть/прочитать ===


@app.route('/notifications')
@login_required
def notifications_page():
    recipient_type, recipient_id = _recipient_key(current_user)
    page = request.args.get('page', type=int, default=1)
    per_page = 30
    accessible_ids = _accessible_ticket_ids_for_notifications(current_user)

    qs = (Notification.query
          .filter_by(recipient_type=recipient_type, recipient_id=recipient_id)
          .order_by(Notification.created_at.desc()))

    if accessible_ids is None:
        qs_vis = qs
    else:
        if accessible_ids:
            qs_vis = qs.filter(or_(Notification.ticket_id.is_(None), Notification.ticket_id.in_(list(accessible_ids))))
        else:
            qs_vis = qs.filter(Notification.ticket_id.is_(None))

    total = qs_vis.count()
    pages = (total + per_page - 1) // per_page if total else 1
    items = (qs_vis
             .offset((page - 1) * per_page)
             .limit(per_page)
             .all())

    # финальная страховка (на случай старых уведомлений без ticket_id)
    items = _filter_notifications_for_user(current_user, items)
    opchat_threads = []
    if isinstance(current_user, User):
        opchat_threads = _opchat_unread_threads_for_user(current_user.id, limit=20)
    return render_template('notifications.html', items=items, page=page, pages=pages, total=total, opchat_threads=opchat_threads)
@app.route('/n/<int:notification_id>')
@login_required
def open_notification(notification_id):
    try:
        recipient_type, recipient_id = _recipient_key(current_user)
        n = Notification.query.filter_by(
            id=notification_id,
            recipient_type=recipient_type,
            recipient_id=recipient_id
        ).first_or_404()

        # отмечаем прочитанным
        n.is_read = True
        db.session.commit()

        # если уведомление ведёт на тикет — проверяем, что доступ ещё есть
        tid = _notification_ticket_id(n)
        if tid:
            ticket = SupportTicket.query.get(tid)
            if not ticket or not _user_has_access_to_ticket_for_notifications(current_user, ticket):
                flash_msg("Доступ к заявке запрещён — уведомление скрыто из списка", "warning")
                return redirect(url_for('notifications_page'))

        if n.url:
            return redirect(n.url)
    except Exception:
        db.session.rollback()

    # fallback
    return redirect(url_for('ticket_list'))


@app.route('/notifications/mark-all-read', methods=['POST'])
@login_required
def notifications_mark_all_read():
    try:
        recipient_type, recipient_id = _recipient_key(current_user)
        Notification.query.filter_by(recipient_type=recipient_type, recipient_id=recipient_id, is_read=False).update({'is_read': True})
        db.session.commit()
        flash_msg("Уведомления отмечены как прочитанные", "success")
    except Exception:
        db.session.rollback()
        flash_msg("Не удалось отметить уведомления", "error")
    # возвращаем туда, откуда пришли
    return redirect(request.referrer or url_for('ticket_list'))


# --- Безопасный алиас endpoint'а ---
# Иногда при переносах/частичных обновлениях шаблоны могут ссылаться на endpoint по имени,
# а Flask по каким-то причинам его не видит (например, если приложение запускают не тем модулем).
# Жёстко регистрируем endpoint "notifications_mark_all_read" на тот же handler.
try:
    if 'notifications_mark_all_read' not in app.view_functions:
        app.add_url_rule(
            '/notifications/mark-all-read',
            endpoint='notifications_mark_all_read',
            view_func=notifications_mark_all_read,
            methods=['POST']
        )
except Exception as e:
    print('Alias registration error:', e)

@app.route('/api/tickets/user')
@login_required
def api_user_tickets():
    if isinstance(current_user, User):
        return jsonify([])
    tickets = SupportTicket.query.filter_by(user_id=current_user.id).order_by(SupportTicket.created_at.desc()).all()
    tickets_list = []
    for t in tickets:
        sla_class = ""
        sla_info = ""
        if not t.is_resolved and t.sla_deadline:
            now = utcnow()
            if now > t.sla_deadline:
                sla_class = "sla-red"
                sla_info = "ПРОСРОЧЕНО!"
            else:
                total = (t.sla_deadline - t.created_at).total_seconds()
                remaining = (t.sla_deadline - now).total_seconds()
                if remaining < total * 0.25:
                    sla_class = "sla-yellow"
                    hrs = int(remaining // 3600)
                    mins = int((remaining % 3600) // 60)
                    sla_info = f"Осталось: ~{hrs} ч {mins} мин"
                else:
                    hrs = int(remaining // 3600)
                    mins = int((remaining % 3600) // 60)
                    sla_info = f"До дедлайна: ~{hrs} ч {mins} мин"
        tickets_list.append({
            'id': t.id,
            'subject': t.subject,
            'message': t.message,
            'created_at': t.created_at.isoformat(),
            'department': t.department,
            'status': t.status,
            'response': t.response,
            'sla_class': sla_class,
            'sla_info': sla_info,
            'organization': t.organization,
            'inn': t.inn
        })
    return jsonify(tickets_list)

@app.route('/ticket/create_manual', methods=['GET', 'POST'])
@login_required
def create_manual_ticket():
    # Устаревший маршрут: используем единую страницу создания заявки
    return redirect(url_for('create_ticket'))


@app.route('/ticket/create', methods=['GET', 'POST'])
@login_required
def create_ticket():
    """Единая страница создания заявки.
    - Клиент создаёт заявку сам (данные организации уже есть, но можно менять).
    - Оператор/админ может создать заявку "по телефону" за клиента (вводит email и подтягивает оргданные из Битрикс).
    """
    # Доступ: все авторизованные. Роль определяет права.
    is_operator = bool(getattr(current_user, 'is_operator', lambda: False)())

    categories = TicketCategory.query.filter_by(is_active=True).order_by(TicketCategory.sort_order, TicketCategory.name).all()

    # Default category for client-created tickets ("Обращение" / code=issue)
    default_issue_category = TicketCategory.query.filter_by(code='issue').first()
    if not default_issue_category:
        default_issue_category = TicketCategory.query.filter(TicketCategory.name.ilike('%обращ%')).first()
    intake_id = get_default_intake_department_id()
    intake_dept = Department.query.get(intake_id) if intake_id else None
    if not intake_dept:
        intake_dept = Department.query.filter_by(name="1ая линия ТП").first()
    intake_department_id = intake_dept.id if intake_dept else (intake_id or 1)

    # Prefill org fields
    org_prefill = {
        'organization': '',
        'inn': '',
        'address': ''
    }
    if not is_operator and getattr(current_user, 'role', None) == 'client':
        org_prefill['organization'] = getattr(current_user, 'organization', '') or ''
        org_prefill['inn'] = getattr(current_user, 'inn', '') or ''
        org_prefill['address'] = getattr(current_user, 'address', '') or ''

    if request.method == 'POST':
        client_name = (request.form.get('client_name') or '').strip()
        client_email = (request.form.get('client_email') or '').strip().lower()
        client_phone = (request.form.get('client_phone') or '').strip()
        subject = (request.form.get('subject') or '').strip()
        message = (request.form.get('message') or '').strip()
        # Приоритеты отключены в системе. Используем внутреннее техническое значение.
        # Важная задача включается отдельной кнопкой и хранится как priority='critical'.
        priority = 'normal'

        organization = (request.form.get('organization') or '').strip()
        inn = (request.form.get('inn') or '').strip()
        address = (request.form.get('address') or '').strip()

        errors = {}

        def _render_form():
            return render_template(
                'create_ticket.html',
                is_operator=is_operator,
                categories=categories,
                default_issue_category=default_issue_category,
                intake_department_id=intake_department_id,
                org_prefill={'organization': organization, 'inn': inn, 'address': address},
                form=request.form,
                errors=errors
            )

        # Категории больше не используем в UI: все заявки по сути "обращения".
        # Оставляем сохранение только в одну дефолтную категорию для совместимости БД.
        category = default_issue_category
        if not category:
            # fallback: любая активная
            category = TicketCategory.query.filter_by(is_active=True).order_by(TicketCategory.sort_order, TicketCategory.id).first()
        if not category:
            flash('Не настроена категория для обращений. Обратитесь к администратору.', 'error')
            return _render_form()

        if not client_email or '@' not in client_email:
            errors['client_email'] = 'Некорректный email'
            flash('Некорректный email', 'error')
            return _render_form()

        # ФИО и телефон не обязательны. Клиент может не заполнять профиль.

        if not subject:
            errors['subject'] = 'Тема заявки обязательна'
            flash('Тема заявки обязательна', 'error')
            return _render_form()
        if not message:
            errors['message'] = 'Описание проблемы обязательно'
            flash('Описание проблемы обязательно', 'error')
            return _render_form()

        # Валидация реквизитов организации делается ниже, после автоподтягивания из профиля/Bitrix.

        # === файлы ===
        saved_files = []
        if 'files' in request.files:
            files = request.files.getlist('files')
            if len(files) > 10:
                flash('Максимум 10 файлов за раз', 'error')
                return _render_form()

            total_size = 0
            allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt'}
            for file in files:
                if not file or not file.filename.strip():
                    continue
                if '.' not in file.filename:
                    flash(f"Файл '{file.filename}' не имеет расширения", 'error')
                    return _render_form()
                ext = file.filename.rsplit('.', 1)[1].lower()
                if ext not in allowed_extensions:
                    flash(f"Недопустимый формат файла: {file.filename}", 'error')
                    return _render_form()
                file.seek(0, os.SEEK_END)
                file_size = file.tell()
                file.seek(0)
                if file_size > 5 * 1024 * 1024:
                    flash(f"Файл '{file.filename}' превышает 5 МБ", 'error')
                    return _render_form()
                total_size += file_size
                if total_size > 50 * 1024 * 1024:
                    flash('Общий размер файлов превышает 50 МБ', 'error')
                    return _render_form()
                filename = secure_filename(file.filename)
                unique_name = f"{uuid.uuid4().hex}_{filename}"
                file_path = os.path.join('static', 'uploads', 'attachments', unique_name)
                os.makedirs(os.path.dirname(file_path), exist_ok=True)
                file.save(file_path)
                saved_files.append(unique_name)

        # user link (единая таблица users)
        client_id = None
        if not is_operator and getattr(current_user, 'role', None) == 'client':
            client_id = current_user.id
        else:
            linked_user = User.query.filter(db.func.lower(User.email) == client_email.lower(), User.role == 'client').first()
            client_id = linked_user.id if linked_user else None
            if linked_user and not client_name:
                fio = " ".join([x for x in [linked_user.last_name, linked_user.name, linked_user.patronymic] if x]).strip()
                client_name = fio or client_email

            # Автоподтягивание данных клиента в тикет (серверный fallback, даже если JS не отработал)
            if linked_user:
                if not client_phone and getattr(linked_user, 'phone', None):
                    client_phone = linked_user.phone
                if not organization and getattr(linked_user, 'organization', None):
                    organization = linked_user.organization
                if not inn and getattr(linked_user, 'inn', None):
                    inn = linked_user.inn
                if not address and getattr(linked_user, 'address', None):
                    address = linked_user.address
            else:
                # Если клиента нет в базе — пробуем подтянуть реквизиты из Битрикс по email
                try:
                    if (not organization and not inn and not address) and client_email:
                        company_data = get_company_from_bitrix(client_email)
                        if company_data:
                            organization = organization or (company_data.get('company') or '').strip()
                            inn = inn or (company_data.get('inn') or '').strip()
                            address = address or (company_data.get('address') or '').strip()
                except Exception:
                    pass


        # === Валидация клиентских реквизитов (анти-ссылки/форматы) ===
        # ФИО (в create_ticket это одно поле): запрещаем ссылки/HTML/стоп-слова
        client_name = _norm_text(client_name)
        if client_name and (_contains_url_like(client_name) or _contains_html_like(client_name) or _contains_bad_words(client_name)):
            errors['client_name'] = 'Недопустимое значение'
            flash('ФИО: недопустимое значение', 'error')
            return _render_form()

        ok, phone_v = normalize_phone(client_phone)
        if not ok:
            errors['client_phone'] = phone_v
            flash(phone_v, 'error')
            return _render_form()
        client_phone = phone_v

        # Данные организации обязательны
        ok, org_v = validate_org(organization)
        if not ok or not _norm_text(org_v):
            errors['organization'] = 'Укажите организацию (без ссылок/HTML)'
        ok, inn_v = validate_inn_ru(inn, required=True)
        if not ok:
            errors['inn'] = inn_v
        ok, addr_v = validate_address(address, required=True)
        if not ok:
            errors['address'] = addr_v
        if errors:
            flash('Проверьте заполнение обязательных полей', 'error')
            return _render_form()
        organization, inn, address = org_v, inn_v, addr_v


        ticket = SupportTicket(
            email=client_email,
            name=client_name or client_email,
            phone=client_phone or None,
            organization=organization or None,
            inn=inn or None,
            address=address or None,
            subject=subject,
            message=message,
            created_at=utcnow(),
            status='Новая',
            department_id=intake_department_id,
            client_id=client_id,
            user_id=None,
            created_by_operator_id=current_user.id if is_operator else None,
            priority=priority,
            ticket_type=category.code if hasattr(category, 'code') else 'issue',
            category_id=category.id
        )

        try:
            db.session.add(ticket)
            db.session.flush()
            recalc_ticket_sla(ticket)
            if saved_files:
                ticket.files = json.dumps(saved_files)
            db.session.commit()

            # === Синхронизация реквизитов с профилем клиента ===
            # 1) Если в профиле клиента реквизитов нет (не подтянулось и не заполнено) —
            #    сохраняем реквизиты из формы создания заявки в профиль.
            # 2) Если в профиле уже есть реквизиты, а в заявке указаны другие —
            #    НЕ перетираем профиль, но покажем уведомление и дадим кнопку "Перенести в профиль".
            org_mismatch = False
            if (not is_operator) and getattr(current_user, 'role', None) == 'client':
                prof_org = (getattr(current_user, 'organization', '') or '').strip()
                prof_inn = (getattr(current_user, 'inn', '') or '').strip()
                prof_addr = (getattr(current_user, 'address', '') or '').strip()

                if not (prof_org or prof_inn or prof_addr):
                    # профиль пустой — заполняем
                    current_user.organization = organization or None
                    current_user.inn = inn or None
                    current_user.address = address or None
                    try:
                        db.session.commit()
                    except Exception:
                        db.session.rollback()
                else:
                    # профиль заполнен — считаем что заявка на другую организацию:
                    # 1) если есть ИНН и у профиля, и у заявки — сравниваем строго ИНН
                    # 2) если ИНН отсутствует хотя бы у одной стороны — сравниваем по названию (нормализованному)
                    if prof_inn and inn:
                        org_mismatch = (prof_inn != inn)
                    else:
                        p = normalize_org_name(prof_org)
                        t = normalize_org_name(organization)
                        if p and t and p != t:
                            org_mismatch = True

            # === In-app уведомления ===
            try:
                # 1) Уведомление клиенту (если есть профиль) — "заявка создана"
                if ticket.client_id:
                    eu = db.session.get(User, ticket.client_id)
                    if eu:
                        create_inapp_notification(
                            eu,
                            'assigned',
                            f"Заявка #{ticket.id} создана",
                            ticket.subject or '',
                            url_for('ticket_detail', ticket_id=ticket.id),
                            dedupe_key=f"created:{ticket.id}:u:{eu.id}"
                        )

                # 2) Уведомление операторам отдела (и админам) — "новая заявка"
                dept = Department.query.get(ticket.department_id) if ticket.department_id else None
                recipients = []
                if dept:
                    # основной отдел + доп. операторы
                    recipients.extend(list(getattr(dept, 'users', []) or []))
                    recipients.extend(list(getattr(dept, 'operators', []) or []))
                # если отдел не найден — уведомим всех админов
                if not recipients:
                    recipients = User.query.filter(User.role.in_(['operator','admin'])).all()

                seen = set()
                for op in recipients:
                    if not op or getattr(op, 'id', None) is None:
                        continue
                    if op.id in seen:
                        continue
                    seen.add(op.id)
                    # не уведомляем создателя-оператора о его же действии
                    if is_operator and isinstance(current_user, User) and op.id == current_user.id:
                        continue
                    if getattr(op, 'role', None) not in ('operator', 'admin'):
                        continue
                    create_inapp_notification(
                        op,
                        'assigned',
                        f"Новая заявка #{ticket.id}",
                        ticket.subject or '',
                        url_for('ticket_detail', ticket_id=ticket.id),
                        dedupe_key=f"created:{ticket.id}:op:{op.id}"
                    )
            except Exception:
                # уведомления не должны ломать создание заявки
                pass

            flash('Заявка создана!', 'success')
            if is_operator:
                return redirect(url_for('admin_kanban'))
            # клиенту удобнее сразу открыть созданную заявку
            if org_mismatch:
                return redirect(url_for('ticket_detail', ticket_id=ticket.id, org_mismatch=1))
            return redirect(url_for('ticket_detail', ticket_id=ticket.id))
        except Exception as e:
            db.session.rollback()
            print(f"[ERROR] Создание заявки: {e}")
            flash('Ошибка при создании заявки. Попробуйте позже.', 'error')
            return _render_form()

    return render_template('create_ticket.html', is_operator=is_operator, categories=categories, default_issue_category=default_issue_category, intake_department_id=intake_department_id, org_prefill=org_prefill, form=None, errors={})
# === АДМИН-СТРАНИЦЫ ===

@app.route('/kanban', endpoint='kanban')
@login_required
def kanban():
    if not isinstance(current_user, User):
        flash("Доступ запрещён", "error")
        return redirect(url_for('ticket_list'))

    is_admin = current_user.role == 'admin'
    is_tp = is_tp_operator(current_user)

    if not (is_admin or is_tp):
        flash("У вас нет доступа к Kanban-доске", "error")
        return redirect(url_for('ticket_list'))

    departments = Department.query.all()
    department_list = [{'id': d.id, 'name': d.name} for d in departments]
    return render_template('admin_kanban.html', departments=department_list)

# --- Backward-compatible alias endpoint name ---
# В старом коде использовался endpoint admin_kanban.
try:
    app.add_url_rule('/admin/kanban', endpoint='admin_kanban', view_func=kanban)
except Exception:
    pass


@app.route('/api/kanban/tickets')
@login_required
def api_kanban_tickets():
    """Возвращает все заявки, сгруппированные по отделам"""
    if not isinstance(current_user, User):
        return jsonify({"error": "Access denied"}), 403

    # Доступ разрешён: админам и операторам техподдержки (1/2 линия)
    is_admin = current_user.role == 'admin'
    is_tp = is_tp_operator(current_user)

    if not (is_admin or is_tp):
        return jsonify({"error": "Access denied"}), 403

    # Загружаем все заявки с отношениями
    tickets = SupportTicket.query.options(
        joinedload(SupportTicket.department_rel),
        joinedload(SupportTicket.locked_by_rel),
        joinedload(SupportTicket.assigned_to_rel)
    ).all()



    # Индикаторы состояний (Bitrix-like) для карточек канбана
    try:
        t_ids = [t.id for t in tickets]
        last_is_op_map = {}
        if t_ids:
            tm = TicketMessage
            last_cte = db.session.query(
                tm.ticket_id.label('t_id'),
                func.max(tm.created_at).label('mx')
            ).filter(tm.ticket_id.in_(t_ids)).group_by(tm.ticket_id).subquery('last_msg_kanban')
            tm_last = aliased(tm, name='tm_last_kanban')
            rows = db.session.query(tm_last.ticket_id, tm_last.is_operator, tm_last.user_id).join(
                last_cte,
                db.and_(tm_last.ticket_id == last_cte.c.t_id, tm_last.created_at == last_cte.c.mx)
            ).all()
            for tid, is_op, user_id in rows:
                last_is_op_map[int(tid)] = (bool(is_op), user_id)

        for t in tickets:
            last_meta = last_is_op_map.get(int(t.id))
            last_is_op = last_meta[0] if last_meta else None
            last_user_id = last_meta[1] if last_meta else None
            code, title = compute_ticket_indicator(t, last_is_op, last_user_id, getattr(current_user, 'id', None))
            t.state_indicator = code
            t.state_indicator_title = title
    except Exception:
        for t in tickets:
            t.state_indicator = 'green'
            t.state_indicator_title = _INDICATOR_LABELS['green']

    # Группируем по department_id
    grouped = {}
    for ticket in tickets:
        dept_id = ticket.department_id or 0  # 0 = без отдела
        if dept_id not in grouped:
            grouped[dept_id] = []
        grouped[dept_id].append(ticket_to_dict(ticket))

    return jsonify(grouped)


def bbcode_to_text(s: str) -> str:
    """Грубое преобразование BBCode -> plain text (для канбана/превью)."""
    if not s:
        return ""
    import re
    s = str(s)
    # url tags
    # [url=https://example.com]текст[/url] -> текст
    s = re.sub(r"\[url=(.*?)\](.*?)\[/url\]", r"\2", s, flags=re.I | re.S)
    # [url]https://example.com[/url] -> https://example.com
    s = re.sub(r"\[url\](.*?)\[/url\]", r"\1", s, flags=re.I | re.S)
    # img tags
    s = re.sub(r"\[img\](.*?)\[/img\]", "", s, flags=re.I | re.S)
    # simple tags like [b], [i], [quote], [code], [list] ...
    s = re.sub(r"\[/?[a-z0-9]+(?:=[^\]]+)?\]", "", s, flags=re.I)
    # cleanup
    s = re.sub(r"\s+", " ", s).strip()
    return s


def ticket_to_dict(ticket):
    now = utcnow()

    def _plural_ru(value, forms):
        value = abs(int(value))
        if value % 10 == 1 and value % 100 != 11:
            return forms[0]
        if 2 <= value % 10 <= 4 and not 12 <= value % 100 <= 14:
            return forms[1]
        return forms[2]

    def _format_sla_hint(deadline, status):
        if not deadline or ticket.status == 'Завершена':
            return ''
        try:
            local_deadline = to_local(deadline)
            local_now = to_local(now)
            delta = local_deadline - local_now
            seconds = int(delta.total_seconds())
            if seconds < 0:
                overdue = abs(seconds)
                days = overdue // 86400
                hours = overdue // 3600
                minutes = max(1, overdue // 60)
                if days >= 30:
                    months = max(1, round(days / 30))
                    return f'- {months} {_plural_ru(months, ("месяц", "месяца", "месяцев"))}'
                if days >= 1:
                    return f'- {days} {_plural_ru(days, ("день", "дня", "дней"))}'
                if hours >= 1:
                    return f'- {hours} {_plural_ru(hours, ("час", "часа", "часов"))}'
                return f'- {minutes} {_plural_ru(minutes, ("минута", "минуты", "минут"))}'

            if local_deadline.date() == local_now.date():
                return local_deadline.strftime('%H:%M')
            tomorrow = (local_now + timedelta(days=1)).date()
            if local_deadline.date() == tomorrow:
                return f'Завтра, {local_deadline.strftime("%H:%M")}'
            if delta.days < 7:
                return local_deadline.strftime('%d.%m, %H:%M')
            return local_deadline.strftime('%d.%m.%Y %H:%M')
        except Exception:
            return ''

    # SLA статус берём из нового SLA-движка, чтобы список и канбан считали одинаково.
    sla_status = "normal"
    try:
        from helpdesk_app.services.sla_service import SLAService
        sla_view = SLAService.build_ticket_view(ticket)
        sla_status = (sla_view or {}).get('summary_status') or 'ok'
        if sla_status not in ('ok', 'overdue'):
            sla_status = 'ok'
    except Exception:
        if ticket.sla_deadline and ticket.status != 'Завершена':
            if now > ticket.sla_deadline:
                sla_status = "overdue"
            else:
                sla_status = "ok"

    is_important = (ticket.priority or '').strip() in ('critical', 'Критический')

    subject = ticket.subject
    if len(subject) > 110:
        subject = subject[:107] + '...'

    msg = bbcode_to_text(ticket.message)
    if len(msg) > 120:
        msg = msg[:120] + "..."

    created_local = to_local(ticket.created_at) if ticket.created_at else None

    return {
        'id': ticket.id,
        'subject': subject,
        'message': msg,
        'status': ticket.status,
        'department_id': ticket.department_id,
        'created_at': format_local(ticket.created_at),
        'created_date': created_local.strftime('%d.%m.%Y') if created_local else '',
        'created_time': created_local.strftime('%H:%M') if created_local else '',
        'email': ticket.email,
        'name': ticket.name,
        'priority': ticket.priority,
        'is_important': is_important,
        'sla_status': sla_status,
        'sla_hint': _format_sla_hint(ticket.sla_deadline, sla_status),
        'is_overdue': ticket.is_overdue,
        'locked_by': ticket.locked_by_rel.username if ticket.locked_by_rel else None,
        'organization': ticket.organization or '',
        'indicator': getattr(ticket, 'state_indicator', None) or 'green',
        'indicator_title': getattr(ticket, 'state_indicator_title', None) or _INDICATOR_LABELS['green'],
        'bitrix_task_url': ticket.bitrix_task_url
    }
@app.route('/api/kanban/ticket/<int:ticket_id>', methods=['PUT'])
@login_required
def api_kanban_update_ticket(ticket_id):
    if not isinstance(current_user, User):
        return jsonify({"error": "Access denied"}), 403

    # Доступ разрешён: админам и операторам техподдержки (1/2 линия)
    is_admin = current_user.role == 'admin'
    is_tp = is_tp_operator(current_user)

    if not (is_admin or is_tp):
        return jsonify({'error': 'Access denied'}), 403

    ticket = SupportTicket.query.get_or_404(ticket_id)
    data = request.get_json()

    changed = False

    # Для сотрудников 1ой линии добавляем проверки прав
    if False:  # ограничения для 1 линии отключены (теперь техподдержка 1/2 линия может распределять)
        # Сотрудник 1ой линии может перемещать только из:
        # 1. "Требуется обработка" (ID: 1) 
        # 2. Своего отдела "1ая линия ТП" (ID: 2)
        if ticket.department_id not in [1, 2]:
            return jsonify({'error': 'Вы можете перемещать только заявки из отделов "Требуется обработка" и "1ая линия ТП"'}), 403
        
        # Проверяем, куда пытается переместить
        if 'department_id' in data:
            new_dept_id = int(data['department_id'])
            # Нельзя перемещать обратно в "Требуется обработка" (ID: 1)
            if new_dept_id == 1:
                return jsonify({'error': 'Вы не можете перемещать заявки обратно в "Требуется обработка"'}), 403
            # Проверяем, что отдел существует
            new_dept = Department.query.get(new_dept_id)
            if not new_dept:
                return jsonify({'error': 'Отдел не найден'}), 404

    if 'status' in data and data['status'] in ['Новая', 'Принята', 'Завершена']:
        ticket.status = data['status']
        changed = True

    if 'priority' in data and data['priority'] in ['high', 'medium', 'low']:
        ticket.priority = new_priority
        changed = True

    if 'department_id' in data:
        try:
            dept_id = int(data['department_id'])
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid department_id'}), 400

        dept = Department.query.get(dept_id)
        if dept:
            old_dept_name = ticket.department.name if ticket.department else 'Без отдела'
            ticket.department_id = dept.id
            # ✅ Используем log_ticket_change
            log_ticket_change(ticket.id, current_user.id, 'department', old_dept_name, dept.name)
            if dept.name != "Требуется обработка" and ticket.status == "Новая":
                ticket.status = "Принята"
            changed = True
        else:
            return jsonify({'error': 'Department not found'}), 404

    if changed:
        db.session.commit()
        return jsonify({'success': True})
    else:
        return jsonify({'error': 'No valid fields to update'}), 400

@app.route('/admin/stats')
@login_required
def admin_stats():
    """Совместимость: редирект на новую страницу."""
    return redirect(url_for('admin_statistics', **request.args))

@app.route('/admin/analytics')
@login_required
def admin_analytics():
    if getattr(current_user, 'role', None) == 'client' or current_user.role != 'admin':
        return redirect(url_for('admin'))

    period = request.args.get('period', '30')
    department_id = request.args.get('department_id', type=int)
    operator_id = request.args.get('operator_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    # Старый URL. Теперь статистика и отчёты разнесены.
    return redirect(url_for(
        'admin_statistics',
        period=period,
        department_id=department_id or '',
        operator_id=operator_id or '',
        date_from=date_from,
        date_to=date_to
    ))


@app.route('/admin/analytics/export.xlsx')
@login_required
def admin_analytics_export():
    if getattr(current_user, 'role', None) == 'client' or current_user.role != 'admin':
        return redirect(url_for('admin'))

    period = request.args.get('period', '30')
    department_id = request.args.get('department_id', type=int)
    operator_id = request.args.get('operator_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    # Старый URL экспорта. Оставлен для совместимости.
    return redirect(url_for(
        'admin_reports_export_xlsx',
        period=period,
        department_id=department_id or '',
        operator_id=operator_id or '',
        date_from=date_from,
        date_to=date_to
    ))


@app.route('/admin/statistics')
@login_required
def admin_statistics():
    """Статистика (дашборд)."""
    if getattr(current_user, 'role', None) == 'client' or current_user.role != 'admin':
        return redirect(url_for('admin'))

    period = request.args.get('period', '30')
    department_id = request.args.get('department_id', type=int)
    operator_id = request.args.get('operator_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    data = _build_analytics_data(period, department_id, operator_id, date_from, date_to)
    return render_template('admin_statistics.html', **data)


@app.route('/admin/reports/legacy')
@login_required
def admin_reports_legacy():
    """Совместимость: редирект на новую страницу."""
    return redirect(url_for('admin_reports', **request.args))

@app.route('/admin/reports/export.xlsx')
@login_required
def admin_reports_export_xlsx():
    """Excel выгрузка для отчётов."""
    if getattr(current_user, 'role', None) == 'client' or current_user.role != 'admin':
        return redirect(url_for('admin'))

    period = request.args.get('period', '30')
    department_id = request.args.get('department_id', type=int)
    operator_id = request.args.get('operator_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    view = request.args.get('view', 'created')
    data = _build_analytics_data(period, department_id, operator_id, date_from, date_to, for_export=True, report_view=view)

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"

    ws.append(["Параметр", "Значение"])
    ws.append(["Период", data.get('period_label', '')])
    ws.append(["Создано", data['sla_metrics']['created_count']])
    ws.append(["Закрыто", data['sla_metrics']['resolved_count']])
    ws.append(["Net flow (создано - закрыто)", data['sla_metrics']['net_flow']])
    ws.append(["Очередь на начало", data['sla_metrics']['backlog_start']])
    ws.append(["Очередь на конец", data['sla_metrics']['backlog_end']])
    ws.append(["Открыто сейчас", data['in_progress']])
    ws.append(["Просрочено сейчас (активные)", data['overdue_active']])
    ws.append(["SLA соблюдено (закрытые в период)", f"{data['sla_compliance_percent']}%"])
    ws.append(["Среднее время решения (MTTR)", data['sla_metrics']['avg_time_to_resolve']])
    ws.append(["Среднее время до первого ответа (FRT)", data['sla_metrics']['frt_avg']])
    ws.append(["Среднее время до принятия (Accept)", data['sla_metrics'].get('avg_time_to_accept','')])
    ws.append(["Среднее время в работе (In work)", data['sla_metrics'].get('avg_time_in_work','')])
    ws.append(["Переоткрыто (reopen)", data.get('events', {}).get('reopened', 0)])
    ws.append(["Доработка от клиента", data.get('events', {}).get('rework', 0)])
    ws.append(["Переводы между отделами", data.get('events', {}).get('department_moves', 0)])
    ws.append(["Удовлетворённость (👍/всего)", f"{data['helpful_pos']}/{data['helpful_total']} ({data['satisfaction']}%)"])

    ws2 = wb.create_sheet("Операторы")
    ws2.append(["Оператор", "Создано", "Закрыто", "Открыто (конец периода)", "Просрочено (сейчас)", "Среднее время решения"])
    for r in data['operator_rows']:
        ws2.append([r['username'], r.get('created', 0), r.get('resolved', 0), r.get('open_end', 0), r.get('overdue_open', 0), r.get('avg_mttr', '')])

    ws3 = wb.create_sheet("Заявки")
    ws3.append(["ID", "Создана", "Закрыта", "Статус", "Приоритет", "Отдел", "Оператор", "Тема", "SLA дедлайн", "Просрочена", "FRT", "Accept", "MTTR", "In work", "Оценка/👍"])
    rows = data.get('export_rows', [])
    for t in rows:
        ws3.append([
            t['id'],
            t['created_at'],
            t['resolved_at'],
            t['status'],
            t['priority'],
            t['department'],
            t['operator'],
            t['subject'],
            t['sla_deadline'],
            t['sla_breached'],
            t.get('frt',''),
            t.get('accept',''),
            t.get('mttr',''),
            t.get('in_work',''),
            t['helpful'],
        ])

    filename = f"reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    tmp_path = os.path.join(app.root_path, 'instance', filename)
    try:
        wb.save(tmp_path)
        return send_file(tmp_path, as_attachment=True, download_name=filename)
    except Exception:
        from io import BytesIO
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _build_analytics_data(period: str, department_id: int | None, operator_id: int | None, date_from: str, date_to: str, for_export: bool=False, report_view: str='created', export_limit: int=500):
    """Считает статистику в стиле ТП: created vs resolved + SLA + FRT/MTTR."""
    now = utcnow()

    # --- диапазон ---
    start = None
    end = None
    period_label = ""
    if period == 'custom' and date_from and date_to:
        try:
            start = datetime.strptime(date_from, '%Y-%m-%d')
            end = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            period_label = f"{date_from}..{date_to}"
        except Exception:
            start = None
            end = None
    if start is None or end is None:
        days = int(period) if str(period).isdigit() else 30
        start = (now - timedelta(days=days)).replace(hour=0, minute=0, second=0, microsecond=0)
        end = now
        period_label = f"последние {days} дней"

    all_departments = Department.query.all()
    operators = User.query.filter(User.role.in_(['admin', 'operator'])).order_by(User.username.asc()).all()
    op_map = {u.id: u.username for u in operators}

    base_q = SupportTicket.query

    if department_id:
        base_q = base_q.filter(
            or_(
                SupportTicket.department_id == department_id,
                SupportTicket.shared_departments_rel.any(Department.id == department_id)
            )
        )

    if operator_id:
        base_q = base_q.filter(
            or_(
                SupportTicket.assigned_to_id == operator_id,
                SupportTicket.locked_by == operator_id,
                SupportTicket.created_by_operator_id == operator_id
            )
        )

    base_q = base_q.filter(SupportTicket.created_at < end).filter(
        or_(
            SupportTicket.created_at >= start,
            SupportTicket.closed_at >= start,
            SupportTicket.marked_as_completed_at >= start,
            SupportTicket.auto_closed_at >= start,
            SupportTicket.closed_at.is_(None)
        )
    )

    tickets = base_q.options(
        joinedload(SupportTicket.department_rel),
        joinedload(SupportTicket.locked_by_rel),
        joinedload(SupportTicket.assigned_to_rel),
        joinedload(SupportTicket.category_rel),
    ).all()
    sla_views = build_ticket_sla_views(tickets)

    def resolved_at(t: SupportTicket):
        return t.closed_at or t.auto_closed_at or t.marked_as_completed_at

    def responsible_id(t: SupportTicket):
        return t.assigned_to_id or t.locked_by or t.created_by_operator_id

    def in_range(d):
        return d is not None and (start <= d < end)

    def open_at(moment: datetime, t: SupportTicket):
        if not t.created_at or t.created_at >= moment:
            return False
        ra = resolved_at(t)
        return (ra is None) or (ra >= moment)

    def sla_deadline_for(t: SupportTicket):
        view = sla_views.get(getattr(t, 'id', None)) or {}
        return view.get('resolve_deadline') or t.sla_deadline

    def sla_summary_status_for(t: SupportTicket):
        view = sla_views.get(getattr(t, 'id', None)) or {}
        return view.get('summary_status') or 'normal'

    created_in_period = [t for t in tickets if in_range(t.created_at)]
    resolved_in_period = [t for t in tickets if in_range(resolved_at(t))]

    created_count = len(created_in_period)
    resolved_count = len(resolved_in_period)
    net_flow = created_count - resolved_count

    backlog_start = sum(1 for t in tickets if open_at(start, t))
    backlog_end = sum(1 for t in tickets if open_at(end, t))
    open_now = sum(1 for t in tickets if open_at(now, t))
    overdue_active = sum(1 for t in tickets if open_at(now, t) and sla_summary_status_for(t) == 'overdue')

    # SLA
    sla_on_time = 0
    sla_overdue_completed = 0
    for t in resolved_in_period:
        deadline = sla_deadline_for(t)
        if not deadline:
            continue
        ra = resolved_at(t)
        if not ra:
            continue
        if ra <= deadline:
            sla_on_time += 1
        else:
            sla_overdue_completed += 1
    total_sla_completed = sla_on_time + sla_overdue_completed
    sla_compliance_percent = (sla_on_time / total_sla_completed * 100) if total_sla_completed else 0.0

    # MTTR
    total_time = 0.0
    completed_cnt = 0
    for t in resolved_in_period:
        ra = resolved_at(t)
        if not (t.created_at and ra):
            continue
        sec = (ra - t.created_at).total_seconds()
        if sec >= 0:
            total_time += sec
            completed_cnt += 1

    def fmt_duration(seconds):
        if seconds is None:
            return 'Нет данных'
        seconds = float(seconds)
        if seconds < 60:
            return f"{int(seconds)} сек"
        if seconds < 3600:
            return f"{int(seconds/60)} мин"
        if seconds < 86400:
            return f"{int(seconds/3600)} ч {int((seconds%3600)/60)} м"
        return f"{int(seconds/86400)} д {int((seconds%86400)/3600)} ч"

    avg_time_to_resolve = fmt_duration((total_time / completed_cnt) if completed_cnt else None)

    # FRT
    frt_avg = 'Нет данных'
    frt_cnt = 0
    frt_total = 0.0
    created_ids = [t.id for t in created_in_period]
    if created_ids:
        first_ops = dict(
            db.session.query(TicketMessage.ticket_id, func.min(TicketMessage.created_at))
            .filter(TicketMessage.ticket_id.in_(created_ids), TicketMessage.is_operator.is_(True))
            .group_by(TicketMessage.ticket_id)
            .all()
        )
        for t in created_in_period:
            first_at = first_ops.get(t.id)
            if first_at and t.created_at:
                sec = (first_at - t.created_at).total_seconds()
                if sec >= 0:
                    frt_total += sec
                    frt_cnt += 1
        if frt_cnt:
            frt_avg = fmt_duration(frt_total / frt_cnt)


    # Время до принятия / взятия в работу (по истории статусов)
    avg_time_to_accept = 'Нет данных'
    avg_time_in_work = 'Нет данных'
    if created_ids:
        accept_map = dict(
            db.session.query(TicketHistory.ticket_id, func.min(TicketHistory.timestamp))
            .filter(TicketHistory.ticket_id.in_(created_ids),
                    TicketHistory.field == 'status',
                    TicketHistory.new_value.in_(['Принята', 'В работе']))
            .group_by(TicketHistory.ticket_id)
            .all()
        )
        work_map = dict(
            db.session.query(TicketHistory.ticket_id, func.min(TicketHistory.timestamp))
            .filter(TicketHistory.ticket_id.in_(created_ids),
                    TicketHistory.field == 'status',
                    TicketHistory.new_value == 'В работе')
            .group_by(TicketHistory.ticket_id)
            .all()
        )

        acc_total = 0.0
        acc_cnt = 0
        for t in created_in_period:
            a = accept_map.get(t.id)
            if a and t.created_at:
                sec = (a - t.created_at).total_seconds()
                if sec >= 0:
                    acc_total += sec
                    acc_cnt += 1
        if acc_cnt:
            avg_time_to_accept = fmt_duration(acc_total / acc_cnt)

        work_total = 0.0
        work_cnt = 0
        for t in resolved_in_period:
            ra = resolved_at(t)
            w = work_map.get(t.id)
            if ra and w:
                sec = (ra - w).total_seconds()
                if sec >= 0:
                    work_total += sec
                    work_cnt += 1
        if work_cnt:
            avg_time_in_work = fmt_duration(work_total / work_cnt)

    # распределения по созданным в период
    status_counts = {}
    priority_counts = {}
    category_counts = {}
    for t in created_in_period:
        status_counts[t.status] = status_counts.get(t.status, 0) + 1
        p = normalize_priority(t.priority, default='Низкий')
        priority_counts[p] = priority_counts.get(p, 0) + 1
        cat = (t.category_rel.name if t.category_rel else 'Без категории')
        category_counts[cat] = category_counts.get(cat, 0) + 1

    # удовлетворённость (по created в период)
    helpful_total = 0
    helpful_pos = 0
    helpful_neg = 0
    for t in created_in_period:
        if t.helpful is None:
            continue
        helpful_total += 1
        if t.helpful:
            helpful_pos += 1
        else:
            helpful_neg += 1
    satisfaction = int(round(helpful_pos / helpful_total * 100)) if helpful_total else 0

    # новые сегодня
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    new_today = SupportTicket.query.filter(SupportTicket.created_at >= today_start).count()

    spam_count = sum(1 for t in created_in_period if getattr(t, 'is_spam', False) or getattr(t, 'close_reason', None) == 'spam')

    kb_total = KnowledgeBaseArticle.query.count()
    kb_published = KnowledgeBaseArticle.query.filter_by(is_published=True).count()
    kb_favs = KnowledgeBaseFavorite.query.count()

    # операторы
    op_stats = {}
    def touch(op_id):
        if not op_id:
            return None
        if op_id not in op_stats:
            op_stats[op_id] = {'username': op_map.get(op_id, f'ID {op_id}'), 'created': 0, 'resolved': 0, 'open_end': 0, 'overdue_open': 0, 'mttr_sec_total': 0.0, 'mttr_cnt': 0}
        return op_stats[op_id]

    for t in created_in_period:
        s = touch(responsible_id(t))
        if s:
            s['created'] += 1
    for t in resolved_in_period:
        s = touch(responsible_id(t))
        if s:
            s['resolved'] += 1
            ra = resolved_at(t)
            if t.created_at and ra:
                sec = (ra - t.created_at).total_seconds()
                if sec >= 0:
                    s['mttr_sec_total'] += sec
                    s['mttr_cnt'] += 1
    for t in tickets:
        s = touch(responsible_id(t))
        if not s:
            continue
        if open_at(end, t):
            s['open_end'] += 1
        if open_at(now, t) and sla_summary_status_for(t) == 'overdue':
            s['overdue_open'] += 1

    operator_rows = []
    for _, s in op_stats.items():
        operator_rows.append({
            'username': s['username'],
            # Для блока "Нагрузка по операторам" показываем число открытых заявок
            # на конец выбранного периода — это ближе к реальной текущей нагрузке.
            'count': s['open_end'],
            'created': s['created'],
            'resolved': s['resolved'],
            'open_end': s['open_end'],
            'overdue_open': s['overdue_open'],
            'avg_mttr': fmt_duration((s['mttr_sec_total'] / s['mttr_cnt']) if s['mttr_cnt'] else None),
        })
    operator_rows.sort(key=lambda r: (r['open_end'], r['created']), reverse=True)

    report_rows = sorted(created_in_period, key=lambda t: (t.created_at or now), reverse=True)[:200]
    for t in report_rows:
        t.sla_view = sla_views.get(getattr(t, 'id', None))
        t.ui_is_overdue = bool(t.sla_view and t.sla_view.get('summary_status') == 'overdue')

    # --- тренды (created vs resolved по дням) ---
    def _date_key(dt: datetime | None):
        return dt.date().isoformat() if dt else None

    d0 = start.date()
    d1 = (end - timedelta(seconds=1)).date() if end else now.date()
    days_span = (d1 - d0).days
    labels = []
    created_series = []
    resolved_series = []
    if days_span >= 0 and days_span <= 4000:
        created_map = {}
        resolved_map = {}
        for t in created_in_period:
            k = _date_key(t.created_at)
            if k:
                created_map[k] = created_map.get(k, 0) + 1
        for t in resolved_in_period:
            k = _date_key(resolved_at(t))
            if k:
                resolved_map[k] = resolved_map.get(k, 0) + 1

        for i in range(days_span + 1):
            day = (d0 + timedelta(days=i)).isoformat()
            labels.append(day)
            created_series.append(int(created_map.get(day, 0)))
            resolved_series.append(int(resolved_map.get(day, 0)))

    # --- SLA по приоритетам (закрытые в период) ---
    sla_by_priority = {}
    for t in resolved_in_period:
        deadline = sla_deadline_for(t)
        if not deadline:
            continue
        p = normalize_priority(t.priority, default='Низкий')
        ra = resolved_at(t)
        if not ra:
            continue
        if p not in sla_by_priority:
            sla_by_priority[p] = {'on_time': 0, 'overdue': 0, 'total': 0}
        sla_by_priority[p]['total'] += 1
        if ra <= deadline:
            sla_by_priority[p]['on_time'] += 1
        else:
            sla_by_priority[p]['overdue'] += 1

    # --- события (из TicketHistory) ---
    ticket_ids = [t.id for t in tickets]
    events = {'reopened': 0, 'rework': 0, 'department_moves': 0}
    try:
        if ticket_ids:
            hq = (TicketHistory.query
                .filter(TicketHistory.ticket_id.in_(ticket_ids))
                .filter(TicketHistory.timestamp >= start)
                .filter(TicketHistory.timestamp < end))

            # Переводы между отделами
            events['department_moves'] = hq.filter(TicketHistory.field == 'department').count()

            # Переоткрытия: из финального статуса -> в нефинальный
            closed_statuses = {'Завершена', 'Спам', 'Дубликат', 'Ошибочная'}
            reopen_rows = (hq.filter(TicketHistory.field == 'status')
                .filter(TicketHistory.old_value.in_(list(closed_statuses)))
                .filter(~TicketHistory.new_value.in_(list(closed_statuses)))
                .all())
            events['reopened'] = len(reopen_rows)

            # Доработка от клиента: "Ожидает подтверждения клиента" -> "В работе"
            rework_rows = (hq.filter(TicketHistory.field == 'status')
                .filter(TicketHistory.old_value == 'Ожидает подтверждения клиента')
                .filter(TicketHistory.new_value == 'В работе')
                .all())
            events['rework'] = len(rework_rows)
    except Exception:
        pass

    export_rows = []
    if for_export:
        view = (report_view or 'created').strip().lower()
        if view not in ('created', 'resolved'):
            view = 'created'

        source = created_in_period if view == 'created' else resolved_in_period

        def _sort_key(t):
            if view == 'resolved':
                ra = resolved_at(t)
                return (ra or datetime.min)
            return (t.created_at or datetime.min)

        source_sorted = sorted(source, key=_sort_key, reverse=True)
        if export_limit and export_limit > 0:
            source_sorted = source_sorted[:export_limit]

        ids = [t.id for t in source_sorted]

        # FRT: min operator msg time
        first_ops = {}
        try:
            if ids:
                first_ops = dict(
                    db.session.query(TicketMessage.ticket_id, func.min(TicketMessage.created_at))
                    .filter(TicketMessage.ticket_id.in_(ids), TicketMessage.is_operator.is_(True))
                    .group_by(TicketMessage.ticket_id)
                    .all()
                )
        except Exception:
            first_ops = {}

        # Accept/In work (history)
        accept_map = {}
        work_map = {}
        try:
            if ids:
                accept_map = dict(
                    db.session.query(TicketHistory.ticket_id, func.min(TicketHistory.timestamp))
                    .filter(TicketHistory.ticket_id.in_(ids),
                            TicketHistory.field == 'status',
                            TicketHistory.new_value.in_(['Принята', 'В работе']))
                    .group_by(TicketHistory.ticket_id)
                    .all()
                )
                work_map = dict(
                    db.session.query(TicketHistory.ticket_id, func.min(TicketHistory.timestamp))
                    .filter(TicketHistory.ticket_id.in_(ids),
                            TicketHistory.field == 'status',
                            TicketHistory.new_value == 'В работе')
                    .group_by(TicketHistory.ticket_id)
                    .all()
                )
        except Exception:
            accept_map, work_map = {}, {}

        def _fmt(seconds):
            return fmt_duration(seconds) if seconds is not None else ''

        for t in source_sorted:
            ra = resolved_at(t)
            op_id = responsible_id(t)

            frt_sec = None
            fo = first_ops.get(t.id)
            if fo and t.created_at:
                frt_sec = max(0.0, (fo - t.created_at).total_seconds())

            mttr_sec = None
            if ra and t.created_at:
                mttr_sec = max(0.0, (ra - t.created_at).total_seconds())

            acc_sec = None
            a = accept_map.get(t.id)
            if a and t.created_at:
                acc_sec = max(0.0, (a - t.created_at).total_seconds())

            inwork_sec = None
            w = work_map.get(t.id)
            if ra and w:
                inwork_sec = max(0.0, (ra - w).total_seconds())

            export_rows.append({
                'id': t.id,
                'created_at': t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
                'resolved_at': ra.strftime('%Y-%m-%d %H:%M') if ra else '',
                'status': t.status or '',
                'priority': normalize_priority(getattr(t, 'priority', None), default='Низкий'),
                'department': (t.department_rel.name if t.department_rel else ''),
                'operator': op_map.get(op_id, '') if op_id else '',
                'subject': (t.subject or ''),
                'sla_deadline': format_local(sla_deadline_for(t)) if sla_deadline_for(t) else '',
                'sla_breached': ('Да' if (sla_deadline_for(t) and ra and ra > sla_deadline_for(t)) else 'Нет') if sla_deadline_for(t) else '',
                'overdue_now': bool((not ra) and sla_summary_status_for(t) == 'overdue'),

                'frt': _fmt(frt_sec),
                'mttr': _fmt(mttr_sec),
                'accept': _fmt(acc_sec),
                'in_work': _fmt(inwork_sec),

                'helpful': ('👍' if t.helpful is True else ('👎' if t.helpful is False else '')),
            })
    # переменные под текущий шаблон (сохраняем имена)
    data = dict(
        period=period,
        period_label=period_label,
        department_id=department_id,
        operator_id=operator_id,
        date_from=date_from,
        date_to=date_to,
        all_departments=all_departments,
        operators=operators,
        total=created_count,  # в шаблоне «Всего» = создано в период
        resolved=resolved_count,
        in_progress=open_now,  # открыто сейчас
        new_count=backlog_end,  # очередь на конец
        new_today=new_today,
        overdue_active=overdue_active,
        satisfaction=satisfaction,
        helpful_total=helpful_total,
        helpful_pos=helpful_pos,
        helpful_neg=helpful_neg,
        sla_metrics={
            'on_time': sla_on_time,
            'overdue_completed': sla_overdue_completed,
            'avg_time_to_resolve': avg_time_to_resolve,
            'frt_avg': frt_avg,
            'avg_time_to_accept': avg_time_to_accept,
            'avg_time_in_work': avg_time_in_work,
            'created_count': created_count,
            'resolved_count': resolved_count,
            'net_flow': net_flow,
            'backlog_start': backlog_start,
            'backlog_end': backlog_end,
        },
        sla_compliance_percent=round(sla_compliance_percent, 1),
        status_counts=status_counts,
        priority_counts=priority_counts,
        category_counts=category_counts,
        trends={
            'labels': labels,
            'created': created_series,
            'resolved': resolved_series,
        },
        sla_by_priority=sla_by_priority,
        events=events,
        operator_rows=operator_rows,
        spam_count=spam_count,
        kb_total=kb_total,
        kb_published=kb_published,
        kb_favs=kb_favs,
        report_rows=report_rows,
        now=now,
        range_start=start,
        range_end=end,
        export_rows=export_rows,
    )
    return data


@app.route('/admin/reports')
@login_required
def admin_reports():
    """Отчёты: таблица + выгрузка. Отдельно от страницы статистики."""
    if getattr(current_user, 'role', None) == 'client' or current_user.role != 'admin':
        flash("Доступ запрещён", "danger")
        return redirect(url_for('admin'))

    period = request.args.get('period', '30')
    department_id = request.args.get('department_id', type=int)
    operator_id = request.args.get('operator_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    # Берём те же расчёты, что и в статистике, но просим подготовить export_rows (500 строк)
    view = request.args.get('view', 'created')
    data = _build_analytics_data(period, department_id, operator_id, date_from, date_to, for_export=True, report_view=view)

    # В отчётах нам важнее таблица: используем export_rows, а также показываем несколько KPI сверху
    return render_template('admin_reports.html', **data)

@app.route('/admin/reports/legacy/export')
@login_required
def export_reports_excel_legacy():
    if getattr(current_user, 'role', None) == 'client' or current_user.role != 'admin':
        flash("Доступ запрещён", "error")
        return redirect(url_for('admin'))
    dept_filter = request.args.get('department', '')
    operator_filter = request.args.get('operator', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    query = SupportTicket.query
    if dept_filter:
        query = query.filter(SupportTicket.department == dept_filter)
    if operator_filter:
        query = query.filter(SupportTicket.locked_by == operator_filter)
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        query = query.filter(SupportTicket.created_at >= start_dt)
    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(SupportTicket.created_at < end_dt)
    tickets = query.order_by(
    db.case(
        (SupportTicket.priority.in_(['Высокий','high']), 1),
        (SupportTicket.priority.in_(['Обычный','medium','Средний']), 2),
        else_=3
    ),
    SupportTicket.created_at.desc()
).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт по заявкам"
    headers = ["ID", "Дата", "Тема", "От кого", "Email", "Отдел", "Статус", "Оператор", "Организация", "ИНН", "SLA"]
    ws.append(headers)
    for t in tickets:
        operator_user = db.session.get(User, t.locked_by) if t.locked_by else None
        operator = operator_user.username if operator_user else ""
        sla_status = "Просрочено" if (not t.is_resolved and t.sla_deadline and utcnow() > t.sla_deadline) else "В срок"
        ws.append([
            t.id,
            t.created_at.strftime('%d.%m.%Y %H:%M'),
            t.subject,
            t.name,
            t.email,
            t.department,
            t.status,
            operator,
            t.organization or "",
            t.inn or "",
            sla_status
        ])
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"Отчёт_техподдержка_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )

@app.route('/api/charts/data')
@login_required
def api_charts_data():
    if getattr(current_user, 'role', None) == 'client' or current_user.role != 'admin':
        return jsonify({}), 403
    
    # Получаем все отделы из базы
    all_depts = Department.query.all()
    dept_names = [dept.name for dept in all_depts]
    
    # Статусы
    status_counts = {s: 0 for s in STATUSES}
    tickets = SupportTicket.query.all()
    for t in tickets:
        status_counts[t.status] = status_counts.get(t.status, 0) + 1
    
    # Отделы — правильный подсчёт
    dept_counts = {dept.name: 0 for dept in all_depts}
    dept_counts['Требуется обработка'] = 0
    
    for t in tickets:
        if t.department_rel:
            dept_name = t.department_rel.name
            if dept_name in dept_counts:
                dept_counts[dept_name] += 1
            else:
                dept_counts['Требуется обработка'] += 1
        else:
            dept_counts['Требуется обработка'] += 1
    
    # Динамика (7 дней)
    today = utcnow().date()
    dates = [(today - timedelta(days=i)).strftime('%d.%m') for i in range(6, -1, -1)]
    daily = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        next_day = day + timedelta(days=1)
        daily.append(SupportTicket.query.filter(
            SupportTicket.created_at >= day,
            SupportTicket.created_at < next_day
        ).count())
    
    # SLA метрики
    sla_metrics = calculate_sla_metrics()
    
    return jsonify({
        'tickets_by_day': {'labels': dates, 'data': daily},
        'status_pie': {'labels': STATUSES, 'data': [status_counts[s] for s in STATUSES]},
        'sla_metrics': sla_metrics,  # <-- Изменено с sla_compliance на sla_metrics
        'department_bar': {
            'labels': dept_names + ['Требуется обработка'],
            'data': [dept_counts.get(name, 0) for name in dept_names] + [dept_counts.get('Требуется обработка', 0)]
        }
    })

def calculate_sla_metrics():
    """Расчёт метрик SLA"""
    from datetime import datetime, timedelta
    
    # Все завершённые заявки
    completed_tickets = SupportTicket.query.filter(
        SupportTicket.status == 'Завершена'
    ).all()
    
    # Заявки выполненные в срок
    on_time_completed = 0
    overdue_completed = 0
    total_resolution_time = timedelta()
    
    for ticket in completed_tickets:
        if ticket.sla_deadline and ticket.marked_as_completed_at:
            if ticket.marked_as_completed_at <= ticket.sla_deadline:
                on_time_completed += 1
            else:
                overdue_completed += 1
            
            # Расчёт времени решения
            if ticket.created_at and ticket.marked_as_completed_at:
                total_resolution_time += ticket.marked_as_completed_at - ticket.created_at
    
    # Активные заявки с просроченным SLA
    overdue_active = SupportTicket.query.filter(
        SupportTicket.status != 'Завершена',
        SupportTicket.sla_deadline < utcnow()
    ).count()
    
    # Активные заявки с контролем SLA
    active_with_sla = SupportTicket.query.filter(
        SupportTicket.status != 'Завершена',
        SupportTicket.sla_deadline.isnot(None)
    ).count()
    
    # Среднее время решения
    avg_time_to_resolve = "—"
    if completed_tickets and total_resolution_time.total_seconds() > 0:
        avg_seconds = total_resolution_time.total_seconds() / len(completed_tickets)
        hours = int(avg_seconds // 3600)
        minutes = int((avg_seconds % 3600) // 60)
        avg_time_to_resolve = f"{hours} ч {minutes} м"
    
    return {
        'on_time': on_time_completed,
        'overdue_completed': overdue_completed,
        'overdue_active': overdue_active,
        'active_with_sla': active_with_sla,
        'avg_time_to_resolve': avg_time_to_resolve
    }



# =========================
# SLA / business calendar helpers
# =========================

SLA_DEFAULTS = {
    'timezone': 'Asia/Yekaterinburg',
    'work_start': '09:00',
    'work_end': '18:00',
    'workdays': '0,1,2,3,4',
    'first_response_minutes': '60',
    'resolve_minutes': '1440',
    'pause_statuses': 'Ожидание клиента',
}


def get_sla_settings():
    cfg = {}
    for key, default in SLA_DEFAULTS.items():
        cfg[key] = get_setting(f'sla.{key}', default) or default
    return cfg


def parse_hhmm(value, fallback=time(9, 0)):
    try:
        parts = (value or '').split(':')
        return time(int(parts[0]), int(parts[1]))
    except Exception:
        return fallback


def normalize_status_name(value):
    return (value or '').strip().lower()


def get_pause_statuses():
    cfg = get_sla_settings()
    raw = cfg.get('pause_statuses', '') or ''
    return {normalize_status_name(x) for x in raw.split(',') if normalize_status_name(x)}


def get_calendar_row(day: date):
    return WorkCalendarDay.query.filter_by(date=day).first()


def is_business_day(day: date) -> bool:
    row = get_calendar_row(day)
    if row is not None:
        return row.day_type in ('workday', 'short_day')
    cfg = get_sla_settings()
    workdays = {int(x) for x in (cfg.get('workdays') or '0,1,2,3,4').split(',') if str(x).strip().isdigit()}
    return day.weekday() in workdays


def business_bounds(day: date):
    cfg = get_sla_settings()
    tz = get_runtime_timezone()
    start_t = parse_hhmm(cfg['work_start'], time(9, 0))
    end_t = parse_hhmm(cfg['work_end'], time(18, 0))
    start = datetime.combine(day, start_t, tzinfo=tz)
    end = datetime.combine(day, end_t, tzinfo=tz)
    row = get_calendar_row(day)
    if row is not None and row.day_type == 'short_day':
        end = end - timedelta(hours=1)
    return start, end


def add_business_minutes(start_dt, minutes: int | None):
    if not start_dt or not minutes:
        return start_dt
    remaining = int(minutes)
    current = to_local(start_dt)
    while remaining > 0:
        if not is_business_day(current.date()):
            current = datetime.combine(current.date() + timedelta(days=1), time(0,0), tzinfo=get_runtime_timezone())
            continue
        start_bound, end_bound = business_bounds(current.date())
        if current < start_bound:
            current = start_bound
        if current >= end_bound:
            current = datetime.combine(current.date() + timedelta(days=1), time(0,0), tzinfo=get_runtime_timezone())
            continue
        chunk = min(remaining, int((end_bound - current).total_seconds() // 60))
        if chunk <= 0:
            current = datetime.combine(current.date() + timedelta(days=1), time(0,0), tzinfo=get_runtime_timezone())
            continue
        current += timedelta(minutes=chunk)
        remaining -= chunk
    return current.astimezone(UTC).replace(tzinfo=None)


def first_operator_reply_at(ticket):
    try:
        row = TicketMessage.query.filter_by(ticket_id=ticket.id, is_operator=True).order_by(TicketMessage.created_at.asc()).first()
        return row.created_at if row else None
    except Exception:
        return None


def first_operator_reply_map(ticket_ids):
    """Batch first operator reply timestamps for a set of ticket IDs."""
    ids = [int(x) for x in (ticket_ids or []) if x]
    if not ids:
        return {}
    try:
        return dict(
            db.session.query(TicketMessage.ticket_id, func.min(TicketMessage.created_at))
            .filter(TicketMessage.ticket_id.in_(ids), TicketMessage.is_operator.is_(True))
            .group_by(TicketMessage.ticket_id)
            .all()
        )
    except Exception:
        return {}


def ticket_closed_at_value(ticket):
    return getattr(ticket, 'closed_at', None) or getattr(ticket, 'auto_closed_at', None) or getattr(ticket, 'marked_as_completed_at', None)


def is_sla_paused(ticket):
    return normalize_status_name(getattr(ticket, 'status', '')) in get_pause_statuses()


def _humanize_minutes(mins: int):
    mins = max(0, int(mins))
    days, rem = divmod(mins, 60*24)
    hours, minutes = divmod(rem, 60)
    parts = []
    if days:
        parts.append(f'{days} д')
    if hours:
        parts.append(f'{hours} ч')
    if minutes or not parts:
        parts.append(f'{minutes} мин')
    return ' '.join(parts[:2])


def ticket_sla_view(ticket, first_reply_at=None, now_local=None):
    # Локальный импорт как страховка для legacy-рендеринга /tickets,
    # даже если верхний импорт в старой копии файла не подтянулся.
    from helpdesk_app.services.sla_service import SLAService
    return SLAService.build_ticket_view(ticket, first_reply_at=first_reply_at, now_local=now_local)


def build_ticket_sla_views(tickets):
    from helpdesk_app.services.sla_service import SLAService
    return SLAService.build_ticket_views(tickets)


def recalc_ticket_sla(ticket):
    return SLAService.sync_deadline_to_ticket(ticket)


def calendar_month_view(year: int, month: int):
    import calendar as _calendar
    cal = _calendar.Calendar(firstweekday=0)
    rows = {r.date: r for r in WorkCalendarDay.query.filter(
        WorkCalendarDay.date >= date(year, month, 1),
        WorkCalendarDay.date <= date(year, month, _calendar.monthrange(year, month)[1])
    ).all()}
    weeks = []
    for week in cal.monthdatescalendar(year, month):
        out = []
        for d in week:
            if d.month != month:
                out.append(None)
                continue
            row = rows.get(d)
            if row is not None:
                day_type = row.day_type
            else:
                day_type = 'workday' if is_business_day(d) else 'weekend'
            out.append({'date': d, 'row': row, 'day_type': day_type})
        weeks.append(out)
    return weeks


@app.context_processor
def inject_sla_helpers():
    return {
        'ticket_sla_view': ticket_sla_view,
        'localtime_fmt': format_local,
    }


@app.route('/admin/sla-calendar', methods=['GET', 'POST'])
@login_required
def admin_sla_calendar():
    if getattr(current_user, 'role', None) != 'admin':
        flash('Доступ запрещён', 'error')
        return redirect(url_for('ticket_list'))

    year = request.values.get('year', type=int) or datetime.now(get_runtime_timezone()).year
    selected_date_raw = (request.values.get('selected_date') or '').strip()
    try:
        selected_date = date.fromisoformat(selected_date_raw) if selected_date_raw else None
    except Exception:
        selected_date = None

    if request.method == 'POST':
        action = request.form.get('action', '').strip()
        try:
            if action == 'save_sla_settings':
                set_setting('system.timezone', (request.form.get('timezone') or SYSTEM_TZ_NAME).strip())
                set_setting('sla.timezone', (request.form.get('timezone') or SYSTEM_TZ_NAME).strip())
                set_setting('sla.work_start', (request.form.get('work_start') or '09:00').strip())
                set_setting('sla.work_end', (request.form.get('work_end') or '18:00').strip())
                set_setting('sla.first_response_minutes', str(request.form.get('first_response_minutes') or '60'))
                set_setting('sla.resolve_minutes', str(request.form.get('resolve_minutes') or '1440'))
                set_setting('sla.pause_statuses', (request.form.get('pause_statuses') or '').strip())
                set_setting('sla.workdays', ','.join(request.form.getlist('workdays') or ['0','1','2','3','4']))
                db.session.commit()
                flash('Настройки SLA сохранены', 'success')
            elif action == 'import_calendar':
                imported = 0
                for item in fetch_production_calendar_year(year):
                    d = date.fromisoformat(item['date'])
                    row = WorkCalendarDay.query.filter_by(date=d).first()
                    if not row:
                        row = WorkCalendarDay(date=d)
                        db.session.add(row)
                    row.day_type = 'workday' if item.get('is_workday') else 'weekend'
                    row.manual_override = False
                    if not row.name:
                        row.name = None
                    imported += 1
                db.session.commit()
                flash(f'Календарь РФ за {year} импортирован: {imported} дней', 'success')
            elif action == 'set_day_type':
                d = date.fromisoformat(request.form.get('date'))
                row = WorkCalendarDay.query.filter_by(date=d).first()
                if not row:
                    row = WorkCalendarDay(date=d)
                    db.session.add(row)
                row.day_type = request.form.get('day_type') or 'workday'
                row.name = (request.form.get('name') or '').strip() or None
                row.manual_override = True
                db.session.commit()
                flash('День календаря обновлён', 'success')
            elif action == 'recalculate_open_tickets':
                cnt = 0
                for t in SupportTicket.query.filter(SupportTicket.is_resolved == False).all():
                    recalc_ticket_sla(t)
                    cnt += 1
                db.session.commit()
                flash(f'Пересчитан SLA для {cnt} открытых заявок', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка SLA/календаря: {e}', 'error')
        return redirect(url_for('admin_sla_calendar', year=year, selected_date=(selected_date.isoformat() if selected_date else None)))

    months = []
    month_titles = ['Январь','Февраль','Март','Апрель','Май','Июнь','Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']
    day_rows = WorkCalendarDay.query.filter(
        WorkCalendarDay.date >= date(year, 1, 1),
        WorkCalendarDay.date <= date(year, 12, 31)
    ).all()
    row_map = {r.date: r for r in day_rows}
    stats = {'workday': 0, 'weekend': 0, 'holiday': 0, 'short_day': 0, 'manual': 0}
    for r in day_rows:
        stats[r.day_type] = stats.get(r.day_type, 0) + 1
        if getattr(r, 'manual_override', False):
            stats['manual'] += 1
    if selected_date is None:
        selected_date = datetime.now(get_runtime_timezone()).date()
    selected_row = row_map.get(selected_date)
    if selected_date.year != year:
        selected_date = date(year, 1, 1)
        selected_row = row_map.get(selected_date)
    for m in range(1,13):
        months.append({'num': m, 'title': month_titles[m-1], 'weeks': calendar_month_view(year, m)})
    return render_template('admin_sla_calendar.html', year=year, months=months, cfg=get_sla_settings(), selected_date=selected_date, selected_row=selected_row, cal_stats=stats)

@app.route('/admin/settings', methods=['GET', 'POST'])
@login_required
def admin_settings():
    if getattr(current_user, 'role', None) == 'client' or getattr(current_user, 'role', None) != 'admin':
        flash("Доступ запрещён", "error")
        # Операторам админка закрыта — отправляем в канбан/заявки
        if isinstance(current_user, User):
            return redirect(url_for('kanban'))
        return redirect(url_for('ticket_list'))

    # Активная вкладка админки
    # Управление пользователями вынесено на отдельную страницу /admin/users
    tab = (request.args.get('tab') or 'bitrix').strip().lower()
    if tab in ('users', 'endusers'):
        return redirect(url_for('admin_users'))

    if request.method == 'POST':
        action = request.form.get('action')

        # --- Mail / parser ---
        if action == 'run_mail_parser_now':
            started = start_mail_check_async()
            if started:
                flash("Проверка почты запущена (в фоне). Обнови страницу через пару секунд.", "success")
            else:
                flash("Проверка почты уже выполняется.", "info")
            return redirect(url_for('admin_settings', tab='mail'))

        if action == 'save_mail':
            # Пока храним настройки только для отображения в админке.
            # Парсер использует IMAP_* из .env (это можно будет связать позже).
            try:
                set_setting('MAIL_SERVER', request.form.get('MAIL_SERVER', '').strip())
                set_setting('MAIL_PORT', request.form.get('MAIL_PORT', '').strip())
                set_setting('MAIL_USE_TLS', request.form.get('MAIL_USE_TLS', 'False'))
                set_setting('MAIL_USERNAME', request.form.get('MAIL_USERNAME', '').strip())
                pwd = request.form.get('MAIL_PASSWORD', '').strip()
                if pwd:
                    set_setting('MAIL_PASSWORD', pwd)
                db.session.commit()
                flash("Почтовые настройки сохранены.", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Ошибка сохранения почтовых настроек: {e}", "error")
            return redirect(url_for('admin_settings', tab='mail'))

        # --- Notifications settings ---
        if action == 'save_notifications_settings':
            try:
                s = NotificationGlobalSettings.get_or_create()
                s.enabled = bool(request.form.get('notif_enabled') == 'on')
                s.enabled_for_operators = bool(request.form.get('notif_ops') == 'on')
                s.enabled_for_clients = bool(request.form.get('notif_clients') == 'on')
                s.event_assigned = bool(request.form.get('event_assigned') == 'on')
                s.event_customer_reply = bool(request.form.get('event_customer_reply') == 'on')
                s.event_status = bool(request.form.get('event_status') == 'on')
                db.session.commit()
                flash_msg('Настройки уведомлений сохранены', 'success')
            except Exception as e:
                db.session.rollback()
                flash_msg(f'Ошибка сохранения настроек уведомлений: {e}', 'error')
            return redirect(url_for('admin_settings', tab='notifications'))

        # Управление пользователями перенесено на /admin/users
        if action in {
            'reset_enduser_password', 'toggle_enduser_status', 'delete_enduser',
            'reset_operator_password', 'add_operator', 'edit_operator', 'delete_operator'
        }:
            flash("Управление пользователями перенесено в раздел 'Пользователи'.", "info")
            return redirect(url_for('admin_users'))

            try:
                dept_ids = [int(x) for x in department_ids if str(x).isdigit()]
                departments = Department.query.filter(Department.id.in_(dept_ids)).all()
                if not departments:
                    flash("Выбранные отделы не существуют", "error")
                    return redirect(url_for('admin_settings', tab='users'))

                user.username = username
                user.name = name
                user.last_name = last_name or None
                user.patronymic = patronymic or None
                user.email = email
                user.phone = phone or None
                user.role = 'operator'
                user.department_id = departments[0].id
                user.departments = departments
                db.session.commit()
                flash("Оператор обновлён", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Ошибка при обновлении оператора: {e}", "error")

            return redirect(url_for('admin_settings', tab='users'))

        # --- Directories: Departments ---
        if action == 'add_department':
            dept_name = request.form.get('dept_name', '').strip()
            if not dept_name:
                flash("Название отдела обязательно", "error")
                return redirect(url_for('admin_settings', tab='departments'))
            if Department.query.filter_by(name=dept_name).first():
                flash("Такой отдел уже существует", "error")
                return redirect(url_for('admin_settings', tab='departments'))

            dept = Department(name=dept_name)
            db.session.add(dept)
            db.session.commit()
            if not BitrixSettings.query.filter_by(department=dept.name).first():
                db.session.add(BitrixSettings(department=dept.name))
                db.session.commit()
            flash("Отдел добавлен", "success")
            return redirect(url_for('admin_settings', tab='departments'))

        if action == 'rename_department':
            dept_id = request.form.get('dept_id')
            new_name = request.form.get('new_name', '').strip()
            dept = Department.query.get(dept_id) if dept_id else None
            if not dept or not new_name:
                flash("Отдел/название не найдено", "error")
                return redirect(url_for('admin_settings', tab='departments'))
            if Department.query.filter(Department.name == new_name, Department.id != dept.id).first():
                flash("Отдел с таким названием уже существует", "error")
                return redirect(url_for('admin_settings', tab='departments'))

            old_name = dept.name
            dept.name = new_name
            # Bitrix settings переименуем
            bs = BitrixSettings.query.filter_by(department=old_name).first()
            if bs:
                bs.department = new_name
            db.session.commit()
            flash("Отдел переименован", "success")
            return redirect(url_for('admin_settings', tab='departments'))

        if action == 'delete_department':
            dept_id = request.form.get('dept_id')
            dept = Department.query.get(dept_id) if dept_id else None
            if not dept:
                flash("Отдел не найден", "error")
                return redirect(url_for('admin_settings', tab='departments'))

            has_tickets = SupportTicket.query.filter_by(department_id=dept.id).first() is not None
            has_users_primary = User.query.filter_by(department_id=dept.id).first() is not None
            has_users_m2m = db.session.execute(
                db.text("SELECT 1 FROM user_departments WHERE department_id = :dept_id LIMIT 1"),
                {"dept_id": dept.id}
            ).first() is not None

            if has_tickets or has_users_primary or has_users_m2m:
                flash("Нельзя удалить отдел: он используется в заявках или назначен операторам.", "error")
                return redirect(url_for('admin_settings', tab='departments'))

            try:
                bs = BitrixSettings.query.filter_by(department=dept.name).first()
                if bs:
                    db.session.delete(bs)
                db.session.delete(dept)
                db.session.commit()
                flash("Отдел удалён", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Ошибка при удалении отдела: {e}", "error")

            return redirect(url_for('admin_settings', tab='departments'))

        # --- Directories: Ticket Categories ---
        if action == 'add_ticket_category':
            code = request.form.get('cat_code', '').strip()
            name = request.form.get('cat_name', '').strip()
            if not code or not name:
                flash("Код и название обязательны", "error")
                return redirect(url_for('admin_settings', tab='categories'))
            if TicketCategory.query.filter_by(code=code).first() or TicketCategory.query.filter_by(name=name).first():
                flash("Категория с таким кодом/названием уже существует", "error")
                return redirect(url_for('admin_settings', tab='categories'))
            db.session.add(TicketCategory(code=code, name=name, sort_order=0, is_active=True))
            db.session.commit()
            flash("Категория добавлена", "success")
            return redirect(url_for('admin_settings', tab='categories'))

        if action == 'edit_ticket_category':
            cat_id = request.form.get('cat_id')
            cat = TicketCategory.query.get(cat_id) if cat_id else None
            if not cat:
                flash("Категория не найдена", "error")
                return redirect(url_for('admin_settings', tab='categories'))
            code = request.form.get('cat_code', '').strip()
            name = request.form.get('cat_name', '').strip()
            is_active = True if request.form.get('is_active') == '1' else False
            if not code or not name:
                flash("Код и название обязательны", "error")
                return redirect(url_for('admin_settings', tab='categories'))
            if TicketCategory.query.filter(TicketCategory.code == code, TicketCategory.id != cat.id).first():
                flash("Код уже используется", "error")
                return redirect(url_for('admin_settings', tab='categories'))
            if TicketCategory.query.filter(TicketCategory.name == name, TicketCategory.id != cat.id).first():
                flash("Название уже используется", "error")
                return redirect(url_for('admin_settings', tab='categories'))
            cat.code = code
            cat.name = name
            cat.is_active = is_active
            db.session.commit()
            flash("Категория обновлена", "success")
            return redirect(url_for('admin_settings', tab='categories'))

        if action == 'delete_ticket_category':
            cat_id = request.form.get('cat_id')
            cat = TicketCategory.query.get(cat_id) if cat_id else None
            if not cat:
                flash("Категория не найдена", "error")
                return redirect(url_for('admin_settings', tab='categories'))
            used = SupportTicket.query.filter_by(category_id=cat.id).first() is not None
            if used:
                flash("Нельзя удалить категорию: она используется в заявках.", "error")
                return redirect(url_for('admin_settings', tab='categories'))
            db.session.delete(cat)
            db.session.commit()
            flash("Категория удалена", "success")
            return redirect(url_for('admin_settings', tab='categories'))

        
        # --- Directories: Tags ---
        if action == 'add_tag':
            name = request.form.get('tag_name', '').strip()
            color = request.form.get('tag_color', '').strip()
            if not name:
                flash("Название тега обязательно", "error")
                return redirect(url_for('admin_settings', tab='tags'))
            if Tag.query.filter_by(name=name).first():
                flash("Такой тег уже существует", "error")
                return redirect(url_for('admin_settings', tab='tags'))
            db.session.add(Tag(name=name, color=color or None, is_active=True))
            db.session.commit()
            flash("Тег добавлен", "success")
            return redirect(url_for('admin_settings', tab='tags'))

        if action == 'edit_tag':
            tag_id = request.form.get('tag_id')
            tag = Tag.query.get(tag_id) if tag_id else None
            if not tag:
                flash("Тег не найден", "error")
                return redirect(url_for('admin_settings', tab='tags'))
            name = request.form.get('tag_name', '').strip()
            color = request.form.get('tag_color', '').strip()
            is_active = True if request.form.get('is_active') == '1' else False
            if not name:
                flash("Название тега обязательно", "error")
                return redirect(url_for('admin_settings', tab='tags'))
            if Tag.query.filter(Tag.name == name, Tag.id != tag.id).first():
                flash("Название уже используется", "error")
                return redirect(url_for('admin_settings', tab='tags'))
            tag.name = name
            tag.color = color or None
            tag.is_active = is_active
            db.session.commit()
            flash("Тег обновлён", "success")
            return redirect(url_for('admin_settings', tab='tags'))

        if action == 'delete_tag':
            tag_id = request.form.get('tag_id')
            tag = Tag.query.get(tag_id) if tag_id else None
            if not tag:
                flash("Тег не найден", "error")
                return redirect(url_for('admin_settings', tab='tags'))
            used = db.session.execute(
                db.text("SELECT 1 FROM ticket_tags WHERE tag_id = :tag_id LIMIT 1"),
                {"tag_id": tag.id}
            ).first() is not None
            if used:
                flash("Нельзя удалить тег: он используется в заявках. Сначала уберите его из заявок.", "error")
                return redirect(url_for('admin_settings', tab='tags'))
            db.session.delete(tag)
            db.session.commit()
            flash("Тег удалён", "success")
            return redirect(url_for('admin_settings', tab='tags'))

        # --- System settings ---
        if action in ['set_default_intake_department','save_system']:
	        dept_id = request.form.get('default_intake_department_id', '').strip()
	        if dept_id.isdigit() and Department.query.get(int(dept_id)):
	            set_setting('default_intake_department_id', dept_id)
	        else:
	            flash_msg("Некорректный отдел", "error")
	            return redirect(url_for('admin_settings', tab='system'))

	        mode = (request.form.get('profile_enforcement_mode', '') or '').strip().lower()
	        if mode in ('strict', 'soft', 'off'):
	            set_setting('profile_enforcement_mode', mode)
	        else:
	            set_setting('profile_enforcement_mode', get_profile_enforcement_mode())

	        db.session.commit()
	        flash_msg("Системные настройки сохранены", "success")
	        return redirect(url_for('admin_settings', tab='system'))

    # --- Load data for tabs ---
    operators = User.query.filter(User.role.in_(['admin', 'operator'])).order_by(User.role.desc(), User.username.asc()).all()
    endusers = User.query.filter_by(role='client').order_by(User.created_at.desc()).all()
    departments = Department.query.order_by(Department.name).all()
    ticket_categories = TicketCategory.query.order_by(TicketCategory.sort_order, TicketCategory.name).all()
    tags = Tag.query.order_by(Tag.name).all()
    bitrix_settings = BitrixSettings.query.order_by(BitrixSettings.department).all()

    system_settings = {
	        'default_intake_department_id': get_default_intake_department_id() or '',
	        'profile_enforcement_mode': get_profile_enforcement_mode(),
    }



    # --- Mail settings for template (чтобы tab=mail не падал) ---
    from types import SimpleNamespace
    # Настройки храним в Settings (для UI). Парсер сейчас использует IMAP_* из .env.
    mail_settings = SimpleNamespace(
        MAIL_SERVER=get_setting('MAIL_SERVER', '') or '',
        MAIL_PORT=int(get_setting('MAIL_PORT', '993') or '993'),
        MAIL_USERNAME=get_setting('MAIL_USERNAME', '') or '',
        MAIL_PASSWORD=get_setting('MAIL_PASSWORD', '') or '',
        MAIL_USE_TLS=(get_setting('MAIL_USE_TLS', 'False') or 'False').lower() == 'true',
        MAIL_USE_SSL=True,
        MAIL_FOLDER='INBOX',
        MAIL_CHECK_INTERVAL=60,
    )

    return render_template(
        'admin_settings.html',
        tab=tab,
        operators=operators,
        endusers=endusers,
        departments=departments,
        ticket_categories=ticket_categories,
        tags=tags,
        bitrix_settings=bitrix_settings,
        system_settings=system_settings,
        mail_settings=mail_settings,
        mail_parser_state=MAIL_PARSER_STATE,
        notif_settings=NotificationGlobalSettings.get_or_create(),
        USER_ROLES=USER_ROLES
    )


@app.route('/admin/users', methods=['GET', 'POST'])
@login_required
def admin_users():
    """Админка: пользователи (единая таблица users, роли client/operator/admin).

    Поток:
      - пользователь может зарегистрироваться сам -> role=client, email_verified=False
      - админ может:
          * создавать любого пользователя
          * назначать роль/отделы
          * включать флажок "email подтверждён" (пропуск верификации)
          * редактировать/удалять
    """
    is_admin = (getattr(current_user, 'role', None) == 'admin')
    can_manage_users = bool(is_admin or is_tp_operator(current_user))

    if not can_manage_users:
        flash("Доступ запрещён", "error")
        return redirect(url_for('ticket_list'))

    departments = Department.query.order_by(Department.name.asc()).all()

    def _render_page(open_add_modal: bool = False, add_errors: dict | None = None, add_data: dict | None = None):
        """Единый рендер страницы (чтобы на ошибках формы не терять состояние и открывать модалку)."""

        q = (request.args.get('q') or '').strip().lower()
        role_filter = (request.args.get('role') or '').strip().lower()
        if role_filter not in ('', 'client', 'operator', 'admin'):
            role_filter = ''
        sort = (request.args.get('sort') or 'created_at').strip().lower()
        direction = (request.args.get('dir') or 'desc').strip().lower()
        if direction not in ('asc', 'desc'):
            direction = 'desc'

        # пагинация (Jira/Zendesk-style lists)
        try:
            page = int(request.args.get('page') or 1)
        except Exception:
            page = 1
        try:
            per_page = int(request.args.get('per_page') or 25)
        except Exception:
            per_page = 25
        if per_page not in (10, 25, 50, 100):
            per_page = 25
        if page < 1:
            page = 1

        base = User.query
        if q:
            base = base.filter(
                db.or_(
                    db.func.lower(User.username).contains(q),
                    db.func.lower(User.email).contains(q),
                    db.func.lower(User.name).contains(q),
                    db.func.lower(User.last_name).contains(q)
                )
            )

        if role_filter:
            base = base.filter(User.role == role_filter)

        # сортировки: id, role, created_at
        if sort == 'id':
            order_col = User.id
        elif sort == 'role':
            # Желаемый порядок ролей: admin -> operator -> client
            order_col = db.case(
                (User.role == 'admin', 0),
                (User.role == 'operator', 1),
                else_=2
            )
        else:
            sort = 'created_at'
            order_col = User.created_at

        if direction == 'asc':
            base = base.order_by(order_col.asc(), User.id.asc())
        else:
            base = base.order_by(order_col.desc(), User.id.desc())

        # paginate
        try:
            pagination = base.paginate(page=page, per_page=per_page, error_out=False)
            users = pagination.items
        except Exception:
            pagination = None
            users = base.all()

        return render_template(
            'admin_users.html',
            users=users,
            pagination=pagination,
            departments=departments,
            q=q,
            role_filter=role_filter,
            sort=sort,
            direction=direction,
            page=page,
            per_page=per_page,
            total_departments=len(departments),
            USER_ROLES=USER_ROLES,
            is_admin=is_admin,
            can_manage_users=can_manage_users,
            # для модалки создания
            open_add_modal=open_add_modal,
            add_errors=add_errors,
            add_data=add_data,
        )

    # сохраняем сортировку/поиск при POST -> redirect обратно
    _back_args = {
        'q': request.args.get('q') or None,
        'role': request.args.get('role') or None,
        'sort': request.args.get('sort') or None,
        'dir': request.args.get('dir') or None,
        'page': request.args.get('page') or None,
        'per_page': request.args.get('per_page') or None,
    }

    def _back():
        return redirect(url_for('admin_users', **_back_args))

    if request.method == 'POST':
        action = (request.form.get('action') or '').strip()

        # --- Универсальные действия ---
        if action == 'reset_user_password':
            if not is_admin:
                flash("Недостаточно прав", "error")
                return _back()
            user_id = request.form.get('user_id')
            u = db.session.get(User, user_id) if user_id else None
            if u and u.id != current_user.id:
                new_pass = 'new_password'
                u.password = generate_password_hash(new_pass, method='pbkdf2:sha256')
                db.session.commit()
                flash(f"Пароль сброшен на: {new_pass}", "success")
            return _back()

        if action == 'delete_user':
            if not is_admin:
                flash("Недостаточно прав", "error")
                return _back()
            user_id = request.form.get('user_id')
            u = db.session.get(User, user_id) if user_id else None
            if not u:
                flash("Пользователь не найден", "error")
                return _back()
            if u.id == current_user.id:
                flash("Нельзя удалить текущего пользователя", "error")
                return _back()
            if u.role == 'admin':
                flash("Нельзя удалить администратора", "error")
                return _back()
            try:
                db.session.delete(u)
                db.session.commit()
                flash("Пользователь удалён", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Ошибка удаления: {e}", "error")
            return _back()

        # --- Новый унифицированный CRUD ---
        if action == 'add_user':
            add_errors = {}
            role = (request.form.get('role') or 'client').strip().lower()
            if role not in ('client', 'operator', 'admin'):
                role = 'client'

            # Операторы ТП 1/2 линии могут добавлять пользователей только как клиентов
            if not is_admin:
                role = 'client'

            email = (request.form.get('email') or '').strip().lower() or None
            username = (request.form.get('username') or '').strip() or (email or '')
            password = (request.form.get('password') or '').strip()

            # базовые поля + валидация (анти-ссылки/анти-мат/форматы)
            name_raw = request.form.get('name') or ''
            last_name_raw = request.form.get('last_name') or ''
            patronymic_raw = request.form.get('patronymic') or ''
            phone_raw = request.form.get('phone') or ''
            organization_raw = request.form.get('organization') or ''
            position = _norm_text(request.form.get('position') or '') or None
            inn_raw = request.form.get('inn') or ''
            address_raw = request.form.get('address') or ''

            # ФИО: для клиента — обязательное (чтобы не было мусора)
            if role == 'client':
                ok, last_name_v = validate_person_part(last_name_raw)
                if not ok:
                    add_errors['last_name'] = last_name_v
                ok, name_v = validate_person_part(name_raw)
                if not ok:
                    add_errors['name'] = name_v
                ok, patr_v = validate_person_part(patronymic_raw)
                if not ok:
                    add_errors['patronymic'] = patr_v
                last_name, name, patronymic = last_name_v, name_v, patr_v
            else:
                # для сотрудников — валидируем только если введено
                last_name = _norm_text(last_name_raw) or None
                name = _norm_text(name_raw) or None
                patronymic = _norm_text(patronymic_raw) or None
                for label, raw, setter in (
                    ("Фамилия", last_name_raw, 'last_name'),
                    ("Имя", name_raw, 'name'),
                    ("Отчество", patronymic_raw, 'patronymic'),
                ):
                    if _norm_text(raw):
                        ok, v = validate_person_part(raw)
                        if not ok:
                            add_errors[setter] = v
                        if setter == 'last_name':
                            last_name = v
                        elif setter == 'name':
                            name = v
                        else:
                            patronymic = v

            ok, phone_v = normalize_phone(phone_raw)
            if not ok:
                add_errors['phone'] = phone_v
            phone = phone_v or None

            ok, org_v = validate_org(organization_raw)
            if not ok:
                add_errors['organization'] = org_v
            organization = org_v or None

            ok, inn_v = validate_inn_ru(inn_raw, required=False)
            if not ok:
                add_errors['inn'] = inn_v
            inn = inn_v or None

            ok, addr_v = validate_address(address_raw, required=False)
            if not ok:
                add_errors['address'] = addr_v
            address = addr_v or None
            email_verified = bool(request.form.get('email_verified') == 'on')

            department_ids = request.form.getlist('department_ids')

            if not is_admin:
                department_ids = []

            if not username:
                add_errors['username'] = "Логин обязателен"
            if not email:
                add_errors['email'] = "Email обязателен"
            if not password:
                add_errors['password'] = "Пароль обязателен"
            elif len(password) < 6:
                add_errors['password'] = "Минимум 6 символов"

            if User.query.filter_by(username=username).first():
                add_errors['username'] = "Логин уже занят"
            if User.query.filter(db.func.lower(User.email) == email.lower()).first():
                add_errors['email'] = "Email уже используется"

            if add_errors:
                flash("Проверьте поля формы", "error")
                # сохраняем введённые значения, чтобы не раздражать пользователя
                add_data = {
                    'role': role,
                    'username': request.form.get('username') or '',
                    'email': request.form.get('email') or '',
                    'last_name': request.form.get('last_name') or '',
                    'name': request.form.get('name') or '',
                    'patronymic': request.form.get('patronymic') or '',
                    'phone': request.form.get('phone') or '',
                    'organization': request.form.get('organization') or '',
                    'position': request.form.get('position') or '',
                    'inn': request.form.get('inn') or '',
                    'address': request.form.get('address') or '',
                }
                return _render_page(open_add_modal=True, add_errors=add_errors, add_data=add_data)

            try:
                u = User(
                    username=username,
                    name=name,
                    last_name=last_name,
                    patronymic=patronymic,
                    email=email,
                    password=generate_password_hash(password, method='pbkdf2:sha256'),
                    role=role,
                    email_verified=email_verified if role == 'client' else True,
                    phone=phone,
                    organization=organization,
                    position=position,
                    inn=inn,
                    address=address,
                )
                db.session.add(u)
                db.session.flush()

                # отделы актуальны только для operator/admin
                if role in ('operator', 'admin') and department_ids:
                    dept_ids = [int(x) for x in department_ids if str(x).isdigit()]
                    depts = Department.query.filter(Department.id.in_(dept_ids)).all()
                    if depts:
                        u.department_id = depts[0].id
                        u.departments = depts

                db.session.commit()
                flash("Пользователь создан", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Ошибка при создании: {e}", "error")
            return _back()

        if action == 'edit_user':
            user_id = request.form.get('user_id')
            u = db.session.get(User, user_id) if user_id else None
            if not u:
                flash("Пользователь не найден", "error")
                return _back()

            # Операторы ТП 1/2 линии могут редактировать только клиентов (без изменения роли/пароля)
            if not is_admin:
                if (u.role or 'client') != 'client':
                    flash("Недостаточно прав для редактирования этого пользователя", "error")
                    return _back()

            role = (request.form.get('role') or u.role or 'client').strip().lower()
            if role not in ('client', 'operator', 'admin'):
                role = u.role

            if not is_admin:
                role = 'client'

            username = (request.form.get('username') or '').strip()
            email = (request.form.get('email') or '').strip().lower() or None
            if not username:
                flash("Логин обязателен", "error")
                return _back()
            if not email:
                flash("Email обязателен", "error")
                return _back()

            if User.query.filter(User.username == username, User.id != u.id).first():
                flash("Логин уже занят", "error")
                return _back()
            if User.query.filter(db.func.lower(User.email) == email.lower(), User.id != u.id).first():
                flash("Email уже используется", "error")
                return _back()

            # Валидация полей
            name_raw = request.form.get('name') or ''
            last_name_raw = request.form.get('last_name') or ''
            patronymic_raw = request.form.get('patronymic') or ''
            phone_raw = request.form.get('phone') or ''
            organization_raw = request.form.get('organization') or ''
            position = _norm_text(request.form.get('position') or '') or None
            inn_raw = request.form.get('inn') or ''
            address_raw = request.form.get('address') or ''

            # Если пользователь (или мы) делаем его клиентом — требуем корректное ФИО
            ln = nm = pt = None
            if role == 'client':
                ok, ln = validate_person_part(last_name_raw)
                if not ok:
                    flash(f"Фамилия: {ln}", "error")
                    return _back()
                ok, nm = validate_person_part(name_raw)
                if not ok:
                    flash(f"Имя: {nm}", "error")
                    return _back()
                ok, pt = validate_person_part(patronymic_raw)
                if not ok:
                    flash(f"Отчество: {pt}", "error")
                    return _back()
            else:
                # для сотрудников валидируем только если заполнено
                for label, raw in (("Фамилия", last_name_raw), ("Имя", name_raw), ("Отчество", patronymic_raw)):
                    if _norm_text(raw):
                        ok, v = validate_person_part(raw)
                        if not ok:
                            flash(f"{label}: {v}", "error")
                            return _back()

            ok, phone_v = normalize_phone(phone_raw)
            if not ok:
                flash(phone_v, "error")
                return _back()

            ok, org_v = validate_org(organization_raw)
            if not ok:
                flash(f"Организация: {org_v}", "error")
                return _back()

            ok, inn_v = validate_inn_ru(inn_raw, required=False)
            if not ok:
                flash(f"ИНН: {inn_v}", "error")
                return _back()

            ok, addr_v = validate_address(address_raw, required=False)
            if not ok:
                flash(f"Адрес: {addr_v}", "error")
                return _back()

            try:
                u.username = username
                u.email = email
                u.role = role

                if role == 'client':
                    u.last_name = ln
                    u.name = nm
                    u.patronymic = pt
                else:
                    u.name = _norm_text(name_raw) or None
                    u.last_name = _norm_text(last_name_raw) or None
                    u.patronymic = _norm_text(patronymic_raw) or None
                u.phone = phone_v or None
                u.organization = org_v or None
                u.position = position
                u.inn = inn_v or None
                u.address = addr_v or None

                # пароль (не обязателен)
                new_pass = (request.form.get('new_password') or '').strip()
                if new_pass and is_admin:
                    u.password = generate_password_hash(new_pass, method='pbkdf2:sha256')

                # email verified — имеет смысл только для клиентов
                if role == 'client':
                    u.email_verified = bool(request.form.get('email_verified') == 'on')
                else:
                    u.email_verified = True

                department_ids = request.form.getlist('department_ids')

                # Операторы ТП не управляют отделами
                if not is_admin:
                    department_ids = []

                # Назначаем отделы только для operator/admin
                if role in ('operator', 'admin') and department_ids:
                    dept_ids = [int(x) for x in department_ids if str(x).isdigit()]
                    depts = Department.query.filter(Department.id.in_(dept_ids)).all()
                    if depts:
                        u.department_id = depts[0].id
                        u.departments = depts
                else:
                    u.department_id = None
                    u.departments = []

                db.session.commit()
                flash("Изменения сохранены", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Ошибка сохранения: {e}", "error")
            return _back()

        # --- Операторы/админы ---
        if action == 'add_operator':
            username = (request.form.get('username') or '').strip()
            name = (request.form.get('name') or '').strip()
            last_name = (request.form.get('last_name') or '').strip() or None
            patronymic = (request.form.get('patronymic') or '').strip() or None
            email = (request.form.get('email') or '').strip().lower()
            password = (request.form.get('password') or '').strip()
            department_ids = request.form.getlist('department_ids')

            if not is_admin:
                department_ids = []
            role = (request.form.get('role') or 'operator').strip()
            if role not in ('operator', 'admin'):
                role = 'operator'

            if not username or not name or not email or not password:
                flash("Логин, имя, email и пароль обязательны", "error")
                return redirect(url_for('admin_users'))
            if not department_ids:
                flash("Нужно выбрать хотя бы один отдел", "error")
                return redirect(url_for('admin_users'))
            if User.query.filter_by(username=username).first():
                flash("Логин уже занят", "error")
                return redirect(url_for('admin_users'))
            if User.query.filter(db.func.lower(User.email) == email.lower()).first():
                flash("Email уже используется", "error")
                return redirect(url_for('admin_users'))

            try:
                dept_ids = [int(x) for x in department_ids if str(x).isdigit()]
                depts = Department.query.filter(Department.id.in_(dept_ids)).all()
                if not depts:
                    flash("Выбранные отделы не существуют", "error")
                    return redirect(url_for('admin_users'))

                u = User(
                    username=username,
                    name=name,
                    last_name=last_name,
                    patronymic=patronymic,
                    email=email,
                    password=generate_password_hash(password, method='pbkdf2:sha256'),
                    role=role,
                    department_id=depts[0].id,
                    email_verified=True,
                )
                db.session.add(u)
                db.session.flush()
                u.departments = depts
                db.session.commit()
                flash("Пользователь добавлен", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Ошибка при создании: {e}", "error")
            return redirect(url_for('admin_users'))

        if action == 'edit_operator':
            user_id = request.form.get('user_id')
            u = db.session.get(User, user_id) if user_id else None
            if not u:
                flash("Пользователь не найден", "error")
                return redirect(url_for('admin_users'))

            username = (request.form.get('username') or '').strip()
            name = (request.form.get('name') or '').strip()
            last_name = (request.form.get('last_name') or '').strip() or None
            patronymic = (request.form.get('patronymic') or '').strip() or None
            email = (request.form.get('email') or '').strip().lower() or None
            phone = (request.form.get('phone') or '').strip() or None
            role = (request.form.get('role') or u.role or 'operator').strip()
            if role not in ('operator', 'admin', 'client'):
                role = u.role
            department_ids = request.form.getlist('department_ids')

            if not is_admin:
                department_ids = []

            if not username or not name:
                flash("Логин и имя обязательны", "error")
                return redirect(url_for('admin_users'))
            if User.query.filter(User.username == username, User.id != u.id).first():
                flash("Логин уже занят", "error")
                return redirect(url_for('admin_users'))
            if email:
                if User.query.filter(db.func.lower(User.email) == email.lower(), User.id != u.id).first():
                    flash("Email уже используется", "error")
                    return redirect(url_for('admin_users'))

            try:
                u.username = username
                u.name = name
                u.last_name = last_name
                u.patronymic = patronymic
                u.email = email
                u.phone = phone
                u.role = role

                # Назначаем отделы только если пользователь — оператор/админ
                if u.role in ('operator', 'admin'):
                    if department_ids:
                        dept_ids = [int(x) for x in department_ids if str(x).isdigit()]
                        depts = Department.query.filter(Department.id.in_(dept_ids)).all()
                        if depts:
                            u.department_id = depts[0].id
                            u.departments = depts
                else:
                    u.department_id = None
                    u.departments = []

                db.session.commit()
                flash("Изменения сохранены", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Ошибка сохранения: {e}", "error")
            return redirect(url_for('admin_users'))

        # --- Клиенты ---
        if action == 'add_client':
            last_name = (request.form.get('last_name') or '').strip() or None
            name = (request.form.get('name') or '').strip() or None
            patronymic = (request.form.get('patronymic') or '').strip() or None
            email = (request.form.get('email') or '').strip().lower()
            password = (request.form.get('password') or '').strip()
            phone = (request.form.get('phone') or '').strip() or None
            organization = (request.form.get('organization') or '').strip() or None
            position = (request.form.get('position') or '').strip() or None
            inn = (request.form.get('inn') or '').strip() or None
            address = (request.form.get('address') or '').strip() or None

            if not email or not password:
                flash("Email и пароль обязательны", "error")
                return redirect(url_for('admin_users'))
            if User.query.filter(db.func.lower(User.email) == email.lower()).first():
                flash("Email уже используется", "error")
                return redirect(url_for('admin_users'))
            try:
                u = User(
                    username=email,
                    name=name,
                    last_name=last_name,
                    patronymic=patronymic,
                    email=email,
                    password=generate_password_hash(password, method='pbkdf2:sha256'),
                    phone=phone,
                    organization=organization,
                    position=position,
                    inn=inn,
                    address=address,
                    role='client',
                    email_verified=bool(request.form.get('email_verified') == 'on')
                )
                db.session.add(u)
                db.session.commit()
                flash("Клиент добавлен", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Ошибка при создании: {e}", "error")
            return redirect(url_for('admin_users'))

        if action == 'edit_client':
            user_id = request.form.get('user_id')
            u = db.session.get(User, user_id) if user_id else None
            if not u or u.role != 'client':
                flash("Пользователь не найден", "error")
                return redirect(url_for('admin_users'))

            email = (request.form.get('email') or '').strip().lower()
            if not email:
                flash("Email обязателен", "error")
                return redirect(url_for('admin_users'))
            if User.query.filter(db.func.lower(User.email) == email.lower(), User.id != u.id).first():
                flash("Email уже используется", "error")
                return redirect(url_for('admin_users'))

            try:
                u.name = (request.form.get('name') or '').strip() or None
                u.last_name = (request.form.get('last_name') or '').strip() or None
                u.patronymic = (request.form.get('patronymic') or '').strip() or None
                u.email = email
                u.username = email
                u.phone = (request.form.get('phone') or '').strip() or None
                u.organization = (request.form.get('organization') or '').strip() or None
                u.position = (request.form.get('position') or '').strip() or None
                u.inn = (request.form.get('inn') or '').strip() or None
                u.address = (request.form.get('address') or '').strip() or None
                u.email_verified = bool(request.form.get('email_verified') == 'on')
                db.session.commit()
                flash("Изменения сохранены", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Ошибка сохранения: {e}", "error")
            return redirect(url_for('admin_users'))

        flash("Неизвестное действие", "error")
        return redirect(url_for('admin_users'))

    # GET
    return _render_page()

@app.route('/admin/reset_password/<int:user_id>')
@login_required
def admin_reset_password(user_id):
    if current_user.role != 'admin':
        flash("Доступ запрещён", "error")
        return redirect(url_for('admin_settings'))
    user = User.query.get_or_404(user_id)
    if user.role == 'admin' and user.id != current_user.id:
        flash("Нельзя сбросить пароль другому админу", "error")
        return redirect(url_for('admin_settings'))
    new_pass = 'new_password'
    user.password = generate_password_hash(new_pass, method='pbkdf2:sha256')
    db.session.commit()
    flash(f"Пароль для {user.username} сброшен на: {new_pass}")
    return redirect(url_for('admin_settings', tab='users'))

@app.route('/admin/delete_operator/<int:user_id>')
@login_required
def admin_delete_operator(user_id):
    if current_user.role != 'admin':
        flash("Доступ запрещён", "error")
        return redirect(url_for('admin_settings'))
    user = User.query.get_or_404(user_id)
    if user.role == 'admin':
        flash("Нельзя удалить администратора", "error")
        return redirect(url_for('admin_settings'))
    db.session.delete(user)
    db.session.commit()
    flash(f"Оператор {user.username} удалён")
    return redirect(url_for('admin_settings', tab='users'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '').strip()
        password2 = request.form.get('password2', '').strip()

        if not email:
            flash("Email обязателен", "error")
        elif not password:
            flash("Пароль обязателен", "error")
        elif len(password) < 6:
            flash("Пароль должен содержать не менее 6 символов", "error")
        elif password != password2:
            flash("Пароли не совпадают", "error")
        elif User.query.filter(db.func.lower(User.email) == email.lower()).first():
            flash("Email уже зарегистрирован", "error")
        else:
            user = User(
                username=email,
                email=email,
                password=generate_password_hash(password, method='pbkdf2:sha256'),
                is_active=True,
                role='client',
                email_verified=False,
            )
            db.session.add(user)
            db.session.commit()

            send_email_verification(user)
            flash("Регистрация успешна! Проверьте почту для подтверждения.", "success")
            return redirect(url_for('login'))

        return render_template('register.html', email=email)

    return render_template('register.html')


@app.route('/tickets', methods=['GET', 'POST'])
@login_required
def ticket_list():
    is_client = getattr(current_user, 'role', None) == 'client'
    is_operator = (not is_client) and (
        current_user.role in ['operator', 'admin'] or (isinstance(current_user, User) and is_tp_operator(current_user))
    )
    is_admin = is_operator and current_user.role == 'admin'
    can_manage_spam = can_manage_spam_user(current_user)

    # === Персистентные настройки списка (как в Bitrix): фильтры/сортировка/кол-во на странице ===
    ui_state = None
    saved = {}
    if isinstance(current_user, User):
        try:
            ui_state = UserUIState.query.get(current_user.id)
            if ui_state is None:
                ui_state = UserUIState(user_id=current_user.id, data='{}')
                db.session.add(ui_state)
                db.session.commit()
            saved = ui_state.get('ticket_list', {}) or {}
        except Exception:
            saved = {}

    # Если пришёл clear=1 — сбрасываем сохранённые настройки
    if request.args.get('clear') == '1' and ui_state is not None:
        try:
            ui_state.set('ticket_list', {})
            db.session.commit()
            saved = {}
        except Exception:
            pass


    # === Получение ID отделов (не используется в этом роуте, но можно оставить) ===
    attention_dept = Department.query.filter_by(name="Требуется обработка").first()
    first_line_dept = Department.query.filter_by(name="1ая линия ТП").first()
    attention_id = attention_dept.id if attention_dept else None
    first_line_id = first_line_dept.id if first_line_dept else None

    # === Базовый запрос ===
    query = SupportTicket.query

    # Спам скрываем по умолчанию (если у пользователя нет прав управлять спамом)
    if not can_manage_spam:
        try:
            query = query.filter(SupportTicket.is_spam == False)
        except Exception:
            pass

    if is_client:
        query = query.filter(SupportTicket.email == current_user.email)
        page_title = "Заявки"

    # === Клиентская фильтрация (минимальная) ===
    # Раньше фильтры работали только у операторов, из-за чего у клиента выбор статуса
    # (например "Принята") не давал результата.
    if is_client:
        ticket_id = request.args.get('id', '').strip()
        subject = request.args.get('subject', '').strip()
        statuses = request.args.getlist('status')

        if ticket_id.isdigit():
            query = query.filter(SupportTicket.id == int(ticket_id))
        if subject:
            query = query.filter(SupportTicket.subject.ilike(f'%{subject}%'))
        if statuses:
            query = query.filter(SupportTicket.status.in_(statuses))
    elif is_admin:
        page_title = "📋 Все обращения"
    else:
        # Оператор: обычно видит заявки своих отделов.
        # Но 1/2 линия ТП должна видеть ВСЕ заявки (включая делегированные в другие отделы).
        if isinstance(current_user, User) and is_tp_operator(current_user):
            page_title = "📋 Все обращения"
        else:
            dep_ids = user_department_ids(current_user)
            if dep_ids:
                # Включаем заявки:
                #  - основного отдела
                #  - а также заявки, где отдел указан как "дополнительный" (multi-delegate)
                try:
                    from helpdesk_app.models.base import ticket_shared_departments
                    shared_ticket_ids = db.session.query(ticket_shared_departments.c.ticket_id).filter(
                        ticket_shared_departments.c.department_id.in_(dep_ids)
                    )
                    query = query.filter(
                        db.or_(SupportTicket.department_id.in_(dep_ids), SupportTicket.id.in_(shared_ticket_ids))
                    )
                except Exception:
                    query = query.filter(SupportTicket.department_id.in_(dep_ids))
            else:
                # если оператору не назначены отделы — ничего не показываем
                query = query.filter(db.text("1=0"))
            page_title = "Заявки"

    # === Фильтрация и сортировка (оператор/админ) ===
    applied_sorting = False
    if is_operator:
        ticket_id = request.args.get('id', '').strip()
        subject = request.args.get('subject', '').strip()
        client_email = request.args.get('email', '').strip()
        client_inn = request.args.get('inn', '').strip()
        client_fio = request.args.get('fio', '').strip()
        statuses = request.args.getlist('status')
        departments = request.args.getlist('department')
        categories = request.args.getlist('category')
        sort = (request.args.get('sort') or '').strip()  # created_desc (default), updated_desc, created_asc, updated_asc
        group = (request.args.get('group') or '').strip()  # important_first
        indicators = request.args.getlist('indicator')  # orange/yellow/green/blue/black (multi)
        # sort_priority убран (приоритеты больше не используем)
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()

        # === Применяем сохранённые настройки, если параметры не переданы в URL ===
        # (Bitrix-поведение: выбранные фильтры/сортировка/показ сохраняются за пользователем)
        if saved:
            if not ticket_id and saved.get('id'):
                ticket_id = str(saved.get('id') or '').strip()
            if not subject and saved.get('subject'):
                subject = str(saved.get('subject') or '').strip()
            if not client_email and saved.get('email'):
                client_email = str(saved.get('email') or '').strip()
            if not client_inn and saved.get('inn'):
                client_inn = str(saved.get('inn') or '').strip()
            if not client_fio and saved.get('fio'):
                client_fio = str(saved.get('fio') or '').strip()

            if not statuses and saved.get('status'):
                statuses = list(saved.get('status') or [])
            if not departments and saved.get('department'):
                departments = list(saved.get('department') or [])
            if not indicators and saved.get('indicator'):
                indicators = list(saved.get('indicator') or [])

        # Нормализуем мультизначения (иногда прилетает строка)
        if isinstance(statuses, str):
            statuses = [s for s in statuses.split(',') if s.strip()]
        if isinstance(departments, str):
            departments = [s for s in departments.split(',') if s.strip()]
        if isinstance(indicators, str):
            indicators = [s for s in indicators.split(',') if s.strip()]
        # Приоритеты убраны. Оставлена только отметка "Критично" (ticket.priority == 'Критический').


        # Фильтр по ID
        if ticket_id.isdigit():
            query = query.filter(SupportTicket.id == int(ticket_id))

        # Фильтр по теме
        if subject:
            query = query.filter(SupportTicket.subject.ilike(f'%{subject}%'))

        # Фильтр по email клиента
        if client_email:
            query = query.filter(SupportTicket.email.ilike(f'%{client_email}%'))

        # Фильтр по ИНН
        if client_inn:
            query = query.filter(SupportTicket.inn.ilike(f'%{client_inn}%'))

        # Фильтр по ФИО/имени (ищем по полю name в заявке)
        if client_fio:
            query = query.filter(SupportTicket.name.ilike(f'%{client_fio}%'))

        # Фильтр по статусам
        if statuses:
            query = query.filter(SupportTicket.status.in_(statuses))

        # Фильтр по индикатору состояния (Bitrix-like) — мультивыбор
        # indicators: orange/yellow/green/blue/black
        if indicators:
            # Subquery: last message created_at per ticket
            tm = TicketMessage
            last_cte = db.session.query(
                tm.ticket_id.label('t_id'),
                func.max(tm.created_at).label('mx')
            ).group_by(tm.ticket_id).subquery('last_msg')

            tm_last = aliased(tm, name='tm_last')
            query = query.outerjoin(last_cte, last_cte.c.t_id == SupportTicket.id)
            query = query.outerjoin(tm_last, db.and_(tm_last.ticket_id == SupportTicket.id, tm_last.created_at == last_cte.c.mx))

            # Responsible expression (assigned_to -> locked_by -> created_by_operator)
            rid = db.func.coalesce(SupportTicket.assigned_to_id, SupportTicket.locked_by, SupportTicket.created_by_operator_id)

            # Closed predicate
            st_lower = db.func.lower(db.func.trim(SupportTicket.status))
            closed_pred = db.or_(
                st_lower.like('закры%'),
                st_lower.like('заверш%'),
                SupportTicket.is_resolved == True,
                SupportTicket.closed_at.isnot(None),
                SupportTicket.auto_closed_at.isnot(None),
                SupportTicket.marked_as_completed_at.isnot(None),
            )

            # No messages means treat as client-last (green)
            no_msg_pred = last_cte.c.mx.is_(None)
            last_is_op = tm_last.is_operator

            conds = []
            for ind in [i.strip() for i in indicators if (i or '').strip()]:
                if ind == 'black':
                    conds.append(closed_pred)
                elif ind == 'blue':
                    conds.append(db.and_(~closed_pred, no_msg_pred == False, last_is_op == True))
                elif ind == 'orange':
                    conds.append(db.and_(~closed_pred, (db.or_(no_msg_pred, last_is_op == False)), rid == current_user.id))
                elif ind == 'yellow':
                    conds.append(db.and_(~closed_pred, (db.or_(no_msg_pred, last_is_op == False)), rid.isnot(None), rid != current_user.id))
                elif ind == 'green':
                    conds.append(db.and_(~closed_pred, db.or_(rid.is_(None), rid == 0), db.or_(no_msg_pred, last_is_op == False)))
            if conds:
                query = query.filter(db.or_(*conds))

        # Фильтр по отделам
        if departments:
            dept_ids = db.session.query(Department.id).filter(Department.name.in_(departments)).all()
            dept_ids = [id for (id,) in dept_ids]
            if dept_ids:
                query = query.filter(SupportTicket.department_id.in_(dept_ids))

        # Фильтр по категориям (мультивыбор)
        if categories:
            try:
                cat_ids = db.session.query(TicketCategory.id).filter(TicketCategory.name.in_(categories)).all()
                cat_ids = [cid for (cid,) in cat_ids]
                if cat_ids:
                    query = query.filter(SupportTicket.category_id.in_(cat_ids))
            except Exception:
                pass

        # Сортировка по приоритету убрана

        # Фильтр по дате "от"
        if date_from:
            try:
                dt = datetime.strptime(date_from, '%Y-%m-%d')
                query = query.filter(SupportTicket.created_at >= dt)
            except ValueError:
                pass

        # Фильтр по дате "до"
        if date_to:
            try:
                dt = datetime.strptime(date_to, '%Y-%m-%d')
                dt = dt.replace(hour=23, minute=59, second=59)
                query = query.filter(SupportTicket.created_at <= dt)
            except ValueError:
                pass

    # === Сортировка (Bitrix-like): по любому столбцу таблицы ===
    # Параметры: sort=<field>&dir=asc|desc
    # Для обратной совместимости понимаем старый sort=created_desc/updated_desc...
    current_sort = (request.args.get('sort') or '').strip() or (saved.get('sort') if isinstance(saved, dict) else None) or 'created'
    current_dir = (request.args.get('dir') or '').strip().lower() or (saved.get('dir') if isinstance(saved, dict) else None) or 'desc'

    if current_sort in ('created_desc', 'created_asc', 'updated_desc', 'updated_asc'):
        legacy = current_sort
        if legacy.startswith('created_'):
            current_sort = 'created'
            current_dir = 'asc' if legacy.endswith('_asc') else 'desc'
        else:
            current_sort = 'updated'
            current_dir = 'asc' if legacy.endswith('_asc') else 'desc'

    if not current_sort:
        current_sort = 'created'
        current_dir = 'desc'

    # Сортировка доступна всем в списке (как в Bitrix).
    if True:
        try:
            # joins for department sorting
            if current_sort == 'department':
                query = query.outerjoin(Department, Department.id == SupportTicket.department_id)

            # ВАЖНО: в нашей модели SupportTicket нет updated_at.
            # Для сортировки по "Обновлено" используем дату последнего события из TicketHistory,
            # а если истории нет — created_at.
            updated_expr = None
            if current_sort == 'updated':
                try:
                    th = TicketHistory
                    last_hist = db.session.query(
                        th.ticket_id.label('t_id'),
                        func.max(th.timestamp).label('mx')
                    ).group_by(th.ticket_id).subquery('last_hist')
                    query = query.outerjoin(last_hist, last_hist.c.t_id == SupportTicket.id)
                    updated_expr = db.func.coalesce(last_hist.c.mx, SupportTicket.created_at)
                except Exception:
                    updated_expr = SupportTicket.created_at

            sort_map = {
                'id': SupportTicket.id,
                'subject': SupportTicket.subject,
                'organization': SupportTicket.organization,
                'inn': SupportTicket.inn,
                'created': SupportTicket.created_at,
                'updated': updated_expr if updated_expr is not None else SupportTicket.created_at,
                'department': Department.name,
                'status': SupportTicket.status,
                'sla': SupportTicket.sla_deadline,
            }
            expr = sort_map.get(current_sort) or SupportTicket.created_at
            if current_dir == 'asc':
                query = query.order_by(expr.asc())
            else:
                query = query.order_by(expr.desc())
            applied_sorting = True
        except Exception:
            pass

    # === Применяем дефолтную сортировку, если не было приоритетной ===
    if not applied_sorting:
        query = query.order_by(SupportTicket.created_at.desc())

    # === Выполняем запрос ===

    # === Пагинация (как в Bitrix): 10 / 25 / 50 на страницу ===
    page = request.args.get('page', 1, type=int)
    show = request.args.get('show', None, type=int)
    if show is None:
        try:
            show = int((saved.get('show') if isinstance(saved, dict) else None) or 10)
        except Exception:
            show = 10
    if show not in (10, 25, 50):
        show = 10
    per_page = show

    # === Сохраняем настройки списка за пользователем (если есть пользователь) ===
    if ui_state is not None:
        try:
            # сохраняем только "настройки", без номера страницы
            new_saved = {
                'id': (request.args.get('id') or (saved.get('id') if isinstance(saved, dict) else '') or '').strip(),
                'subject': (request.args.get('subject') or (saved.get('subject') if isinstance(saved, dict) else '') or '').strip(),
                'email': (request.args.get('email') or (saved.get('email') if isinstance(saved, dict) else '') or '').strip(),
                'inn': (request.args.get('inn') or (saved.get('inn') if isinstance(saved, dict) else '') or '').strip(),
                'fio': (request.args.get('fio') or (saved.get('fio') if isinstance(saved, dict) else '') or '').strip(),
                'status': request.args.getlist('status') or (saved.get('status') if isinstance(saved, dict) else []) or [],
                'department': request.args.getlist('department') or (saved.get('department') if isinstance(saved, dict) else []) or [],
                'indicator': request.args.getlist('indicator') or (saved.get('indicator') if isinstance(saved, dict) else []) or [],
                'sort': current_sort,
                'dir': current_dir,
                'show': show,
            }
            ui_state.set('ticket_list', new_saved)
            db.session.commit()
            saved = new_saved
        except Exception:
            pass

    try:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    except Exception:
        # fallback for newer flask-sqlalchemy paginate helper
        pagination = db.paginate(query, page=page, per_page=per_page, error_out=False)

    tickets = pagination.items
    # ВАЖНО: нельзя присваивать relationship (t.end_user_rel = u), т.к. это выставит FK user_id и вызовет UPDATE.
    # Вместо этого кладём найденного пользователя в временный атрибут t.display_user.
    for t in tickets:
        try:
            t.display_user = getattr(t, 'client_rel', None) or getattr(t, 'end_user_rel', None)
        except Exception:
            t.display_user = None

    try:
        missing = [t for t in tickets if not getattr(t, 'display_user', None) and getattr(t, 'email', None)]
        if missing:
            emails = sorted({(t.email or '').strip().lower() for t in missing if (t.email or '').strip()})
            if emails:
                with db.session.no_autoflush:
                    users = User.query.filter(User.role == 'client', db.func.lower(User.email).in_(emails)).all()
                by_email = {(u.email or '').strip().lower(): u for u in users}
                for t in missing:
                    u = by_email.get((t.email or '').strip().lower())
                    if u:
                        t.display_user = u
    except Exception:
        pass



    # === Индикаторы состояний (как в Bitrix) ===
    # Считаем одним запросом последние сообщения для тикетов текущей страницы.
    try:
        t_ids = [t.id for t in tickets]
        last_is_op_map = {}
        if t_ids:
            tm = TicketMessage
            last_cte = db.session.query(
                tm.ticket_id.label('t_id'),
                func.max(tm.created_at).label('mx')
            ).filter(tm.ticket_id.in_(t_ids)).group_by(tm.ticket_id).subquery('last_msg_page')

            tm_last = aliased(tm, name='tm_last_page')
            rows = db.session.query(tm_last.ticket_id, tm_last.is_operator, tm_last.user_id).join(
                last_cte,
                db.and_(tm_last.ticket_id == last_cte.c.t_id, tm_last.created_at == last_cte.c.mx)
            ).all()
            for tid, is_op, user_id in rows:
                # если несколько совпадений по времени — берём любой
                last_is_op_map[int(tid)] = (bool(is_op), user_id)

        for t in tickets:
            last_meta = last_is_op_map.get(int(t.id))
            last_is_op = last_meta[0] if last_meta else None
            last_user_id = last_meta[1] if last_meta else None
            code, title = compute_ticket_indicator(t, last_is_op, last_user_id, getattr(current_user, 'id', None))
            t.state_indicator = code
            t.state_indicator_title = title
    except Exception:
        # не ломаем список из-за индикаторов
        for t in tickets:
            t.state_indicator = 'green'
            t.state_indicator_title = _INDICATOR_LABELS['green']
    # SLA view for current page
    try:
        sla_views = build_ticket_sla_views(tickets)
    except Exception:
        sla_views = {}
    for t in tickets:
        t.sla_view = sla_views.get(getattr(t, 'id', None))
        t.ui_is_overdue = bool(t.sla_view and t.sla_view.get('summary_status') == 'overdue')

    # Базовые параметры URL (сохраняем фильтры/сортировку/показ), кроме page/ajax
    base_args = request.args.to_dict(flat=False)
    base_args.pop('page', None)
    base_args.pop('ajax', None)

    def _url_with(**kwargs) -> str:
        args = {k: v[:] if isinstance(v, list) else v for k, v in base_args.items()}
        for k, v in kwargs.items():
            if v is None:
                args.pop(k, None)
                continue
            # allow list values
            args[k] = v
        return url_for('ticket_list', **args)

    def page_url(p: int) -> str:
        return _url_with(page=p)

    def show_url(n: int) -> str:
        return _url_with(show=n, page=1)

    def sort_url(field: str) -> str:
        # toggle dir if same field
        cur_field = current_sort or 'created'
        cur_dir = (current_dir or 'desc').strip().lower()
        if cur_field == field:
            ndir = 'asc' if cur_dir == 'desc' else 'desc'
        else:
            ndir = 'asc'
        return _url_with(sort=field, dir=ndir, page=1)

    # === Глобальные списки для фильтров ===
    STATUSES = get_active_statuses()
    DEPARTMENTS = [dept.name for dept in Department.query.order_by(Department.name).all()]
    try:
        CATEGORIES = [c.name for c in TicketCategory.query.filter_by(is_active=True).order_by(TicketCategory.sort_order, TicketCategory.name).all()]
    except Exception:
        CATEGORIES = []

    # AJAX: обновление списка без перезагрузки
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.args.get("ajax") == "1":
        html = render_template(
            "partials/_ticket_list_content.html",
            tickets=tickets,
            pagination=pagination,
            page_url=page_url,
            show_url=show_url,
            sort_url=sort_url,
            current_sort=current_sort,
            current_dir=current_dir,
            show=show,
            is_client=is_client,
            is_operator=is_operator,
            is_admin=is_admin,
        )
        return jsonify(success=True, html=html)

    return render_template(
        'ticket_list.html',
        tickets=tickets,
        pagination=pagination,
        page_url=page_url,
        show_url=show_url,
        sort_url=sort_url,
        current_sort=current_sort,
        current_dir=current_dir,
        show=show,
        page_title=page_title,
        is_client=is_client,
        is_operator=is_operator,
        is_admin=is_admin,
        STATUSES=STATUSES,
        DEPARTMENTS=DEPARTMENTS,
        CATEGORIES=CATEGORIES,
    )


@app.route('/api/ticket/<int:ticket_id>/summary', methods=['GET'])
@login_required
def api_ticket_summary(ticket_id: int):
    """Мини-выжимка для предпросмотра тикета в списке (sidebar preview)."""
    t = SupportTicket.query.get_or_404(ticket_id)

    # Права: клиент может смотреть только свои тикеты
    if getattr(current_user, 'role', None) == 'client':
        if (t.email or '').strip().lower() != (current_user.email or '').strip().lower():
            return jsonify({"success": False, "error": "forbidden"}), 403

    # Берём последние 3 сообщения (если есть TicketMessage)
    items = []
    try:
        msgs = TicketMessage.query.filter_by(ticket_id=t.id).order_by(TicketMessage.created_at.desc()).limit(3).all()
        for m in reversed(msgs):
            items.append({
                "who": (m.username or '').strip() or ("Оператор" if m.is_operator else "Клиент"),
                "ts": m.created_at.strftime('%d.%m.%Y %H:%M') if getattr(m, 'created_at', None) else '',
                "text": (m.message or '')[:500],
                "is_operator": bool(getattr(m, 'is_operator', False)),
            })
    except Exception:
        pass

    # Организация/ИНН — из end_user_rel либо из тикета
    org = (getattr(getattr(t, 'end_user_rel', None), 'organization', None) or getattr(t, 'organization', None) or '')
    inn = (getattr(getattr(t, 'end_user_rel', None), 'inn', None) or getattr(t, 'inn', None) or '')

    return jsonify({
        "success": True,
        "ticket": {
            "id": t.id,
            "subject": t.subject or '',
            "status": t.status or '',
            "created_at": t.created_at.strftime('%d.%m.%Y %H:%M') if t.created_at else '',
            "updated_at": t.updated_at.strftime('%d.%m.%Y %H:%M') if getattr(t, 'updated_at', None) else '',
            "org": org,
            "inn": inn,
            "url": url_for('ticket_detail', ticket_id=t.id),
        },
        "last": items,
    })


@app.route('/api/tickets/presence', methods=['GET'])
@login_required
def api_tickets_presence():
    """Batch presence для списка тикетов: /api/tickets/presence?ids=1,2,3"""
    ids_raw = (request.args.get('ids') or '').strip()
    if not ids_raw:
        return jsonify({"success": True, "presence": {}})
    try:
        ids = [int(x) for x in ids_raw.split(',') if x.strip().isdigit()]
    except Exception:
        ids = []
    if not ids:
        return jsonify({"success": True, "presence": {}})

    cutoff = utcnow() - timedelta(seconds=20)
    rows = TicketPresence.query.filter(
        TicketPresence.ticket_id.in_(ids),
        TicketPresence.last_seen >= cutoff
    ).all()

    pres = {}
    for r in rows:
        d = pres.setdefault(str(r.ticket_id), {"viewers": 0, "typers": []})
        d["viewers"] += 1
        if r.is_typing:
            d["typers"].append(r.display_name)

    return jsonify({"success": True, "presence": pres})


@app.route('/user/create-ticket', methods=['GET', 'POST'])
@login_required
def user_create_ticket():
    # Устаревший маршрут: используем единую страницу создания заявки
    return redirect(url_for('create_ticket'))

# === ЗАПУСК ПАРСЕРА ПОЧТЫ ===
def email_monitor():
    # Тихий режим + backoff: при сетевых проблемах не спамим лог.
    backoff = 30
    while True:
        try:
            _run_mail_check_once()
            backoff = 180
        except Exception as e:
            print(f"[MAIL PARSER] Ошибка: {e}. Следующая попытка через {backoff} сек.")
            backoff = min(backoff * 2, 300)
        pytime.sleep(backoff)

 # Автозапуск парсера отключён: запускай через CLI `flask mail-parser`
 # или кнопкой в админке (tab=mail).

@app.route('/knowledge')
@login_required
def knowledge():
    """База знаний (новая): категории -> статьи. Доступна всем авторизованным."""
    q = (request.args.get('q') or '').strip()
    only_fav = request.args.get('fav') == '1'

    def _kb_user_key(u):
        # Для совместимости с существующими избранными: клиенты = 'user', операторы/админы = 'op'
        try:
            if getattr(u, 'role', None) == 'client':
                return ('user', int(u.id))
        except Exception:
            pass
        return ('op', int(u.id))

    user_type, user_id = _kb_user_key(current_user)

    fav_ids = set()
    try:
        fav_ids = set(
            x.article_id for x in KnowledgeBaseFavorite.query.filter_by(user_type=user_type, user_id=user_id).all()
        )
    except Exception:
        fav_ids = set()

    # Категории
    categories = (KnowledgeBaseCategory.query
                  .filter_by(is_active=True)
                  .order_by(KnowledgeBaseCategory.sort_order.asc(), KnowledgeBaseCategory.title.asc())
                  .all())

    # Статьи: можно искать по заголовку/тегам/тексту (простая реализация)
    art_q = KnowledgeBaseArticle.query.filter_by(is_published=True)
    if q:
        like = f"%{q}%"
        art_q = art_q.filter(or_(KnowledgeBaseArticle.title.ilike(like),
                                 KnowledgeBaseArticle.tags.ilike(like),
                                 KnowledgeBaseArticle.summary.ilike(like),
                                 KnowledgeBaseArticle.body.ilike(like)))
    if only_fav:
        if fav_ids:
            art_q = art_q.filter(KnowledgeBaseArticle.id.in_(list(fav_ids)))
        else:
            art_q = art_q.filter(KnowledgeBaseArticle.id == -1)

    articles = (art_q
                .order_by(KnowledgeBaseArticle.updated_at.desc(), KnowledgeBaseArticle.id.desc())
                .all())

    # Группируем для шаблона
    by_cat = {}
    for a in articles:
        cid = a.category_id or 0
        by_cat.setdefault(cid, []).append(a)

    return render_template(
        'knowledge/index.html',
        q=q,
        only_fav=only_fav,
        categories=categories,
        by_cat=by_cat,
        fav_ids=fav_ids,
        is_operator=current_user.is_operator(),
        is_admin=(getattr(current_user, 'role', '') == 'admin')
    )


@app.route('/knowledge/article/<int:article_id>')
@login_required
def kb_article(article_id: int):
    a = KnowledgeBaseArticle.query.get_or_404(article_id)
    if not a.is_published and getattr(current_user, 'role', '') != 'admin':
        abort(404)

    def _kb_user_key(u):
        # Для совместимости с существующими избранными: клиенты = 'user', операторы/админы = 'op'
        try:
            if getattr(u, 'role', None) == 'client':
                return ('user', int(u.id))
        except Exception:
            pass
        return ('op', int(u.id))

    user_type, user_id = _kb_user_key(current_user)
    is_fav = False
    try:
        is_fav = KnowledgeBaseFavorite.query.filter_by(user_type=user_type, user_id=user_id, article_id=a.id).first() is not None
    except Exception:
        is_fav = False

    return render_template(
        'knowledge/article.html',
        article=a,
        is_favorite=is_fav,
        is_operator=current_user.is_operator(),
        is_admin=(getattr(current_user, 'role', '') == 'admin')
    )


@app.route('/knowledge/favorites')
@login_required
def kb_favorites():
    return redirect(url_for('knowledge', fav='1'))


@app.route('/api/kb/templates')
@login_required
def api_kb_templates():
    """Список шаблонов для вставки в комментарий (краткие описания статей).

    Возвращаем только опубликованные статьи.
    """
    q = (request.args.get('q') or '').strip()
    only_fav = request.args.get('fav') == '1'
    limit = min(int(request.args.get('limit') or 50), 100)

    def _kb_user_key(u):
        # Для совместимости с существующими избранными: клиенты = 'user', операторы/админы = 'op'
        try:
            if getattr(u, 'role', None) == 'client':
                return ('user', int(u.id))
        except Exception:
            pass
        return ('op', int(u.id))

    user_type, user_id = _kb_user_key(current_user)
    fav_ids = set(x.article_id for x in KnowledgeBaseFavorite.query.filter_by(user_type=user_type, user_id=user_id).all())

    qs = KnowledgeBaseArticle.query.filter_by(is_published=True)
    if q:
        like = f"%{q}%"
        qs = qs.filter(or_(KnowledgeBaseArticle.title.ilike(like),
                           KnowledgeBaseArticle.tags.ilike(like),
                           KnowledgeBaseArticle.summary.ilike(like)))
    if only_fav:
        if fav_ids:
            qs = qs.filter(KnowledgeBaseArticle.id.in_(list(fav_ids)))
        else:
            qs = qs.filter(KnowledgeBaseArticle.id == -1)

    rows = (qs.order_by(KnowledgeBaseArticle.updated_at.desc(), KnowledgeBaseArticle.id.desc())
            .limit(limit)
            .all())

    out = []
    for a in rows:
        out.append({
            'id': a.id,
            'title': a.title,
            'category': a.category.title if a.category else '',
            'summary': a.summary or '',
            'url': url_for('kb_article', article_id=a.id),
            'is_favorite': (a.id in fav_ids),
        })
    return jsonify({'success': True, 'items': out})


@app.route('/api/kb/favorite/<int:article_id>', methods=['POST'])
@login_required
def api_kb_toggle_favorite(article_id: int):
    KnowledgeBaseArticle.query.get_or_404(article_id)

    def _kb_user_key(u):
        # Для совместимости с существующими избранными: клиенты = 'user', операторы/админы = 'op'
        try:
            if getattr(u, 'role', None) == 'client':
                return ('user', int(u.id))
        except Exception:
            pass
        return ('op', int(u.id))

    user_type, user_id = _kb_user_key(current_user)
    fav = KnowledgeBaseFavorite.query.filter_by(user_type=user_type, user_id=user_id, article_id=article_id).first()
    if fav:
        db.session.delete(fav)
        db.session.commit()
        return jsonify({'success': True, 'is_favorite': False})
    db.session.add(KnowledgeBaseFavorite(user_type=user_type, user_id=user_id, article_id=article_id))
    db.session.commit()
    return jsonify({'success': True, 'is_favorite': True})


@app.route('/knowledge/manage')
@login_required
def kb_manage():
    """Управление базой знаний: только админ."""
    if getattr(current_user, 'role', '') != 'admin':
        flash_msg('Доступ запрещён', 'danger')
        return redirect(url_for('knowledge'))

    categories = (KnowledgeBaseCategory.query
                  .order_by(KnowledgeBaseCategory.sort_order.asc(), KnowledgeBaseCategory.title.asc())
                  .all())
    articles = (KnowledgeBaseArticle.query
                .order_by(KnowledgeBaseArticle.updated_at.desc(), KnowledgeBaseArticle.id.desc())
                .all())
    return render_template('knowledge/manage.html', categories=categories, articles=articles)


@app.route('/knowledge/manage/category/add', methods=['GET', 'POST'])
@login_required
def kb_category_add():
    if getattr(current_user, 'role', '') != 'admin':
        flash_msg('Доступ запрещён', 'danger')
        return redirect(url_for('knowledge'))

    if request.method == 'POST':
        title = (request.form.get('title') or '').strip()
        sort_order = int(request.form.get('sort_order') or 0)
        is_active = True if request.form.get('is_active') in ('1', 'on', 'true', 'yes') else False
        if not title:
            flash_msg('Введите название категории', 'warning')
            return render_template('knowledge/category_form.html', mode='add', category=None)
        db.session.add(KnowledgeBaseCategory(title=title, sort_order=sort_order, is_active=is_active))
        db.session.commit()
        flash_msg('Категория добавлена', 'success')
        return redirect(url_for('kb_manage'))

    return render_template('knowledge/category_form.html', mode='add', category=None)


@app.route('/knowledge/manage/category/<int:category_id>/edit', methods=['GET', 'POST'])
@login_required
def kb_category_edit(category_id: int):
    if getattr(current_user, 'role', '') != 'admin':
        flash_msg('Доступ запрещён', 'danger')
        return redirect(url_for('knowledge'))

    c = KnowledgeBaseCategory.query.get_or_404(category_id)
    if request.method == 'POST':
        c.title = (request.form.get('title') or '').strip()
        c.sort_order = int(request.form.get('sort_order') or 0)
        c.is_active = True if request.form.get('is_active') in ('1', 'on', 'true', 'yes') else False
        if not c.title:
            flash_msg('Введите название категории', 'warning')
            return render_template('knowledge/category_form.html', mode='edit', category=c)
        db.session.commit()
        flash_msg('Категория сохранена', 'success')
        return redirect(url_for('kb_manage'))

    return render_template('knowledge/category_form.html', mode='edit', category=c)


@app.route('/knowledge/manage/category/<int:category_id>/delete', methods=['POST'])
@login_required
def kb_category_delete(category_id: int):
    if getattr(current_user, 'role', '') != 'admin':
        flash_msg('Доступ запрещён', 'danger')
        return redirect(url_for('knowledge'))
    c = KnowledgeBaseCategory.query.get_or_404(category_id)
    db.session.delete(c)
    db.session.commit()
    flash_msg('Категория удалена', 'success')
    return redirect(url_for('kb_manage'))


@app.route('/knowledge/manage/article/add', methods=['GET', 'POST'])
@login_required
def kb_article_add():
    if getattr(current_user, 'role', '') != 'admin':
        flash_msg('Доступ запрещён', 'danger')
        return redirect(url_for('knowledge'))
    categories = KnowledgeBaseCategory.query.order_by(KnowledgeBaseCategory.sort_order.asc(), KnowledgeBaseCategory.title.asc()).all()

    if request.method == 'POST':
        title = (request.form.get('title') or '').strip()
        summary = (request.form.get('summary') or '').strip()
        body = (request.form.get('body') or '').strip()
        tags = (request.form.get('tags') or '').strip()
        category_id = request.form.get('category_id')
        is_published = True if request.form.get('is_published') in ('1', 'on', 'true', 'yes') else False

        if not title or not summary:
            flash_msg('Заполните название и краткий шаблон', 'warning')
            return render_template('knowledge/article_form.html', mode='add', article=None, categories=categories)

        cid = int(category_id) if (category_id and category_id.isdigit()) else None
        a = KnowledgeBaseArticle(title=title, summary=summary, body=body, tags=tags, category_id=cid, is_published=is_published)
        db.session.add(a)
        db.session.commit()
        flash_msg('Статья добавлена', 'success')
        return redirect(url_for('kb_manage'))

    return render_template('knowledge/article_form.html', mode='add', article=None, categories=categories)


@app.route('/knowledge/manage/article/<int:article_id>/edit', methods=['GET', 'POST'])
@login_required
def kb_article_edit(article_id: int):
    if getattr(current_user, 'role', '') != 'admin':
        flash_msg('Доступ запрещён', 'danger')
        return redirect(url_for('knowledge'))
    a = KnowledgeBaseArticle.query.get_or_404(article_id)
    categories = KnowledgeBaseCategory.query.order_by(KnowledgeBaseCategory.sort_order.asc(), KnowledgeBaseCategory.title.asc()).all()

    if request.method == 'POST':
        a.title = (request.form.get('title') or '').strip()
        a.summary = (request.form.get('summary') or '').strip()
        a.body = (request.form.get('body') or '').strip()
        a.tags = (request.form.get('tags') or '').strip()
        category_id = request.form.get('category_id')
        a.category_id = int(category_id) if (category_id and category_id.isdigit()) else None
        a.is_published = True if request.form.get('is_published') in ('1', 'on', 'true', 'yes') else False

        if not a.title or not a.summary:
            flash_msg('Заполните название и краткий шаблон', 'warning')
            return render_template('knowledge/article_form.html', mode='edit', article=a, categories=categories)

        db.session.commit()
        flash_msg('Статья сохранена', 'success')
        return redirect(url_for('kb_manage'))

    return render_template('knowledge/article_form.html', mode='edit', article=a, categories=categories)


@app.route('/knowledge/manage/article/<int:article_id>/delete', methods=['POST'])
@login_required
def kb_article_delete(article_id: int):
    if getattr(current_user, 'role', '') != 'admin':
        flash_msg('Доступ запрещён', 'danger')
        return redirect(url_for('knowledge'))
    a = KnowledgeBaseArticle.query.get_or_404(article_id)
    db.session.delete(a)
    db.session.commit()
    flash_msg('Статья удалена', 'success')
    return redirect(url_for('kb_manage'))



@app.route('/knowledge/add', methods=['GET', 'POST'], endpoint='add_faq')
@login_required
def add_faq_view():
    if current_user.role != 'admin':
        flash('Доступ запрещён', 'error')
        return redirect(url_for('knowledge'))

    if request.method == 'POST':
        question = (request.form.get('question') or '').strip()
        answer = (request.form.get('answer') or '').strip()
        category = (request.form.get('category') or 'Общее').strip() or 'Общее'
        order = int(request.form.get('order') or 0)
        is_active = True if request.form.get('is_active') in ('1', 'on', 'true', 'yes') else False

        if not question or not answer:
            flash('Заполните вопрос и ответ', 'error')
            return render_template('faq_form.html', mode='add', faq=None)

        faq = FAQ(question=question, answer=answer, category=category, order=order, is_active=is_active)
        db.session.add(faq)
        db.session.commit()
        flash('Статья добавлена', 'success')
        return redirect(url_for('knowledge'))

    return render_template('faq_form.html', mode='add', faq=None)


@app.route('/knowledge/<int:id>/edit', methods=['GET', 'POST'], endpoint='edit_faq')
@login_required
def edit_faq_view(id):
    if current_user.role != 'admin':
        flash('Доступ запрещён', 'error')
        return redirect(url_for('knowledge'))

    faq = FAQ.query.get_or_404(id)

    if request.method == 'POST':
        faq.question = (request.form.get('question') or '').strip()
        faq.answer = (request.form.get('answer') or '').strip()
        faq.category = (request.form.get('category') or 'Общее').strip() or 'Общее'
        faq.order = int(request.form.get('order') or 0)
        faq.is_active = True if request.form.get('is_active') in ('1', 'on', 'true', 'yes') else False

        if not faq.question or not faq.answer:
            flash('Заполните вопрос и ответ', 'error')
            return render_template('faq_form.html', mode='edit', faq=faq)

        db.session.commit()
        flash('Изменения сохранены', 'success')
        return redirect(url_for('knowledge'))

    return render_template('faq_form.html', mode='edit', faq=faq)


@app.route('/user/knowledge-base')
@login_required
def user_knowledge_base():
    """Backwards-compat: старый URL базы знаний."""
    return redirect(url_for('knowledge'))

from flask import request, redirect, url_for, render_template, flash, abort, send_file
from datetime import datetime

# === UPLOAD CONFIG FOR COMMENTS ===
COMMENT_UPLOAD_FOLDER = os.path.join('static', 'uploads', 'attachments')
os.makedirs(COMMENT_UPLOAD_FOLDER, exist_ok=True)

# =========================
# Mail parser runtime state (для кнопки в админке)
# =========================
MAIL_PARSER_STATE = {
    "running": False,
    "last_started_at": None,
    "last_finished_at": None,
    "last_success_at": None,
    "last_error": None,
}
_mail_parser_lock = threading.Lock()


def _run_mail_check_once():
    """Один проход проверки почты. Ошибки не пробрасываем наружу."""
    from mail_parser import check_incoming_emails

    try:
        with app.app_context():
            dept_names = [d.name for d in Department.query.order_by(Department.name).all()]
        check_incoming_emails(
            app=app,
            upload_folder=app.config['UPLOAD_FOLDER'],
            departments=dept_names,
            sla_attention_hours=SLA_ATTENTION_HOURS,
            sla_default_hours=SLA_DEFAULT_HOURS,
            bitrix_webhook_url=BITRIX_WEBHOOK_URL,
        )
        MAIL_PARSER_STATE["last_success_at"] = datetime.now(UTC)
        MAIL_PARSER_STATE["last_error"] = None
    except Exception as e:
        MAIL_PARSER_STATE["last_error"] = str(e)


def start_mail_check_async() -> bool:
    """Запустить проверку почты в фоне (1 проход). Возвращает False, если уже запущено."""
    with _mail_parser_lock:
        if MAIL_PARSER_STATE["running"]:
            return False
        MAIL_PARSER_STATE["running"] = True
        MAIL_PARSER_STATE["last_started_at"] = datetime.now(UTC)
        MAIL_PARSER_STATE["last_finished_at"] = None

    def _worker():
        try:
            _run_mail_check_once()
        finally:
            with _mail_parser_lock:
                MAIL_PARSER_STATE["running"] = False
                MAIL_PARSER_STATE["last_finished_at"] = datetime.now(UTC)

    threading.Thread(target=_worker, daemon=True).start()
    return True


def can_manage_spam_user(user) -> bool:
    """Кто может помечать/снимать Спам: админ и операторы 1/2 линии ТП."""
    role = getattr(user, 'role', None)
    if role == 'admin':
        return True
    # некоторые сборки используют нестандартные роли, но доступ 1/2 линии должен работать
    if role != 'operator':
        try:
            if isinstance(user, User) and is_tp_operator(user):
                return True
        except Exception:
            pass
        return False
    names = set()
    try:
        if getattr(user, 'department', None) is not None and getattr(user.department, 'name', None):
            names.add(user.department.name)
    except Exception:
        pass
    try:
        for d in getattr(user, 'departments', []) or []:
            if getattr(d, 'name', None):
                names.add(d.name)
    except Exception:
        pass
        # tolerate different naming: '1 линия ТП', '1ая линия техподдержки', etc.
    import re
    def _is_line(name: str, n: str) -> bool:
        """Гибкая проверка отдела 1/2 линии ТП.
        Допускаем варианты: '1ая линия ТП', '1 линия', '2-я линия техподдержки' и т.п.
        """
        s = (name or '').lower().replace('ё', 'е')
        if 'линия' not in s and 'line' not in s:
            return False
        if n == '1':
            return bool(re.search(r'(\b1\b|1\s*[-я]?|перва)', s))
        if n == '2':
            return bool(re.search(r'(\b2\b|2\s*[-я]?|втора)', s))
        return False
    for nm in names:
        if _is_line(nm, '1') or _is_line(nm, '2'):
            return True
    return False

@app.route('/ticket/<int:ticket_id>', methods=['GET', 'POST'])
@login_required
def ticket_detail(ticket_id):
    # === Определяем тип пользователя ===
    is_client = getattr(current_user, 'role', None) == 'client'
    is_operator = (not is_client) and isinstance(current_user, User) and (
        current_user.role in ['operator', 'admin'] or is_tp_operator(current_user)
    )
    is_admin = is_operator and current_user.role == 'admin'
    can_manage_spam = can_manage_spam_user(current_user)

    if not is_client and not is_operator:
        return redirect(url_for('login'))

    # === Получаем заявку ===
    ticket = SupportTicket.query.get_or_404(ticket_id)
    
    # Получаем закрепленный результат из связи pinned_result
    pinned_result = ticket.pinned_result
    
    # === Проверка доступа ===
    if is_client:
        if ticket.email != current_user.email:
            flash("Доступ запрещён", "error")
            return redirect(url_for('ticket_list'))
    else:
        if not is_admin:
            # Доступ разрешён, если:
            # - заявка назначена на пользователя ИЛИ
            # - заявка делегирована на отдел, в котором состоит пользователь
            has_access = False
            # 1/2 линия ТП имеет доступ ко всем заявкам
            if isinstance(current_user, User) and is_tp_operator(current_user):
                has_access = True
            elif ticket.assigned_to_id == current_user.id:
                has_access = True
            elif ticket.department_id and ticket.department_id in user_department_ids(current_user):
                has_access = True
            else:
                # Multi-delegate: дополнительные отделы
                try:
                    if getattr(ticket, 'shared_departments_rel', None):
                        dep_ids = set(user_department_ids(current_user))
                        for d in ticket.shared_departments_rel:
                            if d and getattr(d, 'id', None) in dep_ids:
                                has_access = True
                                break
                except Exception:
                    pass

            if not has_access:
                flash("Доступ к заявке запрещён", "error")
                return redirect(url_for('ticket_list'))

    # === Авто-обновление статуса при открытии оператором ===
    # Требование: если оператор открыл задачу, статус должен обновляться.
    # Логика: при первом открытии заявки со статусом "Новая" или "Принята" переводим в "В работе".
    if request.method == 'GET' and is_operator and ticket.status in ('Новая', 'Принята'):
        old_status = ticket.status
        ticket.status = 'В работе'
        try:
            log_ticket_change(ticket.id, current_user.id, 'status', old_status, ticket.status)
        except Exception:
            pass
        # В некоторых окружениях SQLite может быть read-only (права/FS). Не падаем 500 на просмотре.
        try:
            db.session.commit()
        except OperationalError:
            db.session.rollback()
    
    # === Инициализация формы комментария ===
    comment_form = MessageForm()

    # === Уведомление: заявка создана на другую организацию ===
    show_org_mismatch = False
    mismatch_info = None
    if is_client and request.method == 'GET' and request.args.get('org_mismatch'):
        prof_org = (getattr(current_user, 'organization', '') or '').strip()
        prof_inn = (getattr(current_user, 'inn', '') or '').strip()
        prof_addr = (getattr(current_user, 'address', '') or '').strip()
        t_org = (getattr(ticket, 'organization', '') or '').strip()
        t_inn = (getattr(ticket, 'inn', '') or '').strip()
        t_addr = (getattr(ticket, 'address', '') or '').strip()
        # показываем только если действительно отличается
        mismatch = False
        if prof_inn and t_inn:
            mismatch = (prof_inn != t_inn)
        else:
            p = normalize_org_name(prof_org)
            t = normalize_org_name(t_org)
            if p and t and p != t:
                mismatch = True
        if mismatch:
            show_org_mismatch = True
            mismatch_info = {
                'profile_org': prof_org,
                'profile_inn': prof_inn,
                'profile_addr': prof_addr,
                'ticket_org': t_org,
                'ticket_inn': t_inn,
                'ticket_addr': t_addr,
            }

    # === Обработка POST-действий ===
    if request.method == 'POST':
        action = request.form.get('action')
        edit_message_id = request.form.get('edit_message_id')

        # Клиент: перенести реквизиты из заявки в профиль
        if action == 'copy_org_to_profile' and is_client:
            # Берём реквизиты из тикета и валидируем как обязательные
            ok, org_v = validate_org(getattr(ticket, 'organization', '') or '')
            ok_inn, inn_v = validate_inn_ru(getattr(ticket, 'inn', '') or '', required=True)
            ok_addr, addr_v = validate_address(getattr(ticket, 'address', '') or '', required=True)
            if (not ok) or (not _norm_text(org_v)) or (not ok_inn) or (not ok_addr):
                flash('Нельзя перенести реквизиты: в заявке не заполнены корректные данные организации.', 'error')
                return redirect(url_for('ticket_detail', ticket_id=ticket.id))
            current_user.organization = org_v or None
            current_user.inn = inn_v or None
            current_user.address = addr_v or None
            # если были "предложенные" — очищаем, чтобы не путать
            for fld in ('suggested_organization', 'suggested_inn', 'suggested_address'):
                if hasattr(current_user, fld):
                    setattr(current_user, fld, None)
            try:
                db.session.commit()
                flash('Реквизиты перенесены в профиль.', 'success')
            except Exception:
                db.session.rollback()
                flash('Ошибка при сохранении профиля.', 'error')
            return redirect(url_for('ticket_detail', ticket_id=ticket.id))

        # === Единая кнопка "Завершить" (модальное окно) ===
        # finish_choice:
        #   - waiting: перевести в "Ожидание" (на подтверждение клиента)
        #   - spam/duplicate/wrong/withdrawn: закрыть в "Завершена" с причиной
        if action == 'finish_modal':
            finish_choice = (request.form.get('finish_choice') or '').strip()
            finish_comment = (request.form.get('finish_comment') or '').strip()

            if finish_choice == 'waiting':
                # Используем существующую логику: требуется закрепленный результат
                if not pinned_result:
                    flash("Нельзя перевести в ожидание без закрепленного результата в комментариях", "error")
                    return redirect(url_for('ticket_detail', ticket_id=ticket_id))
                old_status = ticket.status
                ticket.status = 'Ожидание'
                ticket.waiting_for_client_feedback = True
                ticket.marked_as_completed_at = utcnow()
                try:
                    ticket.close_reason = None
                    ticket.is_spam = False
                except Exception:
                    pass
                try:
                    log_ticket_change(ticket.id, current_user.id, 'status', old_status, 'Ожидание', 'Оператор отправил на подтверждение клиента')
                except Exception:
                    pass
                try:
                    history = TicketHistory(
                        ticket_id=ticket.id,
                        user_id=current_user.id,
                        field='status',
                        old_value=old_status,
                        new_value='Ожидание',
                        note='Ожидание подтверждения от клиента'
                    )
                    db.session.add(history)
                except Exception:
                    pass
                db.session.commit()
                flash("Заявка переведена в состояние 'Ожидание'", "success")
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))

            # Закрытие с причиной
            allowed = {'spam', 'duplicate', 'wrong', 'withdrawn'}
            if finish_choice not in allowed:
                flash('Выберите причину закрытия', 'error')
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))

            old_status = ticket.status
            ticket.status = 'Завершена'
            ticket.waiting_for_client_feedback = False
            ticket.closed_at = utcnow()
            try:
                ticket.is_resolved = True
            except Exception:
                pass
            try:
                ticket.close_reason = finish_choice
            except Exception:
                pass
            # пометка спамом влияет на показ в списке (скрываем по умолчанию)
            ticket.is_spam = (finish_choice == 'spam')

            # Инфо-комментарий (системный) — всегда, + пользовательский текст если задан
            try:
                fio = _presence_display_name(current_user)
                reason_map = {
                    'spam': 'Спам',
                    'duplicate': 'Дубликат',
                    'wrong': 'Ошибочная',
                    'withdrawn': 'Отозванная',
                }
                rname = reason_map.get(finish_choice, finish_choice)
                msg_html = f"<p><strong>Заявка закрыта оператором:</strong> {fio}. Причина: <strong>{rname}</strong></p>"
                if finish_comment:
                    msg_html += f"<div class='mt-2'><strong>Комментарий:</strong><br>{escape(finish_comment)}</div>"
                comment = TicketMessage(
                    ticket_id=ticket.id,
                    user_id=current_user.id,
                    message=msg_html,
                    is_operator=True
                )
                db.session.add(comment)
            except Exception:
                pass

            try:
                log_ticket_change(ticket.id, current_user.id, 'status', old_status, 'Завершена', f"Закрытие: {finish_choice}")
            except Exception:
                pass
            try:
                history = TicketHistory(
                    ticket_id=ticket.id,
                    user_id=current_user.id,
                    field='status',
                    old_value=old_status,
                    new_value='Завершена',
                    note=f"Закрыто: {finish_choice}"
                )
                db.session.add(history)
            except Exception:
                pass

            db.session.commit()
            flash('Заявка закрыта', 'success')
            return redirect(url_for('ticket_detail', ticket_id=ticket_id))

        # Спам
        # Кнопка "Не спам" удалена — вместо неё используется "Возобновить",
        # которое возвращает заявку обратно в очередь отдела.
        if action in ['mark_spam']:
            if not can_manage_spam:
                flash('Недостаточно прав для изменения статуса Спам', 'error')
                return redirect(url_for('ticket_detail', ticket_id=ticket.id))

            if action == 'mark_spam':
                ticket.is_spam = True
                ticket.status = 'Спам'
                ticket.closed_at = utcnow()
                db.session.commit()
                flash('Заявка закрыта как Спам', 'success')
                return redirect(url_for('ticket_detail', ticket_id=ticket.id))
  # ← Получаем один раз
        # Блокируем добавление комментариев в финальных статусах
        _closed_statuses = ('Завершена',)
        if comment_form.validate_on_submit() and not action:
            if ticket.status in _closed_statuses:
                flash('Нельзя добавлять комментарии в закрытой заявке', 'error')
                return redirect(url_for('ticket_detail', ticket_id=ticket.id))
            new_message = TicketMessage(
                ticket_id=ticket.id,
                user_id=current_user.id,
                message=((comment_form.message.data or '').strip()),
                is_operator=is_operator
            )
            db.session.add(new_message)
            db.session.flush()

            # --- Вложения (если есть) ---
            if 'files' in request.files:
                files = request.files.getlist('files')
                for uploaded_file in files:
                    if uploaded_file and uploaded_file.filename.strip():
                        filename = secure_filename(uploaded_file.filename)
                        unique_name = f"{uuid.uuid4().hex} {filename}"
                        file_path = os.path.join('static', 'uploads', 'attachments', unique_name)
                        os.makedirs(os.path.dirname(file_path), exist_ok=True)
                        uploaded_file.save(file_path)

                        attachment = TicketAttachment(
                            message_id=new_message.id,
                            filename=unique_name,
                            original_name=uploaded_file.filename,
                            size=os.path.getsize(file_path),
                            url=url_for('static', filename=f'uploads/attachments/{unique_name}')
                        )
                        db.session.add(attachment)

            # ВАЖНО: коммитим ВСЕГДА (раньше коммит был только если были файлы)
            db.session.commit()

            # --- In-app уведомления: новый комментарий ---
            try:
                if is_client:
                    # Клиент написал — уведомляем оператора (закрепившего/назначенного), иначе отдел
                    op = None
                    if getattr(ticket, 'locked_by', None):
                        op = db.session.get(User, ticket.locked_by)
                    elif getattr(ticket, 'assigned_to_id', None):
                        op = db.session.get(User, ticket.assigned_to_id)

                    notified_any = False
                    if op:
                        create_inapp_notification(
                            op,
                            'customer_reply',
                            f"Новый комментарий клиента в заявке #{ticket.id}",
                            "",
                            url_for('ticket_detail', ticket_id=ticket.id),
                            dedupe_key=f"customer_reply:{ticket.id}:{new_message.id}:op:{op.id}"
                        )
                        notified_any = True

                    # если заявка ещё не назначена/не закреплена — уведомим операторов отдела
                    recipients = []
                    if (not notified_any) and getattr(ticket, 'department_id', None):
                        dept = Department.query.get(ticket.department_id)
                        if dept:
                            recipients.extend(list(getattr(dept, 'users', []) or []))
                            recipients.extend(list(getattr(dept, 'operators', []) or []))

                    # если отдел пустой/не настроен — уведомим всех операторов/админов (чтобы не терялось)
                    if not recipients and not notified_any:
                        recipients = User.query.filter(User.role.in_(['operator', 'admin'])).all()

                    if not notified_any:
                        seen = set()
                        for u in recipients:
                            if not u or getattr(u, 'id', None) is None:
                                continue
                            if u.id in seen:
                                continue
                            seen.add(u.id)
                            if getattr(u, 'role', None) not in ('operator', 'admin'):
                                continue
                            create_inapp_notification(
                                u,
                                'customer_reply',
                                f"Новый комментарий клиента в заявке #{ticket.id}",
                                "",
                                url_for('ticket_detail', ticket_id=ticket.id),
                                dedupe_key=f"customer_reply:{ticket.id}:{new_message.id}:dept:{u.id}"
                            )
                else:
                    # Оператор написал — уведомляем клиента
                    if getattr(ticket, 'client_id', None):
                        cu = db.session.get(User, ticket.client_id)
                        if cu:
                            create_inapp_notification(
                                cu,
                                'customer_reply',
                                f"Новый комментарий в заявке #{ticket.id}",
                                "",
                                url_for('ticket_detail', ticket_id=ticket.id),
                                dedupe_key=f"comment:{ticket.id}:{new_message.id}:client"
                            )

                    # Оператор написал — уведомляем других операторов с доступом к заявке
                    try:
                        for op in _operators_with_access_to_ticket(ticket):
                            if not op or getattr(op, 'id', None) is None:
                                continue
                            if isinstance(current_user, User) and op.id == current_user.id:
                                continue
                            create_inapp_notification(
                                op,
                                'customer_reply',
                                f"Новый комментарий в заявке #{ticket.id}",
                                "",
                                url_for('ticket_detail', ticket_id=ticket.id),
                                dedupe_key=f"op_comment:{ticket.id}:{new_message.id}:op:{op.id}"
                            )
                    except Exception:
                        pass
            except Exception:
                pass

            flash("Комментарий добавлен", "success")
            return redirect(url_for('ticket_detail', ticket_id=ticket_id))

        # --- Действия клиента ---
        if is_client:
            if action == 'accept_work' and ticket.status == 'Ожидает подтверждения клиента':
                # Клиент принимает работу
                # ВАЖНО: это не "Спам". При подтверждении работы заявка должна стать "Завершена".
                ticket.status = 'Завершена'
                ticket.is_spam = False
                ticket.waiting_for_client_feedback = False
                ticket.closed_at = utcnow()
                ticket.helpful = True
                ticket.helpful_at = utcnow()
                # ЛОГИРУЕМ ПРИНЯТИЕ
                log_client_feedback(ticket.id, current_user.id, 'accepted')

                 # Логируем изменение статуса
                log_ticket_change(
                    ticket.id,
                    current_user.id,
                    'status',
                    'Ожидает подтверждения клиента',
                    'Завершена'
                )
                comment = TicketMessage(
                    ticket_id=ticket.id,
                    user_id=current_user.id,
                    message="<p><strong>Клиент подтвердил:</strong> Проблема решена</p>",
                    is_operator=False
                )
                db.session.add(comment)
                db.session.commit()

                # In-app уведомление оператору: клиент изменил статус
                try:
                    op = None
                    if getattr(ticket, 'locked_by', None):
                        op = db.session.get(User, ticket.locked_by)
                    elif getattr(ticket, 'assigned_to_id', None):
                        op = User.query.get(ticket.assigned_to_id)
                    if op:
                        create_inapp_notification(
                            op,
                            'status',
                            f"Клиент завершил заявку #{ticket.id}",
                            "Ожидает подтверждения клиента → Завершена",
                            url_for('ticket_detail', ticket_id=ticket.id),
                            dedupe_key=f"client_status:{ticket.id}:done"
                        )
                except Exception:
                    pass

                flash("Спасибо за подтверждение! Заявка завершена.", "success")
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))

            elif action == 'request_rework' and ticket.status == 'Ожидает подтверждения клиента':
                # Клиент отправляет на доработку
                ticket.status = 'В работе'
                ticket.waiting_for_client_feedback = False
                ticket.marked_as_completed_at = None
                # ЛОГИРУЕМ ЗАПРОС ДОРАБОТКИ
                log_client_feedback(ticket.id, current_user.id, 'rework')
                # Логируем изменение статуса
                log_ticket_change(
                    ticket.id,
                    current_user.id,
                    'status',
                    'Ожидает подтверждения клиента',
                    'В работе'
                )

                # In-app уведомление оператору: клиент отправил на доработку
                try:
                    op = None
                    if getattr(ticket, 'locked_by', None):
                        op = User.query.get(ticket.locked_by)
                    elif getattr(ticket, 'assigned_to_id', None):
                        op = User.query.get(ticket.assigned_to_id)
                    if op:
                        create_inapp_notification(
                            op,
                            'status',
                            f"Клиент запросил доработку по заявке #{ticket.id}",
                            "Ожидает подтверждения клиента → В работе",
                            url_for('ticket_detail', ticket_id=ticket.id),
                            dedupe_key=f"client_status:{ticket.id}:rework"
                        )
                except Exception:
                    pass
    
                # Получаем комментарий клиента
                client_comment = request.form.get('message', '').strip()
    
                # Создаем комментарий с комментарием клиента
                if client_comment:
                    comment_html = f"""<div>
    <div style="color: #dc3545; font-weight: bold; margin-bottom: 5px;">
        <i class="bi bi-arrow-counterclockwise"></i> Клиент запросил доработку:
    </div>
    <div style="color: #495057; padding: 8px 0;">
        {client_comment}
    </div>
</div>"""
                else:
                    comment_html = "<p><strong>Клиент запросил:</strong> Требуется доработка</p>"
    
                comment = TicketMessage(
                    ticket_id=ticket.id,
                    user_id=current_user.id,
                    message=comment_html,
                    is_operator=False
                )
                db.session.add(comment)
    
                # Уведомление оператора по email
                try:
                    operator = User.query.get(ticket.locked_by) if ticket.locked_by else None
                    if operator and operator.email:
                        msg = Message(
                            subject=f"Требуется доработка заявки #{ticket.id}",
                            recipients=[operator.email],
                            html=f"""
                            <p>Клиент запросил доработку по заявке #{ticket.id}.</p>
                            <p><strong>Тема:</strong> {ticket.subject}</p>
                            <p><strong>Комментарий клиента:</strong> {client_comment}</p>
                            <p>Ссылка: <a href="{url_for('ticket_detail', ticket_id=ticket.id, _external=True)}">Перейти к заявке</a></p>
                            """
                        )
                        mail.send(msg)
                except Exception as e:
                    print(f"Ошибка отправки email оператору: {e}")

                db.session.commit()
                flash("🔧 Заявка отправлена на доработку. Оператор получил уведомление.", "warning")
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))
            
            elif action == 'reopen_ticket_operator' and ticket.status in ('Завершена','Спам','Дубликат','Ошибочная'):
                # "Возобновить" возвращает заявку обратно в очередь отдела:
                # снимаем назначение/закрепление и сбрасываем финальные/спам-метки.
                ticket.status = 'Новая'
                ticket.is_spam = False
                ticket.assigned_to_id = None
                try:
                    ticket.locked_by = None
                except Exception:
                    pass
                ticket.waiting_for_client_feedback = False
                ticket.marked_as_completed_at = None
                ticket.auto_closed_at = None
                ticket.closed_at = None
    
                # Добавляем комментарий
                comment = TicketMessage(
                    ticket_id=ticket.id,
                    user_id=current_user.id,
                    message="<p><strong>Оператор возобновил заявку</strong></p>",
                    is_operator=True
                )
                db.session.add(comment)
    
                db.session.commit()
                flash("Заявка возобновлена", "success")
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))
            
            elif action == 'reopen_ticket' and ticket.status in ('Завершена','Спам','Дубликат','Ошибочная'):
                # Возврат в очередь отдела
                ticket.status = 'Новая'
                ticket.is_spam = False
                ticket.assigned_to_id = None
                try:
                    ticket.locked_by = None
                except Exception:
                    pass
                db.session.commit()
                flash("Заявка возобновлена", "success")
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))
            
            elif action == 'client_complete_ticket' and ticket.status != 'Завершена':
                # Клиент сам завершает заявку
                # Самостоятельное завершение = "Завершена" (не "Спам")
                old_status = ticket.status
                ticket.status = 'Завершена'
                ticket.is_spam = False
                ticket.closed_at = utcnow()
                ticket.helpful = True
                ticket.helpful_at = utcnow()

                # Логируем изменение статуса
                log_ticket_change(
                    ticket.id,
                    current_user.id,
                    'status',
                    old_status,
                    'Завершена'
                )

                # Добавляем комментарий от клиента
                comment = TicketMessage(
                    ticket_id=ticket.id,
                    user_id=current_user.id,
                    message="<p><strong>Клиент завершил заявку самостоятельно</strong></p>",
                    is_operator=False
                )
                db.session.add(comment)
                db.session.commit()

                flash("Заявка успешно завершена.", "success")
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))

        # --- Действия оператора/админа ---
        elif is_operator:
            if action == 'update_ticket_meta':
                # Обновление справочных полей (категория/статус/приоритет/теги)
                new_status = request.form.get('status', '').strip()
                new_priority = request.form.get('priority', '').strip()
                new_category_id = request.form.get('category_id')
                tag_ids = request.form.getlist('tag_ids')

                try:
                    # Статус
                    if new_status and new_status != ticket.status:
                        old = ticket.status
                        ticket.status = new_status
                        log_ticket_change(ticket.id, current_user.id, 'status', old, new_status)


                        # Уведомление клиенту о смене статуса
                        try:
                            if getattr(ticket, 'client_id', None):
                                eu = User.query.get(ticket.client_id)
                                if eu:
                                    create_inapp_notification(
                                        eu,
                                        'status',
                                        f"Статус заявки #{ticket.id} изменён",
                                        f"{old} → {new_status}",
                                        url_for('ticket_detail', ticket_id=ticket.id),
                                        dedupe_key=f"status:{ticket.id}:{new_status}"
                                    )
                        except Exception:
                            pass
                    # Приоритет (храним code)
                    if new_priority and new_priority != (ticket.priority or ''):
                        old = ticket.priority or ''
                        ticket.priority = new_priority
                        log_ticket_change(ticket.id, current_user.id, 'priority', old, new_priority)

                    # Категория
                    if new_category_id and str(new_category_id).isdigit():
                        new_cat = TicketCategory.query.get(int(new_category_id))
                        if new_cat and new_cat.is_active:
                            old_name = ticket.category_rel.name if ticket.category_rel else ''
                            if ticket.category_id != new_cat.id:
                                ticket.category_id = new_cat.id
                                log_ticket_change(ticket.id, current_user.id, 'category', old_name, new_cat.name)

                    # Теги
                    try:
                        new_tag_ids = [int(t) for t in tag_ids if str(t).isdigit()]
                    except Exception:
                        new_tag_ids = []
                    new_tags = Tag.query.filter(Tag.id.in_(new_tag_ids), Tag.is_active == True).all() if new_tag_ids else []
                    old_tags = ', '.join([t.name for t in (ticket.tags_rel or [])])
                    new_tags_str = ', '.join([t.name for t in new_tags])
                    if old_tags != new_tags_str:
                        ticket.tags_rel = new_tags
                        log_ticket_change(ticket.id, current_user.id, 'tags', old_tags, new_tags_str)

                    db.session.commit()
                    flash("Поля заявки обновлены", "success")
                except Exception as e:
                    db.session.rollback()
                    flash(f"Ошибка обновления заявки: {e}", "error")
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))

            # Спам (кнопка "Не спам" удалена — используйте "Возобновить")
            if action == 'mark_spam':
                if not can_manage_spam:
                    flash("Недостаточно прав для изменения статуса Спам.", "error")
                    return redirect(url_for('ticket_detail', ticket_id=ticket_id))

                if not ticket.is_spam:
                    ticket.is_spam = True
                    ticket.waiting_for_client_feedback = False
                    ticket.status = 'Спам'
                    ticket.closed_at = utcnow()

                    # Добавляем системный комментарий
                    try:
                        fio = _presence_display_name(current_user)
                        comment = TicketMessage(
                            ticket_id=ticket.id,
                            user_id=current_user.id,
                            message=f"<p><strong>Заявка закрыта оператором:</strong> {fio} с пометкой <strong>СПАМ</strong></p>",
                            is_operator=True
                        )
                        db.session.add(comment)
                    except Exception:
                        pass

                    try:
                        log_ticket_change(ticket.id, current_user.id, 'spam', '0', '1', 'Заявка закрыта как Спам')
                    except Exception:
                        pass

                    try:
                        history = TicketHistory(
                            ticket_id=ticket.id,
                            user_id=current_user.id,
                            field='spam',
                            old_value='нет',
                            new_value='да',
                            note='Оператор закрыл как Спам'
                        )
                        db.session.add(history)
                    except Exception:
                        pass

                    db.session.commit()
                    flash("Заявка закрыта как Спам", "warning")

                return redirect(url_for('ticket_detail', ticket_id=ticket_id))
            elif action == 'close_mistake':
                # Быстрое закрытие без комментария: Ошибочная
                if not can_manage_spam:
                    flash("Недостаточно прав для быстрого закрытия заявки.", "error")
                    return redirect(url_for('ticket_detail', ticket_id=ticket_id))

                if ticket.status not in ('Завершена','Спам','Дубликат','Ошибочная'):
                    ticket.is_spam = False
                    ticket.waiting_for_client_feedback = False
                    ticket.status = 'Ошибочная'
                    ticket.closed_at = utcnow()
                    try:
                        ticket.is_resolved = True
                    except Exception:
                        pass

                    # Добавляем системный комментарий
                    try:
                        fio = _presence_display_name(current_user)
                        comment = TicketMessage(
                            ticket_id=ticket.id,
                            user_id=current_user.id,
                            message=f"<p><strong>Заявка закрыта оператором:</strong> {fio} с пометкой <strong>Ошибочная</strong></p>", 
                            is_operator=True
                        )
                        db.session.add(comment)
                    except Exception:
                        pass

                    try:
                        log_ticket_change(ticket.id, current_user.id, 'status', '', 'Ошибочная', 'Заявка закрыта как Ошибочная')
                    except Exception:
                        pass

                    try:
                        history = TicketHistory(
                            ticket_id=ticket.id,
                            user_id=current_user.id,
                            field='status',
                            old_value='',
                            new_value='Ошибочная',
                            note='Оператор закрыл как Ошибочная'
                        )
                        db.session.add(history)
                    except Exception:
                        pass

                    db.session.commit()
                    flash("Заявка закрыта как Ошибочная", "warning")
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))

            elif action == 'close_withdrawn':
                # Быстрое закрытие без комментария: Отозвано
                if not can_manage_spam:
                    flash("Недостаточно прав для быстрого закрытия заявки.", "error")
                    return redirect(url_for('ticket_detail', ticket_id=ticket_id))

                if ticket.status not in ('Завершена','Спам','Дубликат','Ошибочная','Отозвано'):
                    ticket.is_spam = False
                    ticket.waiting_for_client_feedback = False
                    ticket.status = 'Отозвано'
                    ticket.closed_at = utcnow()
                    try:
                        ticket.is_resolved = True
                    except Exception:
                        pass

                    # Добавляем системный комментарий
                    try:
                        fio = _presence_display_name(current_user)
                        comment = TicketMessage(
                            ticket_id=ticket.id,
                            user_id=current_user.id,
                            message=f"<p><strong>Заявка отозвана:</strong> {fio} — статус <strong>Отозвано</strong></p>",
                            is_operator=True
                        )
                        db.session.add(comment)
                    except Exception:
                        pass

                    try:
                        log_ticket_change(ticket.id, current_user.id, 'status', '', 'Отозвано', 'Заявка закрыта как Отозвано')
                    except Exception:
                        pass

                    try:
                        history = TicketHistory(
                            ticket_id=ticket.id,
                            user_id=current_user.id,
                            field='status',
                            old_value='',
                            new_value='Отозвано',
                            note='Оператор закрыл как Отозвано'
                        )
                        db.session.add(history)
                    except Exception:
                        pass

                    db.session.commit()
                    flash("Заявка закрыта как Отозвано", "warning")

                return redirect(url_for('ticket_detail', ticket_id=ticket_id))

            elif action == 'close_duplicate':
                # Быстрое закрытие без комментария: Дубликат
                if not can_manage_spam:
                    flash("Недостаточно прав для быстрого закрытия заявки.", "error")
                    return redirect(url_for('ticket_detail', ticket_id=ticket_id))

                if ticket.status not in ('Завершена','Спам','Дубликат','Ошибочная'):
                    ticket.is_spam = False
                    ticket.waiting_for_client_feedback = False
                    ticket.status = 'Дубликат'
                    ticket.closed_at = utcnow()
                    try:
                        ticket.is_resolved = True
                    except Exception:
                        pass

                    # Добавляем системный комментарий
                    try:
                        fio = _presence_display_name(current_user)
                        comment = TicketMessage(
                            ticket_id=ticket.id,
                            user_id=current_user.id,
                            message=f"<p><strong>Заявка закрыта оператором:</strong> {fio} с пометкой <strong>Дубликат</strong></p>", 
                            is_operator=True
                        )
                        db.session.add(comment)
                    except Exception:
                        pass

                    try:
                        log_ticket_change(ticket.id, current_user.id, 'status', '', 'Дубликат', 'Заявка закрыта как Дубликат')
                    except Exception:
                        pass

                    try:
                        history = TicketHistory(
                            ticket_id=ticket.id,
                            user_id=current_user.id,
                            field='status',
                            old_value='',
                            new_value='Дубликат',
                            note='Оператор закрыл как Дубликат'
                        )
                        db.session.add(history)
                    except Exception:
                        pass

                    db.session.commit()
                    flash("Заявка закрыта как Дубликат", "warning")
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))

            elif action == 'complete_ticket' and ticket.status != 'Завершена':
                # Проверяем, есть ли закрепленный результат
                if not pinned_result:
                    flash("Нельзя завершить заявку без закрепленного результата в комментариях", "error")
                    return redirect(url_for('ticket_detail', ticket_id=ticket_id))
                old_status = ticket.status
                # Меняем статус на "Ожидание" (ожидание подтверждения клиента)
                ticket.status = 'Ожидание'
                ticket.waiting_for_client_feedback = True
                ticket.marked_as_completed_at = utcnow()
        
                # ЛОГИРУЕМ ОТПРАВКУ НА ПРОВЕРКУ
                log_ticket_change(
                    ticket.id,
                    current_user.id,
                    'status',
                    old_status,
                    'Ожидание',
                    'Оператор отправил на проверку клиенту'
                )
    
                # Логируем начало ожидания обратной связи
                history = TicketHistory(
                    ticket_id=ticket.id,
                    user_id=current_user.id,
                    field='client_feedback',
                    old_value='не требуется',
                    new_value='ожидает',
                    note='Ожидание подтверждения от клиента (24ч)'
                )
                db.session.add(history)

                db.session.commit()
        
                # Отправляем уведомление клиенту
                try:
                    if ticket.end_user_rel and ticket.end_user_rel.email:
                        msg = Message(
                            subject=f"Заявка #{ticket.id} готова к проверке",
                            recipients=[ticket.end_user_rel.email],
                            html=f"""
                            <p>Здравствуйте!</p>
                            <p>Оператор отметил, что ваша заявка #{ticket.id} решена.</p>
                            <p><strong>Пожалуйста, подтвердите:</strong></p>
                            <ul>
                                <li>✅ Проблема решена — нажмите "Принять работу"</li>
                                <li>🔧 Требуется доработка — нажмите "Доделать"</li>
                            </ul>
                            <p>Если вы не отреагируете в течение 24 часов, заявка будет автоматически завершена.</p>
                            <p>Ссылка на заявку: 
                            <a href="{url_for('ticket_detail', ticket_id=ticket.id, _external=True)}">
                                Перейти к заявке #{ticket.id}
                            </a></p>
                            <p>С уважением,<br>Техническая поддержка</p>
                            """
                        )
                        mail.send(msg)
                except Exception as e:
                    print(f"Ошибка отправки email клиенту: {e}")
        
                flash("Заявка отправлена клиенту на подтверждение. Клиент должен подтвердить решение в течение 24 часов.", "success")
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))
            
            elif action == 'reopen_ticket_operator' and ticket.status in ('Завершена','Спам','Дубликат','Ошибочная'):
                ticket.closed_at = utcnow()
                ticket.status = 'В работе'
                db.session.commit()
                flash("Заявка возобновлена", "success")
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))
            
            elif action == 'delegate_ticket':
                # multi-delegate: можно выбрать несколько отделов
                department_ids = request.form.getlist('department_ids')

            if not is_admin:
                department_ids = []
                if not department_ids:
                    # обратная совместимость (если где-то осталась старая форма)
                    single_id = request.form.get('department_id')
                    if single_id:
                        department_ids = [single_id]
                change_status = request.form.get('change_status') == 'on'  # Проверяем галочку
    
                if department_ids:
                    # Берём основной отдел — первый в списке
                    departments = Department.query.filter(Department.id.in_(department_ids)).all()
                    if not departments:
                        flash("Отдел не найден", "danger")
                        return redirect(url_for('ticket_detail', ticket_id=ticket_id))

                    # Сохраним порядок выбора: departments из IN() могут прийти в другом порядке.
                    dept_by_id = {str(d.id): d for d in departments}
                    ordered = [dept_by_id.get(str(i)) for i in department_ids if dept_by_id.get(str(i))]
                    main_dept = ordered[0]
                    extra_depts = ordered[1:]

                    old_department_name = ticket.department.name if ticket.department else 'Без отдела'
                    ticket.department_id = main_dept.id
                    ticket.assigned_to_id = None

                    # Multi-delegate: сохраняем дополнительные отделы
                    try:
                        ticket.shared_departments_rel = extra_depts
                    except Exception:
                        pass

                    # Логирование
                    log_ticket_change(ticket.id, current_user.id, 'department', old_department_name, main_dept.name)
                    if extra_depts:
                        try:
                            log_ticket_change(
                                ticket.id,
                                current_user.id,
                                'shared_departments',
                                '-',
                                ', '.join(d.name for d in extra_depts)
                            )
                        except Exception:
                            pass

                    # Меняем статус только если галочка установлена
                    if change_status and ticket.status == 'Новая':
                        ticket.status = 'Принята'

                    db.session.commit()

                    if extra_depts:
                        flash(
                            f"Заявка делегирована: основной отдел — {main_dept.name}; доп. отделы — {', '.join(d.name for d in extra_depts)}",
                            "success"
                        )
                    else:
                        flash(f"Заявка делегирована в отдел: {main_dept.name}", "success")
                else:
                    flash("Не выбран отдел", "warning")
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))

            elif action == 'update_shared_departments':
                # Управление дополнительными отделами без смены основного department_id
                department_ids = request.form.getlist('shared_department_ids')
                # Защита: не добавляем основной отдел в shared
                main_id = str(ticket.department_id) if ticket.department_id else None
                if main_id:
                    department_ids = [i for i in department_ids if str(i) != main_id]

                departments = []
                if department_ids:
                    departments = Department.query.filter(Department.id.in_(department_ids)).all()

                old_names = []
                try:
                    if getattr(ticket, 'shared_departments_rel', None):
                        old_names = [d.name for d in ticket.shared_departments_rel]
                except Exception:
                    old_names = []

                try:
                    ticket.shared_departments_rel = departments
                except Exception:
                    pass

                try:
                    log_ticket_change(
                        ticket.id,
                        current_user.id,
                        'shared_departments',
                        ', '.join(old_names) if old_names else '—',
                        ', '.join(d.name for d in departments) if departments else '—',
                        'Обновлены дополнительные отделы'
                    )
                except Exception:
                    pass

                db.session.commit()
                flash('Дополнительные отделы обновлены', 'success')
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))

            elif action == 'update_departments_sidebar':
                # Из правой колонки: смена основного отдела + мультивыбор доп. отделов
                new_main_id = (request.form.get('department_id') or '').strip()
                shared_ids = request.form.getlist('shared_department_ids')

                if not new_main_id:
                    flash('Не выбран основной отдел', 'warning')
                    return redirect(url_for('ticket_detail', ticket_id=ticket_id))

                main_dept = Department.query.get(int(new_main_id))
                if not main_dept:
                    flash('Отдел не найден', 'danger')
                    return redirect(url_for('ticket_detail', ticket_id=ticket_id))

                # Не добавляем основной отдел в shared
                shared_ids = [i for i in shared_ids if str(i) != str(main_dept.id)]
                shared_depts = []
                if shared_ids:
                    shared_depts = Department.query.filter(Department.id.in_(shared_ids)).all()

                old_main_name = ticket.department.name if ticket.department else '—'
                old_shared_names = []
                try:
                    if getattr(ticket, 'shared_departments_rel', None):
                        old_shared_names = [d.name for d in ticket.shared_departments_rel]
                except Exception:
                    old_shared_names = []

                # Основной отдел
                if ticket.department_id != main_dept.id:
                    ticket.department_id = main_dept.id
                    # Если поменяли отдел — сбрасываем назначение на оператора
                    ticket.assigned_to_id = None
                    try:
                        log_ticket_change(ticket.id, current_user.id, 'department', old_main_name, main_dept.name)
                    except Exception:
                        pass

                # Доп. отделы
                try:
                    ticket.shared_departments_rel = shared_depts
                except Exception:
                    pass
                try:
                    log_ticket_change(
                        ticket.id,
                        current_user.id,
                        'shared_departments',
                        ', '.join(old_shared_names) if old_shared_names else '—',
                        ', '.join(d.name for d in shared_depts) if shared_depts else '—',
                        'Обновлены дополнительные отделы'
                    )
                except Exception:
                    pass

                db.session.commit()
                flash('Отделы заявки обновлены', 'success')
                return redirect(url_for('ticket_detail', ticket_id=ticket_id))
        
        # --- Непредвиденное действие или попытка без прав ---
        else:
            flash("Недопустимое действие", "error")
            return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    # === Загрузка сообщений чата с вложениями ===
    messages = []
    for msg in TicketMessage.query.filter_by(ticket_id=ticket_id).order_by(TicketMessage.created_at.desc()):
        author = None
        if msg.is_operator:
            author = db.session.get(User, msg.user_id)
            if author:
                last_name = author.last_name or ''
                name = author.name or ''
                patronymic = author.patronymic or ''
                parts = [last_name, name, patronymic]
                username = ' '.join(part for part in parts if part)
                if not username or username == ' ':
                    username = author.username or 'Оператор'
            else:
                username = 'Оператор'    
        else:
            author = db.session.get(User, msg.user_id)
            if author:
                last_name = author.last_name or ''
                name = author.name or ''
                patronymic = author.patronymic or ''
                parts = [last_name, name, patronymic]
                username = ' '.join(part for part in parts if part).strip()
                if not username:
                    username = author.email or author.username or 'Клиент'
            else:
                username = 'Клиент'
    
        # Получаем вложения из relationship
        attachments = []
        for att in msg.file_attachments:
            attachments.append({
                'id': att.id,
                'filename': att.filename,
                'original_name': att.original_name,
                'url': att.url,
                'size': att.size
            })

        # Имя пользователя, который редактировал (если есть)
        edited_by_name = 'Пользователь'
        edited_by_id = getattr(msg, 'edited_by_id', None)
        if edited_by_id:
            u = User.query.get(edited_by_id)
            if u:
                parts = [u.last_name or '', u.name or '', u.patronymic or '']
                edited_by_name = ' '.join(p for p in parts if p).strip() or (u.username or 'Оператор')
            else:
                edited_by_name = 'Пользователь'

        messages.append({
            'id': msg.id,
            'message': msg.message,
            'raw_message': html_to_bbcode(msg.message),
            'created_at': msg.created_at.isoformat(),
            'is_operator': msg.is_operator,
            'username': username,
            'user_id': msg.user_id,
            'attachments': attachments,
            'edited_at': (msg.edited_at.isoformat() if getattr(msg, 'edited_at', None) else None),
            'edited_by_id': edited_by_id,
            'edited_by_name': edited_by_name,
        })
    
    # Если pinned_result не был получен из связи (например, если есть ошибка в связи)
    # или нужно получить его заново
    if not pinned_result and ticket.pinned_result_id:
        pinned_result = TicketMessage.query.get(ticket.pinned_result_id)
    
    # === SLA ===
    sla_class = ""
    sla_info = ""
    if not ticket.is_resolved and ticket.sla_deadline:
        now = utcnow()
        if now > ticket.sla_deadline:
            sla_class = "sla-red"
            sla_info = "ПРОСРОЧЕНО!"
        else:
            total = (ticket.sla_deadline - ticket.created_at).total_seconds()
            remaining = (ticket.sla_deadline - now).total_seconds()
            hrs = int(remaining // 3600)
            mins = int((remaining % 3600) // 60)
            if remaining < total * 0.25:
                sla_class = "sla-yellow"
                sla_info = f"Осталось: ~{hrs} ч {mins} мин"
            else:
                sla_class = "sla-green"
                sla_info = f"До дедлайна: ~{hrs} ч {mins} мин"

    # === Операторы для делегирования (если нужно) ===
    operators_for_delegate = []
    if is_operator:
        operators_for_delegate = User.query.filter(
            User.role == 'operator',
            User.id != current_user.id
        ).all()

    # === Отделы для делегирования (НОВОЕ) ===
    all_departments = []
    if is_operator:
        all_departments = Department.query.all()

    # === Ответные файлы (если есть) ===
    response_attachments = []

    # === Операторы отдела для отображения в сайдбаре ===
    department_operators = []
    if ticket.department_id:
        department_operators = User.query.filter(
            User.role == 'operator',
            User.department_id == ticket.department_id,
        ).all()

    # Определяем, нужно ли показывать подсказку об авто-повышении
    auto_promoted = (
        ticket.priority == 'high' and
        ticket.created_at < utcnow() - timedelta(hours=24) and
        ticket.status != 'Завершена'
    )
    # Сериализуем закреплённый результат для передачи в JS
    pinned_result_serialized = None
    if pinned_result:
        pinned_result_serialized = {
            'id': pinned_result.id,
            'message': pinned_result.message,
            'created_at_formatted': (pinned_result.created_at.strftime('%d.%m.%Y %H:%M') if pinned_result.created_at else None),
            'user_id': pinned_result.user_id,
            'is_operator': pinned_result.is_operator,
            'attachments': [
                {
                    'original_name': att.original_name,
                    'url': att.url,
                    'size': att.size
                }
                for att in pinned_result.file_attachments
            ] if pinned_result.file_attachments else []
        }
    available_statuses = get_active_statuses()
    available_priorities = get_active_priorities()
    available_tags = get_active_tags()
    available_categories = TicketCategory.query.filter_by(is_active=True).order_by(TicketCategory.sort_order).all()
    close_reasons = get_active_close_reasons()

    # === Рендер ===
    return render_template(
        'ticket_detail.html',
        ticket=ticket,
        messages=messages,
        sla_class=sla_class,
        sla_info=sla_info,
        sla_view=ticket_sla_view(ticket),
        is_client=is_client,
        is_operator=is_operator,
        is_admin=is_admin,
        operators_for_delegate=operators_for_delegate,
        all_departments=all_departments,
        department_operators=department_operators,
        response_attachments=response_attachments,
        current_user_id=current_user.id,
        current_user_name=getattr(current_user, 'username', current_user.email),
        comment_form=comment_form,
        current_user_is_operator=is_operator,
        current_user_is_client=is_client,
        now=utcnow(),
        auto_promoted=auto_promoted,
        pinned_result=pinned_result_serialized,
        available_statuses=available_statuses,
        available_priorities=available_priorities,
        available_tags=available_tags,
        available_categories=available_categories,
        close_reasons=close_reasons,
        can_manage_spam=can_manage_spam,
        show_org_mismatch=show_org_mismatch,
        mismatch_info=mismatch_info
    )

@app.post('/tickets/<int:ticket_id>/delete')
@login_required
def delete_ticket(ticket_id: int):
    # Удалять заявки может только администратор
    is_client = getattr(current_user, 'role', None) == 'client'
    is_operator = not is_client and getattr(current_user, 'role', None) in ['operator', 'admin']
    is_admin = is_operator and getattr(current_user, 'role', None) == 'admin'
    if not is_admin:
        abort(403)

    ticket = SupportTicket.query.get_or_404(ticket_id)

    try:
        # Важно: у тикета есть ссылки на ticket_messages (result_id / pinned_result_id).
        # Их нужно обнулить до удаления сообщений.
        if hasattr(ticket, 'pinned_result_id'):
            ticket.pinned_result_id = None
        if hasattr(ticket, 'result_id'):
            ticket.result_id = None
        db.session.flush()

        # 1) Сообщения тикета + зависимости
        msg_ids = [mid for (mid,) in db.session.query(TicketMessage.id).filter(TicketMessage.ticket_id == ticket.id).all()]
        if msg_ids:
            TicketAttachment.query.filter(TicketAttachment.message_id.in_(msg_ids)).delete(synchronize_session=False)
            CommentLike.query.filter(CommentLike.comment_id.in_(msg_ids)).delete(synchronize_session=False)
            TicketMessage.query.filter(TicketMessage.id.in_(msg_ids)).delete(synchronize_session=False)

        # 2) История / присутствие / опер-чат
        TicketHistory.query.filter_by(ticket_id=ticket.id).delete(synchronize_session=False)
        TicketPresence.query.filter_by(ticket_id=ticket.id).delete(synchronize_session=False)
        TicketOperatorChatRead.query.filter_by(ticket_id=ticket.id).delete(synchronize_session=False)
        TicketOperatorChatMessage.query.filter_by(ticket_id=ticket.id).delete(synchronize_session=False)

        # 3) Сам тикет
        db.session.delete(ticket)
        db.session.commit()
        flash('Заявка удалена', 'success')
    except IntegrityError as e:
        db.session.rollback()
        flash('Не удалось удалить заявку из-за связанных данных (IntegrityError). Связанные записи не были полностью удалены.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash('Ошибка при удалении заявки.', 'danger')

    return redirect(url_for('ticket_list'))


@app.route('/admin/fix-updated-at')
@login_required
def fix_updated_at():
    if current_user.role != 'admin':
        return redirect(url_for('admin'))
    
    tickets = SupportTicket.query.filter(
        SupportTicket.status == 'Завершена',
        SupportTicket.updated_at.is_(None)
    ).all()
    
    fixed = 0
    for ticket in tickets:
        # Для существующих заявок используем created_at + 1 час (примерно)
        ticket.updated_at = ticket.created_at + timedelta(hours=1)
        fixed += 1
    
    db.session.commit()
    return f"Обновлено {fixed} заявок. Теперь updated_at = created_at + 1 час."

@app.route('/admin/fix-sla')
@login_required
def fix_sla():
    if current_user.role != 'admin':
        return redirect(url_for('admin'))
    
    tickets = SupportTicket.query.filter(SupportTicket.sla_deadline.is_(None)).all()
    fixed = 0
    
    for ticket in tickets:
        recalc_ticket_sla(ticket)
        fixed += 1
    
    db.session.commit()
    return f"Исправлено {fixed} заявок"

@app.route('/user/profile', methods=['GET', 'POST'])
@login_required
def user_profile():
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'update_profile':
            # Обновление ФИО (без телефона/уведомлений/реквизитов)
            old_email = current_user.email

            name = request.form.get('name', '')
            last_name = request.form.get('last_name', '')
            patronymic = request.form.get('patronymic', '')
            email = request.form.get('email', '').strip().lower() or old_email

            if getattr(current_user, 'role', None) == 'client':
                # Для клиента почта обязательна, ФИО — опционально (можно заполнить позже в профиле).
                if not email:
                    flash("Email обязателен", "error")
                    return redirect(url_for('user_profile'))

                # Если ФИО заполняется — валидируем только заполненные части (отчество можно не указывать).
                for label, raw in (("Фамилия", last_name), ("Имя", name), ("Отчество", patronymic)):
                    if _norm_text(raw):
                        ok, v = validate_person_part(raw)
                        if not ok:
                            flash(f"{label}: {v}", "error")
                            return redirect(url_for('user_profile'))

                last_name, name, patronymic = _norm_text(last_name), _norm_text(name), _norm_text(patronymic)
            else:
                # для сотрудников валидируем только если заполнено
                for label, raw in (("Фамилия", last_name), ("Имя", name), ("Отчество", patronymic)):
                    if _norm_text(raw):
                        ok, v = validate_person_part(raw)
                        if not ok:
                            flash(f"{label}: {v}", "error")
                            return redirect(url_for('user_profile'))
                last_name, name, patronymic = _norm_text(last_name), _norm_text(name), _norm_text(patronymic)

            # Проверка уникальности email
            if email != old_email:
                existing = User.query.filter_by(email=email).first()
                if existing:
                    flash("Этот email уже используется", "error")
                    return redirect(url_for('user_profile'))

            current_user.name = name or None
            current_user.last_name = last_name or None
            current_user.patronymic = patronymic or None
            current_user.email = email

            try:
                db.session.commit()
                if email != old_email and hasattr(current_user, 'email_verified'):
                    current_user.email_verified = False
                    db.session.commit()
                    flash("На новый email отправлено письмо для подтверждения.", "info")
                else:
                    flash("Профиль обновлён", "success")
            except Exception:
                db.session.rollback()
                flash("Ошибка при сохранении профиля", "error")
            return redirect(url_for('user_profile'))

        if action == 'update_phone':
            phone_raw = request.form.get('phone', '')
            ok, phone = normalize_phone(phone_raw)
            if not ok:
                flash(phone, "error")
                return redirect(url_for('user_profile'))
            current_user.phone = phone or None
            try:
                db.session.commit()
                flash("Телефон обновлён", "success")
            except Exception:
                db.session.rollback()
                flash("Ошибка при сохранении телефона", "error")
            return redirect(url_for('user_profile'))

        if action == 'update_notifications':
            # чекбоксы приходят только если отмечены
            notify_inapp_enabled = request.form.get('notify_inapp_enabled') == 'on'
            notify_event_assigned = request.form.get('notify_event_assigned') == 'on'
            notify_event_customer_reply = request.form.get('notify_event_customer_reply') == 'on'
            notify_event_status = request.form.get('notify_event_status') == 'on'

            current_user.notify_inapp_enabled = bool(notify_inapp_enabled)
            current_user.notify_event_assigned = bool(notify_event_assigned)
            current_user.notify_event_customer_reply = bool(notify_event_customer_reply)
            current_user.notify_event_status = bool(notify_event_status)

            try:
                db.session.commit()
                flash("Настройки уведомлений сохранены", "success")
            except Exception:
                db.session.rollback()
                flash("Ошибка при сохранении настроек уведомлений", "error")
            return redirect(url_for('user_profile'))

        if action == 'update_org':
            if not getattr(current_user, 'role', None) == 'client':
                abort(403)

            organization_raw = request.form.get('organization', '')
            position_raw = request.form.get('position', '')
            inn_raw = request.form.get('inn', '')
            address_raw = request.form.get('address', '')

            ok, org_v = validate_org(organization_raw)
            if not ok:
                flash(f"Организация: {org_v}", "error")
                return redirect(url_for('user_profile'))

            ok, inn_v = validate_inn_ru(inn_raw, required=True)
            if not ok:
                flash(f"ИНН: {inn_v}", "error")
                return redirect(url_for('user_profile'))

            ok, addr_v = validate_address(address_raw, required=True)
            if not ok:
                flash(f"Адрес: {addr_v}", "error")
                return redirect(url_for('user_profile'))

            # должность — свободное поле, но без ссылок/HTML
            pos_v = _norm_text(position_raw)
            if _contains_url_like(pos_v) or _contains_html_like(pos_v):
                flash("Должность: ссылки и HTML запрещены", "error")
                return redirect(url_for('user_profile'))

            current_user.organization = org_v or None
            current_user.position = pos_v or None
            current_user.inn = inn_v or None
            current_user.address = addr_v or None

            try:
                db.session.commit()
                flash("Реквизиты обновлены", "success")
            except Exception:
                db.session.rollback()
                flash("Ошибка при сохранении реквизитов", "error")
            return redirect(url_for('user_profile'))

        if action == 'apply_suggested_org':
            if not getattr(current_user, 'role', None) == 'client':
                abort(403)
            s_org = getattr(current_user, 'suggested_organization', None) or ''
            s_inn = getattr(current_user, 'suggested_inn', None) or ''
            s_addr = getattr(current_user, 'suggested_address', None) or ''
            ok, org_v = validate_org(s_org)
            ok_inn, inn_v = validate_inn_ru(s_inn, required=True)
            ok_addr, addr_v = validate_address(s_addr, required=True)
            if (not ok) or (not _norm_text(org_v)) or (not ok_inn) or (not ok_addr):
                flash('Предложенные реквизиты неполные или некорректные. Заполните реквизиты вручную.', 'error')
                return redirect(url_for('user_profile'))
            current_user.organization = org_v or None
            current_user.inn = inn_v or None
            current_user.address = addr_v or None
            current_user.suggested_organization = None
            current_user.suggested_inn = None
            current_user.suggested_address = None
            try:
                db.session.commit()
                flash('Реквизиты обновлены (принято из CRM).', 'success')
            except Exception:
                db.session.rollback()
                flash('Ошибка при сохранении реквизитов.', 'error')
            return redirect(url_for('user_profile'))

        if action == 'dismiss_suggested_org':
            if not getattr(current_user, 'role', None) == 'client':
                abort(403)
            current_user.suggested_organization = None
            current_user.suggested_inn = None
            current_user.suggested_address = None
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
            flash('Предложенные реквизиты скрыты.', 'info')
            return redirect(url_for('user_profile'))

        elif action == 'change_password':
            current_password = request.form.get('current_password', '').strip()
            new_password = request.form.get('new_password', '').strip()
            confirm_password = request.form.get('confirm_password', '').strip()
            
            if not current_password or not new_password or not confirm_password:
                flash("Все поля обязательны", "error")
                return redirect(url_for('user_profile'))
            
            if not check_password_hash(current_user.password, current_password):
                flash("Текущий пароль неверен", "error")
                return redirect(url_for('user_profile'))
            
            if new_password != confirm_password:
                flash("Новые пароли не совпадают", "error")
                return redirect(url_for('user_profile'))
            
            min_length = 8 if isinstance(current_user, User) else 6
            if len(new_password) < min_length:
                flash(f"Пароль должен быть не менее {min_length} символов", "error")
                return redirect(url_for('user_profile'))
            
            current_user.password = generate_password_hash(new_password, method='pbkdf2:sha256')
            try:
                db.session.commit()
                
                if getattr(current_user, 'role', None) == 'client':
                    logout_user()
                    flash("Пароль изменён. Пожалуйста, войдите снова", "success")
                    return redirect(url_for('login'))
                else:
                    flash("Пароль успешно изменён", "success")
                    return redirect(url_for('user_profile'))
            except Exception as e:
                db.session.rollback()
                flash("Ошибка при смене пароля", "error")
                return redirect(url_for('user_profile'))

    # === GET ===
    tickets_count = 0
    resolved_count = 0
    operator_stats = {}

    if getattr(current_user, 'role', None) == 'client':
        tickets_count = SupportTicket.query.filter_by(email=current_user.email).count()
        resolved_count = SupportTicket.query.filter_by(email=current_user.email, is_resolved=True).count()
    elif isinstance(current_user, User):
        operator_stats = {
            'role_display': dict(USER_ROLES).get(current_user.role, current_user.role),
            'department': current_user.department.name if current_user.department else '—',
            'username': current_user.username
        }

    return render_template(
        'user_profile.html',
        tickets_count=tickets_count,
        resolved_count=resolved_count,
        operator_stats=operator_stats
    )
## /user/instructions удалён — страница не используется

## /changelog удалён — страница не используется




# === Compatibility routes / aliases ===
try:
    app.add_url_rule('/tickets', endpoint='user_tickets', view_func=ticket_list)
except Exception:
    pass

try:
    app.add_url_rule('/admin/support-list', endpoint='admin_support_list', view_func=ticket_list)
except Exception:
    pass

try:
    app.add_url_rule('/admin/user-reset-password/<int:user_id>', endpoint='admin_user_reset_password', view_func=admin_reset_password)
except Exception:
    pass

try:
    app.add_url_rule('/knowledge/add-faq', endpoint='add_faq', view_func=add_faq_view, methods=['GET', 'POST'])
except Exception:
    pass

try:
    app.add_url_rule('/knowledge/edit-faq/<int:id>', endpoint='edit_faq', view_func=edit_faq_view, methods=['GET', 'POST'])
except Exception:
    pass

@app.route('/admin/charts')
@login_required
def admin_charts():
    if getattr(current_user, 'role', None) != 'admin':
        flash('Доступ запрещён', 'error')
        return redirect(url_for('ticket_list'))
    return render_template('admin_charts.html')


@app.route('/admin/audit')
@login_required
def admin_audit():
    if getattr(current_user, 'role', None) != 'admin':
        flash('Доступ запрещён', 'error')
        return redirect(url_for('ticket_list'))
    return render_template('admin_audit.html', logs=[], action=request.args.get('action', ''), actor=request.args.get('actor', ''), target_type=request.args.get('target_type', ''))


@app.route('/admin/endusers/<int:user_id>/edit', methods=['POST'])
@login_required
def admin_edit_enduser(user_id):
    if getattr(current_user, 'role', None) != 'admin':
        flash('Доступ запрещён', 'error')
        return redirect(url_for('admin_users'))
    u = User.query.get_or_404(user_id)
    if u.role != 'client':
        flash('Можно редактировать только конечного пользователя', 'error')
        return redirect(url_for('admin_users'))

    email = (request.form.get('email') or '').strip().lower()
    phone = (request.form.get('phone') or '').strip() or None
    # Допускаем профиль без email, если есть телефон
    if not email and not phone:
        flash('Нужно указать email или телефон', 'error')
        return redirect(url_for('admin_users'))
    if email and User.query.filter(db.func.lower(User.email) == email.lower(), User.id != u.id).first():
        flash('Email уже используется', 'error')
        return redirect(url_for('admin_users'))

    try:
        u.name = (request.form.get('name') or '').strip() or None
        u.last_name = (request.form.get('last_name') or '').strip() or None
        u.patronymic = (request.form.get('patronymic') or '').strip() or None
        u.email = email or None
        u.username = email or phone or (u.username or f'user_{u.id}')
        u.phone = phone
        u.organization = (request.form.get('organization') or '').strip() or None
        u.inn = (request.form.get('inn') or '').strip() or None
        u.address = (request.form.get('address') or '').strip() or None
        u.is_active = bool(request.form.get('is_active') == 'on')
        u.email_verified = bool(request.form.get('email_verified') == 'on')
        db.session.commit()
        flash('Пользователь обновлён', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка сохранения: {e}', 'error')
    return redirect(url_for('admin_users'))


@app.route('/tickets/<int:ticket_id>/accept', methods=['POST'])
@login_required
def accept_ticket(ticket_id):
    ticket = SupportTicket.query.get_or_404(ticket_id)
    if getattr(current_user, 'role', None) not in ('admin', 'operator') and not is_tp_operator(current_user):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('ticket_detail', ticket_id=ticket.id))
    try:
        if getattr(ticket, 'assigned_to_id', None) in (None, 0):
            ticket.assigned_to_id = current_user.id
        old_status = ticket.status
        if ticket.status in ('Новая', 'Принята', None, ''):
            ticket.status = 'В работе'
        try:
            log_ticket_change(ticket.id, current_user.id, 'status', old_status, ticket.status)
        except Exception:
            pass
        db.session.commit()
        flash('Заявка принята в работу', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Не удалось взять заявку в работу: {e}', 'error')
    return redirect(url_for('ticket_detail', ticket_id=ticket.id))


# === ЗАПУСК ===
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # Создание админа (если нужно)
        if not User.query.filter_by(username='admin').first():
            # ... ваш код создания админа ...
            db.session.commit()
        
        # ЗАПУСК ПЛАНИРОВЩИКА
        start_scheduler()
    
    (
    socketio.run(app, host='0.0.0.0', port=5000, debug=True)
    if socketio is not None else app.run(host='0.0.0.0', port=5000, threaded=True, debug=True)
)
