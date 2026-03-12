from __future__ import annotations

import os
from datetime import datetime
from io import BytesIO

from flask import current_app, redirect, render_template, request, send_file, url_for
from flask_login import current_user
try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None
from sqlalchemy.orm import load_only, joinedload

from helpdesk_app.models.tickets import SupportTicket, TicketHistory
from helpdesk_app.services.sla_service import SLAService


class ReportService:
    """Bridge-сервис статистики/отчётов c выравниванием на новый SLA-движок.

    Базовый аналитический payload по-прежнему собирается legacy-слоем для совместимости,
    но SLA-метрики и FRT/MTTR поверх него пересчитываются через новый SLAService.
    """

    @staticmethod
    def _legacy():
        from helpdesk_app.runtime import get_runtime
        return get_runtime()

    @classmethod
    def ensure_admin_or_redirect(cls):
        if getattr(current_user, 'role', None) == 'client' or getattr(current_user, 'role', None) != 'admin':
            return redirect(url_for('admin'))
        return None

    @classmethod
    def current_filters(cls):
        return {
            'period': request.args.get('period', '30'),
            'department_id': request.args.get('department_id', type=int),
            'operator_id': request.args.get('operator_id', type=int),
            'date_from': request.args.get('date_from', ''),
            'date_to': request.args.get('date_to', ''),
            'view': request.args.get('view', 'created'),
        }

    @classmethod
    def _build_sla_ticket_query(cls, filters: dict):
        legacy = cls._legacy()
        query = SupportTicket.query.options(
            load_only(
                SupportTicket.id,
                SupportTicket.created_at,
                SupportTicket.closed_at,
                SupportTicket.auto_closed_at,
                SupportTicket.marked_as_completed_at,
                SupportTicket.status,
                SupportTicket.is_resolved,
                SupportTicket.sla_deadline,
                SupportTicket.department_id,
                SupportTicket.assigned_to_id,
            )
        )

        if filters.get('department_id'):
            query = query.filter(SupportTicket.department_id == filters['department_id'])

        if filters.get('operator_id'):
            query = query.filter(SupportTicket.assigned_to_id == filters['operator_id'])

        date_from = filters.get('date_from')
        date_to = filters.get('date_to')
        period = str(filters.get('period') or '30').strip()
        report_view = (filters.get('view') or 'created').strip()

        dt_from = None
        dt_to = None
        if date_from:
            try:
                dt_from = legacy.datetime.strptime(date_from, '%Y-%m-%d')
            except ValueError:
                dt_from = None
        if date_to:
            try:
                dt_to = legacy.datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            except ValueError:
                dt_to = None

        if not dt_from and not dt_to:
            try:
                days = max(1, int(period))
            except Exception:
                days = 30
            dt_from = legacy.utcnow() - legacy.timedelta(days=days)
            dt_to = legacy.utcnow()

        if report_view == 'resolved':
            resolved_expr = legacy.func.coalesce(
                SupportTicket.closed_at,
                SupportTicket.auto_closed_at,
                SupportTicket.marked_as_completed_at,
            )
            if dt_from is not None:
                query = query.filter(resolved_expr >= dt_from)
            if dt_to is not None:
                query = query.filter(resolved_expr <= dt_to)
            query = query.filter(resolved_expr.isnot(None))
        else:
            if dt_from is not None:
                query = query.filter(SupportTicket.created_at >= dt_from)
            if dt_to is not None:
                query = query.filter(SupportTicket.created_at <= dt_to)

        return query

    @classmethod
    def _build_ticket_scope(cls, filters: dict):
        legacy = cls._legacy()
        query = SupportTicket.query.options(
            joinedload(SupportTicket.department_rel),
            joinedload(SupportTicket.assigned_to_rel),
            joinedload(SupportTicket.locked_by_rel),
            joinedload(SupportTicket.category_rel),
        )

        if filters.get('department_id'):
            query = query.filter(SupportTicket.department_id == filters['department_id'])

        if filters.get('operator_id'):
            query = query.filter(SupportTicket.assigned_to_id == filters['operator_id'])

        date_from = filters.get('date_from')
        date_to = filters.get('date_to')
        period = str(filters.get('period') or '30').strip()
        report_view = (filters.get('view') or 'created').strip()

        dt_from = None
        dt_to = None
        if date_from:
            try:
                dt_from = legacy.datetime.strptime(date_from, '%Y-%m-%d')
            except ValueError:
                dt_from = None
        if date_to:
            try:
                dt_to = legacy.datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            except ValueError:
                dt_to = None

        if not dt_from and not dt_to:
            try:
                days = max(1, int(period))
            except Exception:
                days = 30
            dt_from = legacy.utcnow() - legacy.timedelta(days=days)
            dt_to = legacy.utcnow()

        if report_view == 'resolved':
            resolved_expr = legacy.func.coalesce(
                SupportTicket.closed_at,
                SupportTicket.auto_closed_at,
                SupportTicket.marked_as_completed_at,
            )
            if dt_from is not None:
                query = query.filter(resolved_expr >= dt_from)
            if dt_to is not None:
                query = query.filter(resolved_expr <= dt_to)
            query = query.filter(resolved_expr.isnot(None))
        else:
            if dt_from is not None:
                query = query.filter(SupportTicket.created_at >= dt_from)
            if dt_to is not None:
                query = query.filter(SupportTicket.created_at <= dt_to)

        return query

    @staticmethod
    def _avg_minutes(values):
        cleaned = [v for v in values if isinstance(v, (int, float))]
        if not cleaned:
            return 0
        return int(round(sum(cleaned) / len(cleaned)))

    @staticmethod
    def _format_minutes(total_minutes: int) -> str:
        if not total_minutes:
            return '0м'
        hours, minutes = divmod(int(total_minutes), 60)
        days, hours = divmod(hours, 24)
        parts = []
        if days:
            parts.append(f'{days}д')
        if hours:
            parts.append(f'{hours}ч')
        parts.append(f'{minutes}м')
        return ' '.join(parts)

    @classmethod
    def _apply_new_sla_metrics(cls, data: dict, filters: dict):
        tickets = cls._build_sla_ticket_query(filters).all()
        views = SLAService.build_ticket_views(tickets)

        compliance = {
            'ok': 0,
            'paused': 0,
            'first_response_breach': 0,
            'resolve_breach': 0,
        }
        fr_minutes = []
        mttr_minutes = []

        for ticket in tickets:
            view = views.get(ticket.id) or {}
            fr = view.get('first_response') or {}
            rs = view.get('resolve') or {}

            if view.get('summary_status') == 'paused':
                compliance['paused'] += 1
            elif fr.get('status') in {'overdue', 'breached'}:
                compliance['first_response_breach'] += 1
            elif rs.get('status') in {'overdue', 'breached'}:
                compliance['resolve_breach'] += 1
            else:
                compliance['ok'] += 1

            fr_completed = fr.get('completed_at')
            if ticket.created_at and fr_completed:
                fr_minutes.append(max(0, int((fr_completed - ticket.created_at).total_seconds() // 60)))

            rs_completed = rs.get('completed_at')
            if ticket.created_at and rs_completed:
                mttr_minutes.append(max(0, int((rs_completed - ticket.created_at).total_seconds() // 60)))

        total_closed = compliance['ok'] + compliance['first_response_breach'] + compliance['resolve_breach']
        percent = int(round((compliance['ok'] / total_closed) * 100)) if total_closed else 100

        data['sla_alignment_meta'] = {
            'source': 'new_sla_service',
            'tickets_evaluated': len(tickets),
        }
        data['sla_compliance_breakdown'] = compliance
        data['sla_compliance_percent'] = percent

        metrics = dict(data.get('sla_metrics') or {})
        metrics['on_time'] = compliance['ok']
        metrics['overdue_completed'] = compliance['first_response_breach'] + compliance['resolve_breach']
        metrics['frt_avg'] = cls._format_minutes(cls._avg_minutes(fr_minutes))
        metrics['avg_time_to_resolve'] = cls._format_minutes(cls._avg_minutes(mttr_minutes))
        data['sla_metrics'] = metrics
        return data

    @classmethod
    def _apply_step9_analytics(cls, data: dict, filters: dict):
        tickets = cls._build_ticket_scope(filters).all()
        ticket_ids = [t.id for t in tickets]

        def resolved_at(ticket):
            return ticket.closed_at or ticket.auto_closed_at or ticket.marked_as_completed_at

        closed_tickets = [t for t in tickets if resolved_at(t)]

        top_clients_map = {}
        for t in tickets:
            key = (t.inn or '').strip() or (t.organization or '').strip() or 'Без организации'
            row = top_clients_map.setdefault(key, {
                'name': (t.organization or '').strip() or 'Без названия',
                'inn': (t.inn or '').strip(),
                'tickets': 0,
                'open': 0,
                'overdue': 0,
            })
            row['tickets'] += 1
            if not resolved_at(t):
                row['open'] += 1
            if getattr(t, 'is_overdue', False) and not resolved_at(t):
                row['overdue'] += 1
        top_clients = sorted(top_clients_map.values(), key=lambda x: (x['tickets'], x['open']), reverse=True)[:7]

        close_reason_map = {}
        for t in closed_tickets:
            reason = (t.close_reason or '').strip() or 'Обычное закрытие'
            close_reason_map[reason] = close_reason_map.get(reason, 0) + 1
        top_close_reasons = [
            {'reason': k, 'count': v}
            for k, v in sorted(close_reason_map.items(), key=lambda item: item[1], reverse=True)[:7]
        ]

        events = dict(data.get('events') or {})
        reopened = int(events.get('reopened', 0) or 0)
        reopen_rate = int(round((reopened / len(closed_tickets)) * 100)) if closed_tickets else 0

        top_categories = [
            {'name': k, 'count': v}
            for k, v in sorted((data.get('category_counts') or {}).items(), key=lambda item: item[1], reverse=True)[:7]
        ]

        queue_hotspots = sorted(data.get('operator_rows') or [], key=lambda row: (row.get('open_end', 0), row.get('overdue_open', 0)), reverse=True)[:7]

        data['step9'] = {
            'reopen_rate': reopen_rate,
            'reopened_count': reopened,
            'closed_count': len(closed_tickets),
            'top_clients': top_clients,
            'top_close_reasons': top_close_reasons,
            'top_categories': top_categories,
            'queue_hotspots': queue_hotspots,
        }
        return data

    @classmethod
    def build_data(cls, *, for_export: bool = False):
        legacy = cls._legacy()
        f = cls.current_filters()
        data = legacy._build_analytics_data(
            f['period'],
            f['department_id'],
            f['operator_id'],
            f['date_from'],
            f['date_to'],
            for_export=for_export,
            report_view=f['view'],
        )
        data = cls._apply_new_sla_metrics(data, f)
        return cls._apply_step9_analytics(data, f)

    @classmethod
    def render_statistics(cls):
        denied = cls.ensure_admin_or_redirect()
        if denied is not None:
            return denied
        data = cls.build_data(for_export=False)
        return render_template('admin_statistics.html', **data)

    @classmethod
    def render_reports(cls):
        denied = cls.ensure_admin_or_redirect()
        if denied is not None:
            return denied
        data = cls.build_data(for_export=False)
        return render_template('admin_reports.html', **data)

    @classmethod
    def export_xlsx(cls):
        denied = cls.ensure_admin_or_redirect()
        if denied is not None:
            return denied
        data = cls.build_data(for_export=True)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Сводка'
        ws.append(['Параметр', 'Значение'])
        ws.append(['Период', data.get('period_label', '')])
        ws.append(['Создано', data['sla_metrics']['created_count']])
        ws.append(['Закрыто', data['sla_metrics']['resolved_count']])
        ws.append(['Net flow (создано - закрыто)', data['sla_metrics']['net_flow']])
        ws.append(['Очередь на начало', data['sla_metrics']['backlog_start']])
        ws.append(['Очередь на конец', data['sla_metrics']['backlog_end']])
        ws.append(['Открыто сейчас', data['in_progress']])
        ws.append(['Просрочено сейчас (активные)', data['overdue_active']])
        ws.append(['SLA соблюдено (закрытые в период)', f"{data['sla_compliance_percent']}%"])
        ws.append(['Среднее время решения (MTTR)', data['sla_metrics']['avg_time_to_resolve']])
        ws.append(['Среднее время до первого ответа (FRT)', data['sla_metrics']['frt_avg']])
        ws.append(['Среднее время до принятия (Accept)', data['sla_metrics'].get('avg_time_to_accept', '')])
        ws.append(['Среднее время в работе (In work)', data['sla_metrics'].get('avg_time_in_work', '')])
        ws.append(['Переоткрыто (reopen)', data.get('events', {}).get('reopened', 0)])
        ws.append(['Доработка от клиента', data.get('events', {}).get('rework', 0)])
        ws.append(['Переводы между отделами', data.get('events', {}).get('department_moves', 0)])
        ws.append(['Удовлетворённость (👍/всего)', f"{data['helpful_pos']}/{data['helpful_total']} ({data['satisfaction']}%)"])

        ws2 = wb.create_sheet('Операторы')
        ws2.append(['Оператор', 'Создано', 'Закрыто', 'Открыто (конец периода)', 'Просрочено (сейчас)', 'Среднее время решения'])
        for row in data['operator_rows']:
            ws2.append([
                row['username'],
                row.get('created', 0),
                row.get('resolved', 0),
                row.get('open_end', 0),
                row.get('overdue_open', 0),
                row.get('avg_mttr', ''),
            ])

        ws3 = wb.create_sheet('Заявки')
        ws3.append(['ID', 'Создана', 'Закрыта', 'Статус', 'Приоритет', 'Отдел', 'Оператор', 'Тема', 'SLA дедлайн', 'Просрочена', 'FRT', 'Accept', 'MTTR', 'In work', 'Оценка/👍'])
        for ticket in data.get('export_rows', []):
            ws3.append([
                ticket['id'],
                ticket['created_at'],
                ticket['resolved_at'],
                ticket['status'],
                ticket['priority'],
                ticket['department'],
                ticket['operator'],
                ticket['subject'],
                ticket['sla_deadline'],
                ticket['sla_breached'],
                ticket.get('frt', ''),
                ticket.get('accept', ''),
                ticket.get('mttr', ''),
                ticket.get('in_work', ''),
                ticket['helpful'],
            ])

        filename = f"reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        instance_dir = os.path.join(current_app.root_path, 'instance')
        os.makedirs(instance_dir, exist_ok=True)
        tmp_path = os.path.join(instance_dir, filename)
        try:
            wb.save(tmp_path)
            return send_file(tmp_path, as_attachment=True, download_name=filename)
        except Exception:
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(
                bio,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
